#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from __future__ import absolute_import

from sican.celery import app
import openpyxl
from matrices.models import CargaMasiva, Beneficiario
from informes.functions import construir_reporte
from datetime import datetime
from usuarios.models import User
from django.core.files import File
import pytz
from productos.models import Diplomado
from region.models import Region
from formadores.models import Formador
from formadores.models import Grupos
from matrices.models import Beneficiario, Area, Grado
from radicados.models import Radicado
from sican.settings import base as settings
import PIL
from PIL import ImageFont
from PIL import Image
from PIL import ImageDraw
import StringIO
from django.core.files import File
from django.utils import timezone

@app.task
def carga_masiva_matrices(id,email):
    carga = CargaMasiva.objects.get(id = id)
    wb = openpyxl.load_workbook(filename = carga.archivo.file.name,read_only=True)
    sheets = wb.get_sheet_names()

    if 'InnovaTIC' in sheets and 'TecnoTIC' in sheets and 'DirecTIC' in sheets:
        carga.estado = 'Iniciando carga masiva ...'

        titulos = ['DIPLOMADO',
                   'RESULTADO',
                   'REGION',
                   'DEPARTAMENTO',
                   'SECRETARIA DE EDUCACION',
                   'RADICADO',
                   'CODIGO DANE INSTITUCION EDUCATIVA',
                   'NOMBRE INSTITUCION EDUCATIVA',
                   'CODIGO DANE SEDE EDUCATIVA',
                   'NOMBRE DE LA SEDE EDUCATIVA',
                   'MUNICIPIO',
                   'ZONA (URBANA/RURAL)',
                   'CODIGO DEL GRUPO',
                   'NOMBRE DEL FORMADOR',
                   'NUMERO DE CEDULA DEL FORMADOR',
                   'APELLIDOS DEL DOCENTE',
                   'NOMBRES DEL DOCENTE',
                   'NUMERO DE CEDULA DEL DOCENTE',
                   'CORREO ELECTRONICO',
                   'TELEFONO FIJO',
                   'TELEFONO CELULAR',
                   'AREA CURRICULAR',
                   'GRADO',
                   'TIPO DE BENEFICIARIO',
                   'GENERO',
                   'ESTADO']

        formatos = ['General',
                    'General',
                   '0',
                   'General',
                   'General',
                   '0',
                   '0',
                   'General',
                   '0',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   '0',
                   'General',
                   'General',
                   '0',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General']

        ancho_columnas =  [30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           ]

        contenidos = []

        for name in ['InnovaTIC','TecnoTIC','DirecTIC']:
            ws = wb.get_sheet_by_name(name)

            for fila in ws.iter_rows(row_offset=5):

                resultado = ''

                if Diplomado.objects.filter(nombre__icontains = name).count() == 1:
                    diplomado = Diplomado.objects.get(nombre__icontains = name)

                    try:
                        region = Region.objects.get(numero = fila[0].value if fila[0].value != None else 0)
                    except:
                        resultado = 'No existe el código de región'

                    else:

                        if Formador.objects.filter(cedula = fila[12].value if fila[12].value != None else 0).count() == 1:
                            formador = Formador.objects.get(cedula = fila[12].value if fila[12].value != None else 0)

                            try:
                                grupo_list = fila[10].value.split('-') if fila[10].value != None else ['']
                            except:
                                resultado = 'No se puede identificar el grupo'
                            else:
                                if len(grupo_list) > 1:
                                    try:
                                        grupo_numero = int(grupo_list[-1])
                                    except:
                                        resultado = 'No se puede identificar el grupo'
                                    else:
                                        ruta = ''

                                        for t in grupo_list[:-1]:
                                            ruta += str(t) + '-'

                                        ruta = ruta[:-1]

                                        grupo, grupo_creado = Grupos.objects.get_or_create(formador=formador,nombre = grupo_numero)

                                        if grupo_creado:
                                            resultado += 'Grupo creado, '

                                        if fila[13].value != None:

                                            if fila[14].value != None:

                                                try:
                                                    radicado_numero = int(fila[3].value)
                                                except:
                                                    radicado_numero = -1

                                                if Radicado.objects.filter(numero = radicado_numero).count() == 1:
                                                    radicado = Radicado.objects.get(numero = radicado_numero)
                                                else:
                                                    radicado = None


                                                if Area.objects.filter(numero__icontains = str(fila[19].value).split('.')[0] if fila[19].value != None else '').count() == 1:
                                                    area = Area.objects.get(numero__icontains = str(fila[19].value).split('.')[0])
                                                else:
                                                    area = None

                                                if Grado.objects.filter(numero__icontains = str(fila[20].value).split('.')[0] if fila[20].value != None else '').count() == 1:
                                                    grado = Grado.objects.get(numero__icontains = str(fila[20].value).split('.')[0])
                                                else:
                                                    grado = None

                                                radicado_text = str(radicado.numero) if radicado != None else ''

                                                try:
                                                    cedula = long(fila[15].value)
                                                except:
                                                    resultado = 'Error en el numero de cedula'
                                                else:

                                                    if Beneficiario.objects.filter(cedula = fila[15].value if fila[15].value != None else '').count() == 1:
                                                        beneficiario = Beneficiario.objects.get(cedula = fila[15].value)
                                                        beneficiario.region = region
                                                        beneficiario.radicado = radicado

                                                        beneficiario.departamento_text = fila[1].value
                                                        beneficiario.secretaria_text = fila[2].value
                                                        beneficiario.dane_ie_text = fila[4].value
                                                        beneficiario.ie_text = fila[5].value
                                                        beneficiario.dane_sede_text = fila[6].value
                                                        beneficiario.sede_text = fila[7].value
                                                        beneficiario.municipio_text = fila[8].value
                                                        beneficiario.zona_text = fila[9].value

                                                        beneficiario.radicado_text = radicado_text
                                                        beneficiario.formador = formador
                                                        beneficiario.ruta = ruta
                                                        beneficiario.grupo = grupo
                                                        beneficiario.apellidos =fila[13].value
                                                        beneficiario.nombres = fila[14].value
                                                        beneficiario.cedula=fila[15].value
                                                        beneficiario.correo=fila[16].value
                                                        beneficiario.telefono_fijo=fila[17].value
                                                        beneficiario.telefono_celular=fila[18].value
                                                        beneficiario.area = area
                                                        beneficiario.grado = grado
                                                        beneficiario.genero=fila[22].value
                                                        beneficiario.estado=fila[23].value
                                                        beneficiario.save()

                                                        resultado += 'Docente actualizado'

                                                    else:
                                                        Beneficiario.objects.create(diplomado = diplomado,region=region,radicado=radicado,radicado_text=radicado_text,
                                                                                    departamento_text=fila[1].value,secretaria_text=fila[2].value,
                                                                                    dane_ie_text=fila[4].value,ie_text=fila[5].value,dane_sede_text=fila[6].value,
                                                                                    sede_text=fila[7].value,municipio_text=fila[8].value,zona_text=fila[9].value,
                                                                                    formador=formador,grupo=grupo,apellidos=fila[13].value,ruta=ruta,
                                                                                    nombres=fila[14].value,cedula=fila[15].value,correo=fila[16].value,
                                                                                    telefono_fijo=fila[17].value,telefono_celular=fila[18].value,
                                                                                    area=area,grado=grado,genero=fila[22].value,estado=fila[23].value)
                                                        resultado += 'Docente creado'

                                            else:
                                                resultado += 'No hay nombres del docente'

                                        else:
                                            resultado += 'No hay apellidos del docente'

                                else:
                                    resultado = 'No se puede identificar el grupo'
                        else:
                            resultado = 'No hay ningun formador con el numero de cedula'


                else:
                    resultado = 'No existe el diplomado'




                contenidos.append([
                    name,
                    resultado,
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    fila[12].value,
                    fila[13].value,
                    fila[14].value,
                    fila[15].value,
                    fila[16].value,
                    fila[17].value,
                    fila[18].value,
                    fila[19].value,
                    fila[20].value,
                    fila[21].value,
                    fila[22].value,
                    fila[23].value
                ])

        usuario = User.objects.get(email=email)
        nombre = "Resultado carga masiva matrices"
        proceso = "FOR-MAS01"
        fecha = pytz.utc.localize(datetime.now())
        output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
        filename = unicode(fecha) + '.xlsx'
        carga.resultado.save(filename,File(output))
        carga.estado = 'Proceso concluido'

    elif 'Matriz revisión documental' in sheets:
        carga.estado = 'Iniciando carga masiva ...'

        titulos = ['DIPLOMADO',
                   'RESULTADO',
                   'REGION',
                   'DEPARTAMENTO',
                   'SECRETARIA DE EDUCACION',
                   'RADICADO',
                   'CODIGO DANE INSTITUCION EDUCATIVA',
                   'NOMBRE INSTITUCION EDUCATIVA',
                   'CODIGO DANE SEDE EDUCATIVA',
                   'NOMBRE DE LA SEDE EDUCATIVA',
                   'MUNICIPIO',
                   'ZONA (URBANA/RURAL)',
                   'CODIGO DEL GRUPO',
                   'NOMBRE DEL FORMADOR',
                   'NUMERO DE CEDULA DEL FORMADOR',
                   'APELLIDOS DEL DOCENTE',
                   'NOMBRES DEL DOCENTE',
                   'NUMERO DE CEDULA DEL DOCENTE',
                   'CORREO ELECTRONICO',
                   'TELEFONO FIJO',
                   'TELEFONO CELULAR',
                   'AREA CURRICULAR',
                   'GRADO',
                   'TIPO DE BENEFICIARIO',
                   'GENERO',
                   'ESTADO']

        formatos = ['General',
                    'General',
                   '0',
                   'General',
                   'General',
                   '0',
                   '0',
                   'General',
                   '0',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   '0',
                   'General',
                   'General',
                   '0',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General']

        ancho_columnas =  [30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           ]

        contenidos = []

        ws = wb.get_sheet_by_name('Matriz revisión documental')

        for fila in ws.iter_rows(row_offset=9):

            resultado = ''

            diplomado = Diplomado.objects.get(id = 4)

            try:
                region = Region.objects.get(numero = fila[0].value if fila[0].value != None else 0)
            except:
                resultado = 'No existe el código de región'

            else:

                if Formador.objects.filter(cedula = fila[12].value if fila[12].value != None else 0).count() == 1:
                    formador = Formador.objects.get(cedula = fila[12].value if fila[12].value != None else 0)

                    try:
                        grupo_list = fila[10].value.split('-') if fila[10].value != None else ['']
                    except:
                        resultado = 'No se puede identificar el grupo'
                    else:

                        if len(grupo_list) > 1:
                            try:
                                grupo_numero = int(grupo_list[-1])
                            except:
                                resultado = 'No se puede identificar el grupo'
                            else:
                                try:
                                    grupo, grupo_creado = Grupos.objects.get_or_create(formador=formador,nombre = grupo_numero)
                                except:
                                    resultado = 'Hay mas de un grupo con el codigo'

                                else:

                                    if grupo_creado:
                                        resultado += 'Grupo creado, '

                                    if fila[13].value != None:

                                        if fila[14].value != None:

                                            try:
                                                radicado_numero = int(fila[3].value)
                                            except:
                                                radicado_numero = -1

                                            if Radicado.objects.filter(numero = radicado_numero).count() == 1:
                                                radicado = Radicado.objects.get(numero = radicado_numero)
                                            else:
                                                radicado = None


                                            area = None
                                            grado = None

                                            radicado_text = str(radicado.numero) if radicado != None else ''

                                            try:
                                                cedula = long(fila[15].value)
                                            except:
                                                resultado = 'Error en el numero de cedula'

                                            else:

                                                if Beneficiario.objects.filter(cedula = fila[15].value if fila[15].value != None else '').count() == 1:
                                                    beneficiario = Beneficiario.objects.get(cedula = fila[15].value)
                                                    if beneficiario.diplomado.id == 4:
                                                        beneficiario.region = region
                                                        beneficiario.radicado = radicado
                                                        beneficiario.radicado_text = radicado_text

                                                        beneficiario.departamento_text = fila[1].value
                                                        beneficiario.secretaria_text = fila[2].value
                                                        beneficiario.dane_ie_text = fila[4].value
                                                        beneficiario.ie_text = fila[5].value
                                                        beneficiario.dane_sede_text = fila[6].value
                                                        beneficiario.sede_text = fila[7].value
                                                        beneficiario.municipio_text = fila[8].value
                                                        beneficiario.zona_text = fila[9].value

                                                        beneficiario.formador = formador
                                                        beneficiario.grupo = grupo
                                                        beneficiario.apellidos =fila[13].value
                                                        beneficiario.nombres = fila[14].value
                                                        beneficiario.cedula=fila[15].value
                                                        beneficiario.correo=fila[16].value
                                                        beneficiario.telefono_fijo=fila[17].value
                                                        beneficiario.telefono_celular=fila[18].value
                                                        beneficiario.area = area
                                                        beneficiario.grado = grado
                                                        beneficiario.genero=fila[22].value
                                                        beneficiario.estado=fila[23].value
                                                        beneficiario.save()

                                                        resultado += 'Participante actualizado'

                                                    else:
                                                        resultado += 'El participante se encuentra registrado como docente'

                                                else:
                                                    Beneficiario.objects.create(diplomado = diplomado,region=region,radicado=radicado,radicado_text=radicado_text,
                                                                                departamento_text=fila[1].value,secretaria_text=fila[2].value,
                                                                                dane_ie_text=fila[4].value,ie_text=fila[5].value,dane_sede_text=fila[6].value,
                                                                                sede_text=fila[7].value,municipio_text=fila[8].value,zona_text=fila[9].value,
                                                                                formador=formador,grupo=grupo,apellidos=fila[13].value,
                                                                                nombres=fila[14].value,cedula=fila[15].value,correo=fila[16].value,
                                                                                telefono_fijo=fila[17].value,telefono_celular=fila[18].value,
                                                                                area=area,grado=grado,genero=fila[22].value,estado=fila[23].value)
                                                    resultado += 'Participante creado'

                                        else:
                                            resultado += 'No hay nombres del participante'

                                    else:
                                        resultado += 'No hay apellidos del participante'

                        else:
                            resultado = 'No se puede identificar el grupo'
                else:
                    resultado = 'No hay ningun formador con el numero de cedula'



            contenidos.append([
                'Escuela TIC',
                resultado,
                fila[0].value,
                fila[1].value,
                fila[2].value,
                fila[3].value,
                fila[4].value,
                fila[5].value,
                fila[6].value,
                fila[7].value,
                fila[8].value,
                fila[9].value,
                fila[10].value,
                fila[11].value,
                fila[12].value,
                fila[13].value,
                fila[14].value,
                fila[15].value,
                fila[16].value,
                fila[17].value,
                fila[18].value,
                fila[19].value,
                fila[20].value,
                fila[21].value,
                fila[22].value,
                fila[23].value
            ])

        usuario = User.objects.get(email=email)
        nombre = "Resultado carga masiva matrices"
        proceso = "FOR-MAS01"
        fecha = pytz.utc.localize(datetime.now())
        output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
        filename = unicode(fecha) + '.xlsx'
        carga.resultado.save(filename,File(output))
        carga.estado = 'Proceso concluido'


    else:
        carga.estado = 'El archivo no tiene la estructura necesaria.'

    carga.save()
    return "Carga masiva exitosa"

@app.task
def diplomas_escuela_tic():
    for obj in Beneficiario.objects.filter(diploma = '').filter(diplomado__id = 4):
        nombre_beneficiario = obj.get_full_name().upper()
        cedula = 'IDENTIFICADO(A) CON CÉDULA DE CIUDADANÍA NÚMERO ' + str(obj.cedula)

        municipio = ''

        if obj.radicado != None:
            municipio = obj.radicado.municipio.nombre.upper()
        else:
            if obj.municipio_text != None:
                municipio = obj.municipio_text.upper()

        date = timezone.now()

        fecha = municipio + ' ' + date.strftime('%d de %B de %Y').upper()

        fuente_primaria = ImageFont.truetype(settings.STATICFILES_DIRS[0] + '\\documentos\\DK_Lemon_Yellow_Sun.otf', 432)
        fuente_secundaria = ImageFont.truetype(settings.STATICFILES_DIRS[0] + '\\documentos\\DK_Lemon_Yellow_Sun.otf', 100)
        diploma = Image.open(settings.STATICFILES_DIRS[0]+'\\documentos\\Diploma.png')
        W,H = diploma.size
        nombre = ImageDraw.Draw(diploma)
        w,h = nombre.textsize(nombre_beneficiario,font=fuente_primaria,fill='white')

        nombre.text(((W-w)/2,1450),nombre_beneficiario,fill='white',font=fuente_primaria)

        w,h = nombre.textsize(cedula,font=fuente_secundaria,fill='white')

        nombre.text(((W-w)/2,1950),cedula,fill='white',font=fuente_secundaria)

        w,h = nombre.textsize(fecha,font=fuente_secundaria,fill='white')

        nombre.text(((W-w)/2,2970),fecha,fill='white',font=fuente_secundaria)


        diploma.thumbnail((1600,1281),Image.ANTIALIAS)

        output = StringIO.StringIO()

        diploma.save(output,format = 'PNG')

        obj.diploma.save(str(obj.cedula) + '.png',File(output))
    return "Verificacion completa de diplomas"
