#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.contrib import admin
from matrices.models import Area, Grado, Beneficiario
from matrices.models import CargaMasiva
from formadores.models import Grupos, Formador
import openpyxl
from sican.settings import base as settings
from administrativos.models import Administrativo
from sican.settings import base as settings
import PIL
from PIL import ImageFont
from PIL import Image
from PIL import ImageDraw
import StringIO
from django.core.files import File
from django.utils import timezone

# Register your models here.

admin.site.register(Area)
admin.site.register(Grado)

def grupos_colombia_aprende(modeladmin, request, queryset):
    beneficiarios = Beneficiario.objects.exclude(usuario_colombia_aprende = '').exclude(radicado = None)
    grupos_lista_id = beneficiarios.values_list('grupo__id',flat=True).distinct()
    grupos = Grupos.objects.filter(id__in = grupos_lista_id).exclude(formador__usuario_colombia_aprende = '')
    formadores_lista_id = grupos.values_list('formador__id',flat=True).distinct()
    formadores = Formador.objects.filter(id__in = formadores_lista_id)

    for grupo in grupos:
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Formato GD.xlsx')
        ws = wb.get_sheet_by_name('Hoja1')

        fila = 2
        for beneficiario in Beneficiario.objects.filter(id__in = beneficiarios.values_list('id',flat=True)).filter(grupo = grupo):
            ws['A'+str(fila)] = 'ASOANDES'
            ws['B'+str(fila)] = beneficiario.region.numero
            ws['C'+str(fila)] = beneficiario.radicado.municipio.departamento.nombre
            ws['D'+str(fila)] = beneficiario.radicado.municipio.nombre
            ws['E'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).cedula
            ws['F'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).nombres
            ws['G'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).apellidos
            ws['H'+str(fila)] = beneficiario.formador.lider.email
            ws['I'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).celular_personal
            ws['J'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).usuario_colombia_aprende
            ws['K'+str(fila)] = beneficiario.formador.cedula
            ws['L'+str(fila)] = beneficiario.formador.nombres
            ws['M'+str(fila)] = beneficiario.formador.apellidos
            ws['N'+str(fila)] = beneficiario.formador.correo_personal
            ws['O'+str(fila)] = beneficiario.formador.celular_personal
            ws['P'+str(fila)] = beneficiario.formador.usuario_colombia_aprende
            ws['Q'+str(fila)] = beneficiario.diplomado.nombre
            ws['R'+str(fila)] = beneficiario.cedula
            ws['S'+str(fila)] = beneficiario.nombres
            ws['T'+str(fila)] = beneficiario.apellidos
            ws['U'+str(fila)] = beneficiario.correo
            ws['V'+str(fila)] = beneficiario.telefono_celular
            ws['W'+str(fila)] = beneficiario.usuario_colombia_aprende
            fila += 1


        wb.save('C:\\Temp\\Colombia\\'+grupo.formador.codigo_ruta + '-' + grupo.nombre + '.xlsx')
grupos_colombia_aprende.short_description = 'Usuarios colombia aprende'

def diploma(modeladmin, request, queryset):

    for obj in Beneficiario.objects.filter(diploma = '').filter(diplomado__id = 4):
        nombre_beneficiario = obj.get_full_name().upper()
        cedula = 'IDENTIFICADO(A) CON CÉDULA DE CIUDADANÍA NÚMERO ' + str(obj.cedula)

        municipio = ''

        if obj.radicado != None:
            municipio = obj.radicado.municipio.nombre.upper()
        else:
            if obj.municipio_text != None:
                municipio = obj.municipio_text.upper()

        date = timezone.now()

        fecha = municipio + ' ' + date.strftime('%d de %B de %Y').upper()

        fuente_primaria = ImageFont.truetype(settings.STATICFILES_DIRS[0] + '\\documentos\\DK_Lemon_Yellow_Sun.otf', 432)
        fuente_secundaria = ImageFont.truetype(settings.STATICFILES_DIRS[0] + '\\documentos\\DK_Lemon_Yellow_Sun.otf', 100)
        diploma = Image.open(settings.STATICFILES_DIRS[0]+'\\documentos\\Diploma.png')
        W,H = diploma.size
        nombre = ImageDraw.Draw(diploma)
        w,h = nombre.textsize(nombre_beneficiario,font=fuente_primaria,fill='white')

        nombre.text(((W-w)/2,1450),nombre_beneficiario,fill='white',font=fuente_primaria)

        w,h = nombre.textsize(cedula,font=fuente_secundaria,fill='white')

        nombre.text(((W-w)/2,1950),cedula,fill='white',font=fuente_secundaria)

        w,h = nombre.textsize(fecha,font=fuente_secundaria,fill='white')

        nombre.text(((W-w)/2,2970),fecha,fill='white',font=fuente_secundaria)


        diploma.thumbnail((1600,1281),Image.ANTIALIAS)

        output = StringIO.StringIO()

        diploma.save(output,format = 'PNG')

        obj.diploma.save(str(obj.cedula) + '.png',File(output))



diploma.short_description = 'Generar Diploma'

class BeneficiarioAdmin(admin.ModelAdmin):
    list_display = ['nombres','apellidos']
    ordering = ['nombres']
    actions = [grupos_colombia_aprende,diploma]

admin.site.register(Beneficiario, BeneficiarioAdmin)
admin.site.register(CargaMasiva)