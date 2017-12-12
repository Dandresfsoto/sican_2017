#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from __future__ import absolute_import

from sican.celery import app
import openpyxl
from informes.functions import construir_reporte
from datetime import datetime
from usuarios.models import User
from django.core.files import File
import pytz
from productos.models import Diplomado
from region.models import Region
from formadores.models import Formador
from formadores.models import Grupos
from radicados.models import Radicado
from sican.settings import base as settings
import PIL
from PIL import ImageFont
from PIL import Image
from PIL import ImageDraw
import StringIO
from django.core.files import File
from django.utils import timezone
from vigencia2017.models import CargaMatriz, DaneSEDE, Grupos, Beneficiario, BeneficiarioCambio, Evidencia
from formadores.models import Contrato
from validate_email import validate_email
from formadores.models import Contrato as ContratoVigencia2017
from vigencia2017.models import Grupos as GruposVigencia2017
from vigencia2017.models import Beneficiario as BeneficiarioVigencia2017
from vigencia2017.models import Evidencia as EvidenciaVigencia2017
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from informes.models import InformesExcel
from productos.models import Entregable
from municipios.models import Municipio
from django.core.files import File
from sican.settings import base as settings
from django.core.files.uploadedfile import SimpleUploadedFile
from zipfile import ZipFile
from vigencia2017.models import CargaMasiva2017
from vigencia2017.models import Red
from vigencia2017.models import Pago
import json
from vigencia2017.models import Rechazo as RechazoVigencia2017
from openpyxl.comments import Comment
import xlsxwriter

@app.task
def carga_masiva_matrices(id,email_user):
    print(id)
    print(email_user)
    carga = CargaMatriz.objects.get(id = id)
    wb = openpyxl.load_workbook(filename = carga.archivo.file.name,read_only=True)
    sheets = wb.get_sheet_names()

    if 'InnovaTIC' in sheets and 'TecnoTIC' in sheets and 'DirecTIC' in sheets and 'ESCUELATIC DOCENTES INNOVADORES' in sheets and 'DocenTIC' in sheets and 'ESCUELA TIC FAMILIA' in sheets and 'SAN ANDRES' in sheets:

        titulos = ['DIPLOMADO',
                   'RESULTADO',
                   'REGION',
                   'CODIGO DANE SEDE EDUCATIVA',
                   'NOMBRE DE LA SEDE EDUCATIVA',
                   'CODIGO DANE INSTITUCION EDUCATIVA',
                   'NOMBRE INSTITUCION EDUCATIVA',
                   'CODIGO MUNICIPIO',
                   'MUNICIPIO',
                   'CODIGO DEPARTAMENTO',
                   'DEPARTAMENTO',
                   'SECRETARIA DE EDUCACION',
                   'ZONA (URBANA/RURAL)',
                   'CODIGO DEL GRUPO',
                   'NOMBRE DEL FORMADOR',
                   'NUMERO DE CEDULA DEL FORMADOR',
                   'APELLIDOS DEL DOCENTE',
                   'NOMBRES DEL DOCENTE',
                   'NUMERO DE CEDULA DEL DOCENTE',
                   'CORREO ELECTRONICO',
                   'TELEFONO FIJO',
                   'TELEFONO CELULAR',
                   'AREA CURRICULAR',
                   'GRADO',
                   'TIPO DE BENEFICIARIO',
                   'GENERO']

        formatos = ['General',
                   'General',
                   'General',
                   '0',
                   'General',
                   '0',
                   'General',
                   '0',
                   'General',
                   '0',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   '0',
                   'General',
                   'General',
                   '0',
                   'General',
                   'General',
                   'General',
                   '0',
                   '0',
                   'General',
                   'General']

        ancho_columnas =  [30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           ]

        contenidos = []

        for name in ['InnovaTIC','TecnoTIC','DirecTIC','ESCUELATIC DOCENTES INNOVADORES','DocenTIC']:
            ws = wb.get_sheet_by_name(name)

            for fila in ws.iter_rows(row_offset=5):

                resultado = ''


                try:
                    diplomado = Diplomado.objects.get(nombre__icontains = name)
                except:
                    resultado = 'No existe el diplomado'
                else:

                    try:
                        region = Region.objects.get(id = fila[0].value)
                    except:
                        resultado = 'Codigo invalido de región'
                    else:

                        try:
                            dane_sede = DaneSEDE.objects.get(dane_sede = fila[1].value)
                        except:
                            resultado = 'No existe el codigo DANE de la sede'
                        else:

                            ruta_archivo = fila[11].value.split('-')

                            if len(ruta_archivo) == 3:

                                ruta = ruta_archivo[0]+"-"+ruta_archivo[1]

                                try:
                                    formador = Formador.objects.get(cedula=fila[13].value)
                                except:
                                    resultado = 'No existe el numero de cedula del formador'
                                else:
                                    try:
                                        contrato = Contrato.objects.get(formador=formador, codigo_ruta = ruta)
                                    except:
                                        resultado = "No se pudo identificar el contrato del formador (contacte a sistemas)"
                                    else:
                                        try:
                                            numero_grupo = int(ruta_archivo[2])
                                        except:
                                            resultado = "Numero de grupo invalido"
                                        else:
                                            grupo, created = Grupos.objects.get_or_create(contrato = contrato,
                                                                                 diplomado = diplomado,
                                                                                 numero = numero_grupo)
                                            try:
                                                cedula = long(fila[16].value)
                                            except:
                                                resultado = "Error en el numero de cedula"
                                            else:

                                                if fila[14].value != None and fila[15].value != None:

                                                    nombres = fila[15].value.upper()
                                                    apellidos = fila[14].value.upper()

                                                    email = ''
                                                    telefono_fijo = fila[18].value
                                                    telefono_celular = fila[19].value

                                                    genero = fila[23].value

                                                    if validate_email(fila[17].value):
                                                        email = fila[17].value

                                                    try:
                                                        area = int(fila[20].value)
                                                    except:
                                                        area = None

                                                    try:
                                                        grado = int(fila[21].value)
                                                    except:
                                                        grado = None

                                                    try:
                                                        beneficiario = Beneficiario.objects.get(cedula = cedula)
                                                    except:
                                                        resultado = "Beneficiario creado"
                                                        Beneficiario.objects.create(region=region,
                                                                                    dane_sede=dane_sede,
                                                                                    grupo=grupo,
                                                                                    apellidos=apellidos,
                                                                                    nombres=nombres,
                                                                                    cedula=cedula,
                                                                                    correo=email,
                                                                                    telefono_fijo=telefono_fijo,
                                                                                    telefono_celular=telefono_celular,
                                                                                    area=area,
                                                                                    grado=grado,
                                                                                    genero=genero
                                                                                    )
                                                    else:

                                                        if beneficiario.nombres != nombres and beneficiario.apellidos != apellidos:
                                                            if beneficiario.dane_sede != dane_sede and beneficiario.grupo != grupo:
                                                                resultado = "Solicitud de cambio"
                                                                BeneficiarioCambio.objects.create(original = beneficiario,
                                                                                                  masivo = carga,
                                                                                                  region = region,
                                                                                                  dane_sede = dane_sede,
                                                                                                  grupo = grupo,
                                                                                                  apellidos = apellidos,
                                                                                                  nombres = nombres,
                                                                                                  cedula=cedula,
                                                                                                  correo=email,
                                                                                                  telefono_fijo=telefono_fijo,
                                                                                                  telefono_celular=telefono_celular,
                                                                                                  area=area,
                                                                                                  grado=grado,
                                                                                                  genero=genero
                                                                                                  )
                                                        else:
                                                            resultado = "Beneficirio actualizado"
                                                            beneficiario.dane_sede = dane_sede
                                                            beneficiario.grupo = grupo
                                                            beneficiario.apellidos = apellidos
                                                            beneficiario.nombres = nombres
                                                            beneficiario.correo = email
                                                            beneficiario.telefono_fijo = telefono_fijo
                                                            beneficiario.telefono_celular = telefono_celular
                                                            beneficiario.area = area
                                                            beneficiario.grado = grado
                                                            beneficiario.genero = genero
                                                            beneficiario.save()

                                                else:
                                                    resultado = "Error en los nombres o apellidos del docente"


                            else:
                                resultado = 'El codigo de ruta no se encuentra bien parametrizado'






                contenidos.append([
                    name,
                    resultado,
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    fila[12].value,
                    fila[13].value,
                    fila[14].value,
                    fila[15].value,
                    fila[16].value,
                    fila[17].value,
                    fila[18].value,
                    fila[19].value,
                    fila[20].value,
                    fila[21].value,
                    fila[22].value,
                    fila[23].value
                ])

        for name in ['ESCUELA TIC FAMILIA']:
            ws = wb.get_sheet_by_name(name)

            for fila in ws.iter_rows(row_offset=9):

                resultado = ''

                try:
                    diplomado = Diplomado.objects.get(id = 4)
                except:
                    resultado = 'No existe el diplomado'
                else:

                    try:
                        region = Region.objects.get(id=fila[0].value)
                    except:
                        resultado = 'Codigo invalido de región'
                    else:

                        try:
                            dane_sede = None
                        except:
                            resultado = 'No existe el codigo DANE de la sede'
                        else:

                            try:
                                municipio = Municipio.objects.get(codigo_municipio = fila[5].value)
                            except:
                                resultado = 'Error en el codigo de municipio'
                            else:
                                ruta_archivo = fila[11].value.split('-')

                                if len(ruta_archivo) == 3:

                                    ruta = ruta_archivo[0] + "-" + ruta_archivo[1]

                                    try:
                                        formador = Formador.objects.get(cedula=fila[13].value)
                                    except:
                                        resultado = 'No existe el numero de cedula del formador'
                                    else:
                                        try:
                                            contrato = Contrato.objects.get(formador=formador, codigo_ruta=ruta)
                                        except:
                                            resultado = "No se pudo identificar el contrato del formador (contacte a sistemas)"
                                        else:
                                            try:
                                                numero_grupo = int(ruta_archivo[2])
                                            except:
                                                resultado = "Numero de grupo invalido"
                                            else:
                                                grupo, created = Grupos.objects.get_or_create(contrato=contrato,
                                                                                              diplomado=diplomado,
                                                                                              numero=numero_grupo)
                                                try:
                                                    cedula = long(fila[16].value)
                                                except:
                                                    resultado = "Error en el numero de cedula"
                                                else:

                                                    if fila[14].value != None and fila[15].value != None:

                                                        nombres = fila[15].value.upper()
                                                        apellidos = fila[14].value.upper()

                                                        email = ''
                                                        telefono_fijo = fila[18].value
                                                        telefono_celular = fila[19].value

                                                        genero = fila[23].value

                                                        if validate_email(fila[17].value):
                                                            email = fila[17].value

                                                        try:
                                                            area = int(fila[20].value)
                                                        except:
                                                            area = None

                                                        try:
                                                            grado = int(fila[21].value)
                                                        except:
                                                            grado = None

                                                        try:
                                                            beneficiario = Beneficiario.objects.get(cedula=cedula)
                                                        except:
                                                            resultado = "Beneficiario creado"
                                                            Beneficiario.objects.create(region=region,
                                                                                        municipio = municipio,
                                                                                        dane_sede=dane_sede,
                                                                                        grupo=grupo,
                                                                                        apellidos=apellidos,
                                                                                        nombres=nombres,
                                                                                        cedula=cedula,
                                                                                        correo=email,
                                                                                        telefono_fijo=telefono_fijo,
                                                                                        telefono_celular=telefono_celular,
                                                                                        area=area,
                                                                                        grado=grado,
                                                                                        genero=genero
                                                                                        )
                                                        else:
                                                            if beneficiario.nombres != nombres and beneficiario.apellidos != apellidos:
                                                                if beneficiario.grupo != grupo:
                                                                    resultado = "Solicitud de cambio"
                                                                    BeneficiarioCambio.objects.create(
                                                                        original=beneficiario,
                                                                        masivo=carga,
                                                                        region=region,
                                                                        municipio=municipio,
                                                                        dane_sede=dane_sede,
                                                                        grupo=grupo,
                                                                        apellidos=apellidos,
                                                                        nombres=nombres,
                                                                        cedula=cedula,
                                                                        correo=email,
                                                                        telefono_fijo=telefono_fijo,
                                                                        telefono_celular=telefono_celular,
                                                                        area=area,
                                                                        grado=grado,
                                                                        genero=genero
                                                                        )
                                                            else:
                                                                resultado = "Beneficiario actualizado"
                                                                beneficiario.dane_sede = dane_sede
                                                                beneficiario.municipio = municipio
                                                                beneficiario.grupo = grupo
                                                                beneficiario.apellidos = apellidos
                                                                beneficiario.nombres = nombres
                                                                beneficiario.correo = email
                                                                beneficiario.telefono_fijo = telefono_fijo
                                                                beneficiario.telefono_celular = telefono_celular
                                                                beneficiario.area = area
                                                                beneficiario.grado = grado
                                                                beneficiario.genero = genero
                                                                beneficiario.save()

                                                    else:
                                                        resultado = "Error en los nombres o apellidos del docente"


                                else:
                                    resultado = 'El codigo de ruta no se encuentra bien parametrizado'

                contenidos.append([
                    name,
                    resultado,
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    fila[12].value,
                    fila[13].value,
                    fila[14].value,
                    fila[15].value,
                    fila[16].value,
                    fila[17].value,
                    fila[18].value,
                    fila[19].value,
                    fila[20].value,
                    fila[21].value,
                    fila[22].value,
                    fila[23].value
                ])

        usuario = User.objects.get(email=email_user)
        nombre = "Resultado carga masiva matrices"
        proceso = "FOR-MAS01"
        fecha = pytz.utc.localize(datetime.now())
        output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, nombre, fecha, usuario, proceso)
        filename = unicode(fecha) + '.xlsx'
        carga.resultado.save(filename, File(output))


    return "Carga masiva exitosa"

@app.task
def matriz_chequeo_vigencia_2017(email,id_contrato):
    usuario = User.objects.get(email=email)

    contrato = ContratoVigencia2017.objects.get(id = int(id_contrato))
    id_diplomado = ''
    nombre = ''


    grupos = GruposVigencia2017.objects.filter(contrato = contrato)


    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    informe.nombre = "Matriz de chequeo - " + contrato.formador.get_full_name() + " - " +contrato.nombre
    informe.save()
    fecha = informe.creacion
    output = StringIO.StringIO()
    dict_productos = []


    wb = openpyxl.load_workbook(filename = settings.STATICFILES_DIRS[0]+'/documentos/chequeo_2017.xlsx')
    ws_innovatic = wb.get_sheet_by_name('InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('TecnoTIC')
    ws_directic = wb.get_sheet_by_name('DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('EscuelaTIC')
    ws_escuelatic_innovadores = wb.get_sheet_by_name('ESCUELATIC DOCENTES INNOVADORES')
    ws_docentic = wb.get_sheet_by_name('DocenTIC')
    ws_san_andres = wb.get_sheet_by_name('SAN ANDRES')


    contadores = {'1':6,'2':6,'3':6,'4':6,'7':6,'8':6,'9':6}


    number = Style(font=Font(name='Calibri', size=12),
                   alignment=Alignment(horizontal='right', vertical='center', wrap_text=False),
                   number_format='0',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
                   )

    text = Style(font=Font(name='Calibri', size=12),
                 alignment=Alignment(horizontal='left', vertical='center', wrap_text=False),
                 number_format='General',
                 border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                               bottom=Side(style='thin'))
                 )

    validado = Style(font=Font(name='Calibri', size=12),
                     alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                     number_format='General',
                     fill=PatternFill(fill_type='solid', start_color='FF00B050', end_color='FF00B050')
                     )

    enviado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='General',
                    fill=PatternFill(fill_type='solid', start_color='FFC65911', end_color='FFC65911')
                    )

    cargado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='General',
                    fill=PatternFill(fill_type='solid', start_color='FFFFC000', end_color='FFFFC000')
                    )

    rechazado = Style(font=Font(name='Calibri', size=12),
                      alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                      number_format='General',
                      fill=PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                      )



    for beneficiario in BeneficiarioVigencia2017.objects.filter(grupo__in = grupos):

        id_diplomado = beneficiario.grupo.diplomado.id

        if id_diplomado == 1:
            ws = ws_innovatic
        elif id_diplomado == 2:
            ws = ws_tecnotic
        elif id_diplomado == 3:
            ws = ws_directic
        elif id_diplomado == 4:
            ws = ws_escuelatic
        elif id_diplomado == 7:
            ws = ws_escuelatic_innovadores
        elif id_diplomado == 8:
            ws = ws_docentic
        elif id_diplomado == 9:
            ws = ws_san_andres

        ws.cell(row=contadores[str(id_diplomado)], column=1, value = beneficiario.region.nombre.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=2, value = beneficiario.dane_sede.dane_sede if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=3, value = beneficiario.dane_sede.nombre_sede.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=4, value = beneficiario.dane_sede.dane_ie if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=5, value = beneficiario.dane_sede.nombre_ie.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=6, value = beneficiario.dane_sede.municipio.codigo_municipio if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=7, value = beneficiario.dane_sede.municipio.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=8, value = beneficiario.dane_sede.municipio.departamento.codigo_departamento if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=9, value = beneficiario.dane_sede.municipio.departamento.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=10, value = beneficiario.dane_sede.secretaria.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=11, value = beneficiario.dane_sede.zona.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=12, value = beneficiario.grupo.get_nombre_grupo())
        ws.cell(row=contadores[str(id_diplomado)], column=13, value = beneficiario.grupo.contrato.formador.get_full_name().upper())
        ws.cell(row=contadores[str(id_diplomado)], column=14, value = beneficiario.grupo.contrato.formador.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=15, value = beneficiario.apellidos.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=16, value = beneficiario.nombres.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=17, value = beneficiario.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=18, value = beneficiario.correo)
        ws.cell(row=contadores[str(id_diplomado)], column=19, value = beneficiario.telefono_fijo)
        ws.cell(row=contadores[str(id_diplomado)], column=20, value = beneficiario.telefono_celular)
        ws.cell(row=contadores[str(id_diplomado)], column=21, value = beneficiario.area)
        ws.cell(row=contadores[str(id_diplomado)], column=22, value = beneficiario.grado)
        ws.cell(row=contadores[str(id_diplomado)], column=23, value = beneficiario.grupo.diplomado.nombre)
        ws.cell(row=contadores[str(id_diplomado)], column=24, value = beneficiario.genero)

        for entregable in Entregable.objects.filter(sesion__nivel__diplomado__id = id_diplomado):

            estado = beneficiario.get_evidencia_state(id_entregable = entregable.id)

            if estado['state'] == 'cargado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="C")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = cargado
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).comment = Comment(estado['codigo'], "SICAN")
            elif estado['state'] == 'enviado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="E")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = enviado
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).comment = Comment(estado['red'] + " \n" + estado['codigo'], "SICAN")
            elif estado['state'] == 'validado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="V")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = validado
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).comment = Comment(estado['red'] + " \n" + estado['codigo'], "SICAN")
            elif estado['state'] == 'rechazado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="R")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = rechazado
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).comment = Comment(estado['red'] + ": " + " \n" + estado['codigo'] + ": " + " \n" + estado['observacion'], "SICAN")

        contadores[str(id_diplomado)] += 1


    if contadores['1'] == 6:
        wb.remove_sheet(ws_innovatic)
    if contadores['2'] == 6:
        wb.remove_sheet(ws_tecnotic)
    if contadores['3'] == 6:
        wb.remove_sheet(ws_directic)
    if contadores['4'] == 6:
        wb.remove_sheet(ws_escuelatic)
    if contadores['7'] == 6:
        wb.remove_sheet(ws_escuelatic_innovadores)
    if contadores['8'] == 6:
        wb.remove_sheet(ws_docentic)
    if contadores['9'] == 6:
        wb.remove_sheet(ws_san_andres)

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_valores_vigencia_2017(email,id_contrato):
    usuario = User.objects.get(email=email)

    contrato = ContratoVigencia2017.objects.get(id = int(id_contrato))
    id_diplomado = ''
    nombre = ''


    grupos = GruposVigencia2017.objects.filter(contrato = contrato)


    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    informe.nombre = "Matriz de pago - " + contrato.formador.get_full_name() + " - " +contrato.nombre
    informe.save()
    fecha = informe.creacion
    output = StringIO.StringIO()
    dict_productos = []


    wb = openpyxl.load_workbook(filename = settings.STATICFILES_DIRS[0]+'/documentos/chequeo_2017.xlsx')
    ws_innovatic = wb.get_sheet_by_name('InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('TecnoTIC')
    ws_directic = wb.get_sheet_by_name('DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('EscuelaTIC')
    ws_escuelatic_innovadores = wb.get_sheet_by_name('ESCUELATIC DOCENTES INNOVADORES')
    ws_docentic = wb.get_sheet_by_name('DocenTIC')
    ws_san_andres = wb.get_sheet_by_name('SAN ANDRES')


    contadores = {'1':6,'2':6,'3':6,'4':6,'7':6,'8':6,'9':6}




    validado = Style(font=Font(name='Calibri', size=12),
                     alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                     number_format='$ #,##0.00',
                     fill=PatternFill(fill_type='solid', start_color='FF00B050', end_color='FF00B050')
                     )

    enviado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='$ #,##0.00',
                    fill=PatternFill(fill_type='solid', start_color='FFFFC000', end_color='FFFFC000')
                    )


    rechazado = Style(font=Font(name='Calibri', size=12),
                      alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                      number_format='$ #,##0.00',
                      fill=PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                      )



    for beneficiario in BeneficiarioVigencia2017.objects.filter(grupo__in = grupos):

        id_diplomado = beneficiario.grupo.diplomado.id

        if id_diplomado == 1:
            ws = ws_innovatic
        elif id_diplomado == 2:
            ws = ws_tecnotic
        elif id_diplomado == 3:
            ws = ws_directic
        elif id_diplomado == 4:
            ws = ws_escuelatic
        elif id_diplomado == 7:
            ws = ws_escuelatic_innovadores
        elif id_diplomado == 8:
            ws = ws_docentic
        elif id_diplomado == 9:
            ws = ws_san_andres

        ws.cell(row=contadores[str(id_diplomado)], column=1, value = beneficiario.region.nombre.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=2, value = beneficiario.dane_sede.dane_sede if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=3, value = beneficiario.dane_sede.nombre_sede.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=4, value = beneficiario.dane_sede.dane_ie if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=5, value = beneficiario.dane_sede.nombre_ie.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=6, value = beneficiario.dane_sede.municipio.codigo_municipio if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=7, value = beneficiario.dane_sede.municipio.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=8, value = beneficiario.dane_sede.municipio.departamento.codigo_departamento if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=9, value = beneficiario.dane_sede.municipio.departamento.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=10, value = beneficiario.dane_sede.secretaria.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=11, value = beneficiario.dane_sede.zona.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=12, value = beneficiario.grupo.get_nombre_grupo())
        ws.cell(row=contadores[str(id_diplomado)], column=13, value = beneficiario.grupo.contrato.formador.get_full_name().upper())
        ws.cell(row=contadores[str(id_diplomado)], column=14, value = beneficiario.grupo.contrato.formador.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=15, value = beneficiario.apellidos.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=16, value = beneficiario.nombres.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=17, value = beneficiario.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=18, value = beneficiario.correo)
        ws.cell(row=contadores[str(id_diplomado)], column=19, value = beneficiario.telefono_fijo)
        ws.cell(row=contadores[str(id_diplomado)], column=20, value = beneficiario.telefono_celular)
        ws.cell(row=contadores[str(id_diplomado)], column=21, value = beneficiario.area)
        ws.cell(row=contadores[str(id_diplomado)], column=22, value = beneficiario.grado)
        ws.cell(row=contadores[str(id_diplomado)], column=23, value = beneficiario.grupo.diplomado.nombre)
        ws.cell(row=contadores[str(id_diplomado)], column=24, value = beneficiario.genero)

        for entregable in Entregable.objects.filter(sesion__nivel__diplomado__id = id_diplomado):

            estado = beneficiario.get_pago_state(id_entregable = entregable.id)

            if estado['state'] == 'reportado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value = beneficiario.get_pago_valor_entregable(entregable.id))
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = enviado
            elif estado['state'] == 'pago':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value = beneficiario.get_pago_valor_entregable(entregable.id))
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = validado
        contadores[str(id_diplomado)] += 1


    if contadores['1'] == 6:
        wb.remove_sheet(ws_innovatic)
    if contadores['2'] == 6:
        wb.remove_sheet(ws_tecnotic)
    if contadores['3'] == 6:
        wb.remove_sheet(ws_directic)
    if contadores['4'] == 6:
        wb.remove_sheet(ws_escuelatic)
    if contadores['7'] == 6:
        wb.remove_sheet(ws_escuelatic_innovadores)
    if contadores['8'] == 6:
        wb.remove_sheet(ws_docentic)
    if contadores['9'] == 6:
        wb.remove_sheet(ws_san_andres)

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_chequeo_vigencia_2017_total(email):
    usuario = User.objects.get(email=email)

    nombre = 'Matriz de chequeo total'


    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    informe.save()
    output = StringIO.StringIO()


    wb = openpyxl.load_workbook(filename = settings.STATICFILES_DIRS[0]+'/documentos/chequeo_2017.xlsx')
    ws_innovatic = wb.get_sheet_by_name('InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('TecnoTIC')
    ws_directic = wb.get_sheet_by_name('DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('EscuelaTIC')
    ws_escuelatic_innovadores = wb.get_sheet_by_name('ESCUELATIC DOCENTES INNOVADORES')
    ws_docentic = wb.get_sheet_by_name('DocenTIC')
    ws_san_andres = wb.get_sheet_by_name('SAN ANDRES')


    contadores = {'1':6,'2':6,'3':6,'4':6,'7':6,'8':6,'9':6}


    number = Style(font=Font(name='Calibri', size=12),
                   alignment=Alignment(horizontal='right', vertical='center', wrap_text=False),
                   number_format='0',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
                   )

    text = Style(font=Font(name='Calibri', size=12),
                 alignment=Alignment(horizontal='left', vertical='center', wrap_text=False),
                 number_format='General',
                 border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                               bottom=Side(style='thin'))
                 )

    validado = Style(font=Font(name='Calibri', size=12),
                     alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                     number_format='General',
                     fill=PatternFill(fill_type='solid', start_color='FF00B050', end_color='FF00B050')
                     )

    enviado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='General',
                    fill=PatternFill(fill_type='solid', start_color='FFC65911', end_color='FFC65911')
                    )

    cargado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='General',
                    fill=PatternFill(fill_type='solid', start_color='FFFFC000', end_color='FFFFC000')
                    )

    rechazado = Style(font=Font(name='Calibri', size=12),
                      alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                      number_format='General',
                      fill=PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                      )



    for beneficiario in BeneficiarioVigencia2017.objects.filter():

        id_diplomado = beneficiario.grupo.diplomado.id

        if id_diplomado == 1:
            ws = ws_innovatic
        elif id_diplomado == 2:
            ws = ws_tecnotic
        elif id_diplomado == 3:
            ws = ws_directic
        elif id_diplomado == 4:
            ws = ws_escuelatic
        elif id_diplomado == 7:
            ws = ws_escuelatic_innovadores
        elif id_diplomado == 8:
            ws = ws_docentic
        elif id_diplomado == 9:
            ws = ws_san_andres

        ws.cell(row=contadores[str(id_diplomado)], column=1, value = beneficiario.region.nombre.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=2, value = beneficiario.dane_sede.dane_sede if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=3, value = beneficiario.dane_sede.nombre_sede.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=4, value = beneficiario.dane_sede.dane_ie if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=5, value = beneficiario.dane_sede.nombre_ie.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=6, value = beneficiario.dane_sede.municipio.codigo_municipio if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=7, value = beneficiario.dane_sede.municipio.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=8, value = beneficiario.dane_sede.municipio.departamento.codigo_departamento if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=9, value = beneficiario.dane_sede.municipio.departamento.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=10, value = beneficiario.dane_sede.secretaria.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=11, value = beneficiario.dane_sede.zona.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=12, value = beneficiario.grupo.get_nombre_grupo())
        ws.cell(row=contadores[str(id_diplomado)], column=13, value = beneficiario.grupo.contrato.formador.get_full_name().upper())
        ws.cell(row=contadores[str(id_diplomado)], column=14, value = beneficiario.grupo.contrato.formador.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=15, value = beneficiario.apellidos.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=16, value = beneficiario.nombres.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=17, value = beneficiario.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=18, value = beneficiario.correo)
        ws.cell(row=contadores[str(id_diplomado)], column=19, value = beneficiario.telefono_fijo)
        ws.cell(row=contadores[str(id_diplomado)], column=20, value = beneficiario.telefono_celular)
        ws.cell(row=contadores[str(id_diplomado)], column=21, value = beneficiario.area)
        ws.cell(row=contadores[str(id_diplomado)], column=22, value = beneficiario.grado)
        ws.cell(row=contadores[str(id_diplomado)], column=23, value = beneficiario.grupo.diplomado.nombre)
        ws.cell(row=contadores[str(id_diplomado)], column=24, value = beneficiario.genero)

        for entregable in Entregable.objects.filter(sesion__nivel__diplomado__id = id_diplomado):

            estado = beneficiario.get_evidencia_state(id_entregable = entregable.id)

            if estado['state'] == 'cargado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="C")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = cargado
            elif estado['state'] == 'enviado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="E")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = enviado
            elif estado['state'] == 'validado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="V")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = validado
            elif estado['state'] == 'rechazado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="R")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = rechazado

        contadores[str(id_diplomado)] += 1


    if contadores['1'] == 6:
        wb.remove_sheet(ws_innovatic)
    if contadores['2'] == 6:
        wb.remove_sheet(ws_tecnotic)
    if contadores['3'] == 6:
        wb.remove_sheet(ws_directic)
    if contadores['4'] == 6:
        wb.remove_sheet(ws_escuelatic)
    if contadores['7'] == 6:
        wb.remove_sheet(ws_escuelatic_innovadores)
    if contadores['8'] == 6:
        wb.remove_sheet(ws_docentic)
    if contadores['9'] == 6:
        wb.remove_sheet(ws_san_andres)

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_valores_vigencia_2017_total(email):
    usuario = User.objects.get(email=email)

    nombre = 'Matriz de pago total'

    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    informe.save()
    output = StringIO.StringIO()


    wb = openpyxl.load_workbook(filename = settings.STATICFILES_DIRS[0]+'/documentos/chequeo_2017.xlsx')
    ws_innovatic = wb.get_sheet_by_name('InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('TecnoTIC')
    ws_directic = wb.get_sheet_by_name('DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('EscuelaTIC')
    ws_escuelatic_innovadores = wb.get_sheet_by_name('ESCUELATIC DOCENTES INNOVADORES')
    ws_docentic = wb.get_sheet_by_name('DocenTIC')
    ws_san_andres = wb.get_sheet_by_name('SAN ANDRES')



    contadores = {'1':6,'2':6,'3':6,'4':6,'7':6,'8':6,'9':6}




    validado = Style(font=Font(name='Calibri', size=12),
                     alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                     number_format='$ #,##0.00',
                     fill=PatternFill(fill_type='solid', start_color='FF00B050', end_color='FF00B050')
                     )

    enviado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='$ #,##0.00',
                    fill=PatternFill(fill_type='solid', start_color='FFFFC000', end_color='FFFFC000')
                    )


    rechazado = Style(font=Font(name='Calibri', size=12),
                      alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                      number_format='$ #,##0.00',
                      fill=PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                      )



    for beneficiario in BeneficiarioVigencia2017.objects.filter():

        id_diplomado = beneficiario.grupo.diplomado.id

        if id_diplomado == 1:
            ws = ws_innovatic
        elif id_diplomado == 2:
            ws = ws_tecnotic
        elif id_diplomado == 3:
            ws = ws_directic
        elif id_diplomado == 4:
            ws = ws_escuelatic
        elif id_diplomado == 7:
            ws = ws_escuelatic_innovadores
        elif id_diplomado == 8:
            ws = ws_docentic
        elif id_diplomado == 9:
            ws = ws_san_andres

        ws.cell(row=contadores[str(id_diplomado)], column=1, value = beneficiario.region.nombre.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=2, value = beneficiario.dane_sede.dane_sede if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=3, value = beneficiario.dane_sede.nombre_sede.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=4, value = beneficiario.dane_sede.dane_ie if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=5, value = beneficiario.dane_sede.nombre_ie.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=6, value = beneficiario.dane_sede.municipio.codigo_municipio if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=7, value = beneficiario.dane_sede.municipio.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=8, value = beneficiario.dane_sede.municipio.departamento.codigo_departamento if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=9, value = beneficiario.dane_sede.municipio.departamento.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=10, value = beneficiario.dane_sede.secretaria.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=11, value = beneficiario.dane_sede.zona.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=12, value = beneficiario.grupo.get_nombre_grupo())
        ws.cell(row=contadores[str(id_diplomado)], column=13, value = beneficiario.grupo.contrato.formador.get_full_name().upper())
        ws.cell(row=contadores[str(id_diplomado)], column=14, value = beneficiario.grupo.contrato.formador.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=15, value = beneficiario.apellidos.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=16, value = beneficiario.nombres.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=17, value = beneficiario.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=18, value = beneficiario.correo)
        ws.cell(row=contadores[str(id_diplomado)], column=19, value = beneficiario.telefono_fijo)
        ws.cell(row=contadores[str(id_diplomado)], column=20, value = beneficiario.telefono_celular)
        ws.cell(row=contadores[str(id_diplomado)], column=21, value = beneficiario.area)
        ws.cell(row=contadores[str(id_diplomado)], column=22, value = beneficiario.grado)
        ws.cell(row=contadores[str(id_diplomado)], column=23, value = beneficiario.grupo.diplomado.nombre)
        ws.cell(row=contadores[str(id_diplomado)], column=24, value = beneficiario.genero)

        for entregable in Entregable.objects.filter(sesion__nivel__diplomado__id = id_diplomado):

            estado = beneficiario.get_pago_state(id_entregable = entregable.id)

            if estado['state'] == 'reportado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value = beneficiario.get_pago_valor_entregable(entregable.id))
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = enviado
            elif estado['state'] == 'pago':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value = beneficiario.get_pago_valor_entregable(entregable.id))
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = validado
        contadores[str(id_diplomado)] += 1


    if contadores['1'] == 6:
        wb.remove_sheet(ws_innovatic)
    if contadores['2'] == 6:
        wb.remove_sheet(ws_tecnotic)
    if contadores['3'] == 6:
        wb.remove_sheet(ws_directic)
    if contadores['4'] == 6:
        wb.remove_sheet(ws_escuelatic)
    if contadores['7'] == 6:
        wb.remove_sheet(ws_escuelatic_innovadores)
    if contadores['8'] == 6:
        wb.remove_sheet(ws_docentic)
    if contadores['9'] == 6:
        wb.remove_sheet(ws_san_andres)

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def carga_masiva_evidencia(carga_id,id_contrato,id_entregable,user_id):
    carga = CargaMasiva2017.objects.get(id = carga_id)
    user = User.objects.get(id=user_id)
    contrato = Contrato.objects.get(id=id_contrato)
    entregable = Entregable.objects.get(id=id_entregable)

    soportes = ZipFile(carga.archivo, 'r')

    for soporte_info in soportes.infolist():
        soporte = soporte_info.filename
        try:
            cedula = soporte.split('/')[-1].split('.')[-2]
        except:
            pass
        else:
            try:
                beneficiario = BeneficiarioVigencia2017.objects.get(cedula=cedula)
            except:
                pass
            else:
                evidencias = EvidenciaVigencia2017.objects.filter(entregable=entregable, contrato=contrato)
                if evidencias.filter(beneficiarios_validados=beneficiario).count() == 0:
                    if evidencias.filter(beneficiarios_cargados=beneficiario).count() > 0:
                        evidencias_cargadas = evidencias.filter(beneficiarios_cargados=beneficiario)

                        for evidencia_cargada in evidencias_cargadas:
                            evidencia_cargada.beneficiarios_cargados.remove(beneficiario)
                            beneficiario.delete_pago_entregable(id_entregable=entregable.id)

                    archivo = SimpleUploadedFile(name=soporte, content=soportes.read(soporte_info))
                    evidencia = EvidenciaVigencia2017.objects.create(usuario=user, archivo=archivo,
                                                                     entregable=entregable,
                                                                     contrato=contrato)
                    evidencia.beneficiarios_cargados.add(beneficiario)
                    beneficiario.set_pago_entregable(id_entregable=entregable.id, evidencia_id=evidencia.id)

    return "Evidencias cargadas"

@app.task
def build_red(id_red):

    red = Red.objects.get(id = id_red)
    output = StringIO.StringIO()

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ids = []
    inicia = 0

    if red.diplomado.numero == 1:
        ids = [{'id':8,'letter':'Q'},
               {'id':9,'letter':'R'},
               {'id':11,'letter':'S'},
               {'id':20,'letter':'T'},
               {'id':12,'letter':'U'},
               {'id':262,'letter':'V'},
               {'id':14,'letter':'W'},
               {'id':15,'letter':'X'},
               {'id':16,'letter':'Y'},
               {'id':17,'letter':'Z'},
               {'id':27,'letter':'AA'},
               {'id':30,'letter':'AB'},
               {'id':33,'letter':'AC'},
               {'id':267,'letter':'AD'},
               {'id':36,'letter':'AE'},
               {'id':46,'letter':'AF'},
               {'id':58,'letter':'AG'},
               {'id':49,'letter':'AH'},
               {'id':59,'letter':'AI'},
               {'id':52,'letter':'AJ'},
               {'id':60,'letter':'AK'},
               {'id':55,'letter':'AL'},
               {'id':63,'letter':'AM'},
               {'id':66,'letter':'AN'},
               {'id':67,'letter':'AO'}]
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED INNOVATIC 2017.xlsx')
        ws = wb.get_sheet_by_name('RED InnovaTIC')
        inicia = 6
    elif red.diplomado.numero == 2:
        ids = [{'id':72,'letter':'M'},
               {'id':73,'letter':'N'},
               {'id':75,'letter':'O'},
               {'id':74,'letter':'P'},
               {'id':76,'letter':'Q'},
               {'id':77,'letter':'R'},
               {'id':84,'letter':'S'},
               {'id':85,'letter':'T'},
               {'id':78,'letter':'U'},
               {'id':89,'letter':'V'},
               {'id':97,'letter':'W'},
               {'id':98,'letter':'X'},
               {'id':93,'letter':'Y'},
               {'id':92,'letter':'Z'},
               {'id':99,'letter':'AA'},
               {'id':94,'letter':'AB'},
               {'id':100,'letter':'AC'},
               {'id':95,'letter':'AD'},
               {'id':104,'letter':'AE'},
               {'id':112,'letter':'AF'},
               {'id':106,'letter':'AG'},
               {'id':109,'letter':'AH'},
               {'id':108,'letter':'AI'},
               {'id':110,'letter':'AJ'},
               {'id':119,'letter':'AK'},
               {'id':124,'letter':'AL'},
               {'id':118,'letter':'AM'},
               {'id':120,'letter':'AN'},
               {'id':121,'letter':'AO'}]
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED TECNOTIC.xlsx')
        ws = wb.get_sheet_by_name('RED TecnoTIC')
        inicia = 6
    elif red.diplomado.numero == 3:
        ids = [{'id':127,'letter':'Q'},
               {'id':128,'letter':'R'},
               {'id':131,'letter':'S'},
               {'id':132,'letter':'T'},
               {'id':134,'letter':'U'},
               {'id':133,'letter':'V'},
               {'id':142,'letter':'W'},
               {'id':143,'letter':'X'},
               {'id':135,'letter':'Y'},
               {'id':144,'letter':'Z'},
               {'id':137,'letter':'AA'},
               {'id':145,'letter':'AB'},
               {'id':139,'letter':'AC'},
               {'id':147,'letter':'AD'},
               {'id':146,'letter':'AE'},
               {'id':156,'letter':'AF'},
               {'id':148,'letter':'AG'},
               {'id':149,'letter':'AH'},
               {'id':151,'letter':'AI'},
               {'id':150,'letter':'AJ'},
               {'id':294,'letter':'AK'},
               {'id':155,'letter':'AL'},
               {'id':157,'letter':'AM'},
               {'id':164,'letter':'AN'},
               {'id':165,'letter':'AO'},
               {'id': 159, 'letter': 'AP'},
               {'id': 162, 'letter': 'AQ'},
               {'id': 161, 'letter': 'AR'},
               {'id': 166, 'letter': 'AS'},
               {'id': 167, 'letter': 'AT'},
               {'id': 171, 'letter': 'AU'},
               {'id': 170, 'letter': 'AV'},
               {'id': 169, 'letter': 'AW'},
               ]
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED DIRECTIC 2017.xlsx')
        ws = wb.get_sheet_by_name('RED DirecTIC')
        inicia = 6

    elif red.diplomado.numero == 4:
        ids = [{'id':224,'letter':'R'},
               {'id':228,'letter':'S'}]

        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED FAMILIA 2017.xlsx')
        ws = wb.get_sheet_by_name('RED FAMILIA')
        inicia = 6

    elif red.diplomado.numero == 6:
        ids = [{'id':258,'letter':'M'},
               {'id':234,'letter':'N'},
               {'id':235,'letter':'O'},
               {'id':236,'letter':'P'},
               {'id':239,'letter':'Q'},
               {'id':242,'letter':'R'},
               {'id':244,'letter':'S'},
               {'id':247,'letter':'T'},
               {'id':248,'letter':'U'},
               {'id':255,'letter':'V'},
               {'id':256,'letter':'W'}]

        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED BOGOTA.xlsx')
        ws = wb.get_sheet_by_name('RED BOGOTA')
        inicia = 6

    elif red.diplomado.numero == 7:
        ids = [{'id':304,'letter':'Q'},
               {'id':305,'letter':'R'},
               {'id':306,'letter':'S'},
               {'id':307,'letter':'T'},
               {'id':308,'letter':'U'},
               {'id':309,'letter':'V'},
               {'id':310,'letter':'W'},
               {'id':311,'letter':'X'},
               {'id':312,'letter':'Y'},
               {'id':313,'letter':'Z'},
               {'id':314,'letter':'AA'},
               {'id':315,'letter': 'AB'},
               {'id':316,'letter': 'AC'},
               {'id':317,'letter': 'AD'},
               {'id':318,'letter': 'AE'},
               {'id':319,'letter': 'AF'},
               {'id': 335, 'letter': 'AG'},
               {'id': 336, 'letter': 'AH'},
               {'id': 337, 'letter': 'AI'},
               {'id': 338, 'letter': 'AJ'}
               ]

        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/ESCUELATIC DOCENTES INNOVADORES 2017.xlsx')
        ws = wb.get_sheet_by_name('ESCUELATIC DOCENTES INNOVADORES')
        inicia = 6

    elif red.diplomado.numero == 8:
        ids = [{'id':334,'letter':'R'},
               {'id':320,'letter':'S'},
               {'id':321,'letter':'T'},
               {'id':322,'letter':'U'},
               {'id':323,'letter':'V'},
               {'id':324,'letter':'W'},
               {'id':325,'letter':'X'},
               {'id':326,'letter':'Y'},
               {'id':327,'letter':'Z'},
               {'id':328,'letter':'AA'},
               {'id':329,'letter':'AB'},
               {'id':330,'letter': 'AC'},
               {'id':331,'letter': 'AD'},
               {'id':332,'letter': 'AE'},
               {'id':333,'letter': 'AF'}
               ]

        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED DOCENTIC 2017.xlsx')
        ws = wb.get_sheet_by_name('DOCENTIC')
        inicia = 6

    elif red.diplomado.numero == 9:
        ids = [{'id':339,'letter':'R'},
               {'id':340,'letter':'S'},
               {'id':341,'letter':'T'},
               {'id':342,'letter':'U'},
               {'id':345,'letter':'V'},
               {'id':348,'letter':'W'},
               {'id':350,'letter':'X'},
               {'id':353,'letter':'Y'},
               {'id':354,'letter':'Z'},
               {'id':361,'letter':'AA'},
               {'id':362,'letter':'AB'}
               ]

        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED SAN ANDRES 2017.xlsx')
        ws = wb.get_sheet_by_name('SAN ANDRES')
        inicia = 6

    evidencias_total = EvidenciaVigencia2017.objects.filter(red_id=id_red)
    beneficiarios_id = evidencias_total.exclude(beneficiarios_cargados = None).values_list('beneficiarios_cargados__id',flat=True).distinct()


    i = 0 + inicia
    for beneficiario_id in beneficiarios_id:
        beneficiario = BeneficiarioVigencia2017.objects.get(id = beneficiario_id)
        evidencias = evidencias_total.filter(beneficiarios_cargados__id = beneficiario_id)


        ws.cell('A'+str(i)).value = i - inicia + 1
        ws.cell('B'+str(i)).value = beneficiario.region.nombre.upper()
        ws.cell('C'+str(i)).value = beneficiario.dane_sede.dane_sede if red.diplomado.id != 4 else 'N/A'
        ws.cell('C' + str(i)).number_format = '0' if red.diplomado.id != 4 else 'General'
        ws.cell('D'+str(i)).value = beneficiario.dane_sede.nombre_sede if red.diplomado.id != 4 else 'N/A'
        ws.cell('E'+str(i)).value = beneficiario.dane_sede.dane_ie if red.diplomado.id != 4 else 'N/A'
        ws.cell('E' + str(i)).number_format = '0' if red.diplomado.id != 4 else 'General'
        ws.cell('F'+str(i)).value = beneficiario.dane_sede.nombre_ie if red.diplomado.id != 4 else 'N/A'
        ws.cell('G'+str(i)).value = beneficiario.dane_sede.municipio.codigo_municipio if red.diplomado.id != 4 else beneficiario.municipio.codigo_municipio
        ws.cell('G'+str(i)).number_format = '0'
        ws.cell('H'+str(i)).value = beneficiario.dane_sede.municipio.nombre.upper() if red.diplomado.id != 4 else beneficiario.municipio.nombre.upper()
        ws.cell('I'+str(i)).value = beneficiario.dane_sede.municipio.departamento.codigo_departamento if red.diplomado.id != 4 else beneficiario.municipio.departamento.codigo_departamento
        ws.cell('I' + str(i)).number_format = '0'
        ws.cell('J'+str(i)).value = beneficiario.dane_sede.municipio.departamento.nombre.upper() if red.diplomado.id != 4 else beneficiario.municipio.departamento.nombre.upper()

        ws.cell('K'+str(i)).value = beneficiario.grupo.get_nombre_grupo()
        ws.cell('L'+str(i)).value = beneficiario.grupo.contrato.formador.get_full_name().upper()
        ws.cell('M' + str(i)).value = beneficiario.grupo.contrato.formador.cedula
        ws.cell('N' + str(i)).value = beneficiario.apellidos.upper()
        ws.cell('O' + str(i)).value = beneficiario.nombres.upper()
        ws.cell('P' + str(i)).value = beneficiario.cedula

        for id in ids:
            evidencia = evidencias.filter(entregable__id = id['id']).order_by('id')
            if evidencia.count() == 1:
                ws.cell( id['letter'] + str(i)).value = 'SIC-' + str(evidencia[0].id)
                ws.cell( id['letter'] + str(i)).hyperlink = 'https://sican.asoandes.org' + evidencia[0].get_archivo_url()

        i += 1


    wb.save(output)
    filename = 'RED-VIG2017-' + unicode(red.id) + '-'+ red.region.nombre +'.xlsx'
    red.archivo.save(filename,File(output))

    return "Generado RED-" + str(id_red)

@app.task
def retroalimentacion_red(id_red):

    red = Red.objects.get(id = id_red)
    evidencias_red = EvidenciaVigencia2017.objects.filter(red_id = id_red)

    evidencias_red.filter(beneficiarios_cargados=None).update(completa = True)

    wb = openpyxl.load_workbook(red.archivo_retroalimentacion.file)
    ws = wb.get_active_sheet()
    ids = []
    inicia = 0

    if not red.producto_final:

        if red.diplomado.numero == 1:
            ids = [{'id': 8, 'letter': 'Q'},
                   {'id': 9, 'letter': 'R'},
                   {'id': 11, 'letter': 'S'},
                   {'id': 20, 'letter': 'T'},
                   {'id': 12, 'letter': 'U'},
                   {'id': 262, 'letter': 'V'},
                   {'id': 14, 'letter': 'W'},
                   {'id': 15, 'letter': 'X'},
                   {'id': 16, 'letter': 'Y'},
                   {'id': 17, 'letter': 'Z'},
                   {'id': 27, 'letter': 'AA'},
                   {'id': 30, 'letter': 'AB'},
                   {'id': 33, 'letter': 'AC'},
                   {'id': 267, 'letter': 'AD'},
                   {'id': 36, 'letter': 'AE'},
                   {'id': 46, 'letter': 'AF'},
                   {'id': 58, 'letter': 'AG'},
                   {'id': 49, 'letter': 'AH'},
                   {'id': 59, 'letter': 'AI'},
                   {'id': 52, 'letter': 'AJ'},
                   {'id': 60, 'letter': 'AK'},
                   {'id': 55, 'letter': 'AL'},
                   {'id': 63, 'letter': 'AM'},
                   {'id': 66, 'letter': 'AN'},
                   {'id': 67, 'letter': 'AO'}]

            inicia = 6
        elif red.diplomado.numero == 2:
            ids = [{'id':72,'letter':'M'},
                   {'id':73,'letter':'N'},
                   {'id':75,'letter':'O'},
                   {'id':74,'letter':'P'},
                   {'id':76,'letter':'Q'},
                   {'id':77,'letter':'R'},
                   {'id':84,'letter':'S'},
                   {'id':85,'letter':'T'},
                   {'id':78,'letter':'U'},
                   {'id':89,'letter':'V'},
                   {'id':97,'letter':'W'},
                   {'id':98,'letter':'X'},
                   {'id':93,'letter':'Y'},
                   {'id':92,'letter':'Z'},
                   {'id':99,'letter':'AA'},
                   {'id':94,'letter':'AB'},
                   {'id':100,'letter':'AC'},
                   {'id':95,'letter':'AD'},
                   {'id':104,'letter':'AE'},
                   {'id':112,'letter':'AF'},
                   {'id':106,'letter':'AG'},
                   {'id':109,'letter':'AH'},
                   {'id':108,'letter':'AI'},
                   {'id':110,'letter':'AJ'},
                   {'id':119,'letter':'AK'},
                   {'id':124,'letter':'AL'},
                   {'id':118,'letter':'AM'},
                   {'id':120,'letter':'AN'},
                   {'id':121,'letter':'AO'}]
            inicia = 6
        elif red.diplomado.numero == 3:
            ids = [{'id': 127, 'letter': 'Q'},
                   {'id': 128, 'letter': 'R'},
                   {'id': 131, 'letter': 'S'},
                   {'id': 132, 'letter': 'T'},
                   {'id': 134, 'letter': 'U'},
                   {'id': 133, 'letter': 'V'},
                   {'id': 142, 'letter': 'W'},
                   {'id': 143, 'letter': 'X'},
                   {'id': 135, 'letter': 'Y'},
                   {'id': 144, 'letter': 'Z'},
                   {'id': 137, 'letter': 'AA'},
                   {'id': 145, 'letter': 'AB'},
                   {'id': 139, 'letter': 'AC'},
                   {'id': 147, 'letter': 'AD'},
                   {'id': 146, 'letter': 'AE'},
                   {'id': 156, 'letter': 'AF'},
                   {'id': 148, 'letter': 'AG'},
                   {'id': 149, 'letter': 'AH'},
                   {'id': 151, 'letter': 'AI'},
                   {'id': 150, 'letter': 'AJ'},
                   {'id': 294, 'letter': 'AK'},
                   {'id': 155, 'letter': 'AL'},
                   {'id': 157, 'letter': 'AM'},
                   {'id': 164, 'letter': 'AN'},
                   {'id': 165, 'letter': 'AO'},
                   {'id': 159, 'letter': 'AP'},
                   {'id': 162, 'letter': 'AQ'},
                   {'id': 161, 'letter': 'AR'},
                   {'id': 166, 'letter': 'AS'},
                   {'id': 167, 'letter': 'AT'},
                   {'id': 171, 'letter': 'AU'},
                   {'id': 170, 'letter': 'AV'},
                   {'id': 169, 'letter': 'AW'},
                   ]
            inicia = 6
        elif red.diplomado.numero == 4:
            ids = [{'id':224,'letter':'R'},
               {'id':228,'letter':'S'}]
            inicia = 6

        elif red.diplomado.numero == 6:
            ids = [{'id': 258, 'letter': 'M'},
                   {'id': 234, 'letter': 'N'},
                   {'id': 235, 'letter': 'O'},
                   {'id': 236, 'letter': 'P'},
                   {'id': 239, 'letter': 'Q'},
                   {'id': 242, 'letter': 'R'},
                   {'id': 244, 'letter': 'S'},
                   {'id': 247, 'letter': 'T'},
                   {'id': 248, 'letter': 'U'},
                   {'id': 255, 'letter': 'V'},
                   {'id': 256, 'letter': 'W'}]
            inicia = 6

        elif red.diplomado.numero == 7:
            ids = [{'id': 304, 'letter': 'Q'},
                   {'id': 305, 'letter': 'R'},
                   {'id': 306, 'letter': 'S'},
                   {'id': 307, 'letter': 'T'},
                   {'id': 308, 'letter': 'U'},
                   {'id': 309, 'letter': 'V'},
                   {'id': 310, 'letter': 'W'},
                   {'id': 311, 'letter': 'X'},
                   {'id': 312, 'letter': 'Y'},
                   {'id': 313, 'letter': 'Z'},
                   {'id': 314, 'letter': 'AA'},
                   {'id': 315, 'letter': 'AB'},
                   {'id': 316, 'letter': 'AC'},
                   {'id': 317, 'letter': 'AD'},
                   {'id': 318, 'letter': 'AE'},
                   {'id': 319, 'letter': 'AF'},
                   {'id': 335, 'letter': 'AG'},
                   {'id': 336, 'letter': 'AH'},
                   {'id': 337, 'letter': 'AI'},
                   {'id': 338, 'letter': 'AJ'}
                   ]
            inicia = 6

        elif red.diplomado.numero == 8:
            ids = [{'id': 334, 'letter': 'R'},
                   {'id': 320, 'letter': 'S'},
                   {'id': 321, 'letter': 'T'},
                   {'id': 322, 'letter': 'U'},
                   {'id': 323, 'letter': 'V'},
                   {'id': 324, 'letter': 'W'},
                   {'id': 325, 'letter': 'X'},
                   {'id': 326, 'letter': 'Y'},
                   {'id': 327, 'letter': 'Z'},
                   {'id': 328, 'letter': 'AA'},
                   {'id': 329, 'letter': 'AB'},
                   {'id': 330, 'letter': 'AC'},
                   {'id': 331, 'letter': 'AD'},
                   {'id': 332, 'letter': 'AE'},
                   {'id': 333, 'letter': 'AF'}
                   ]
            inicia = 6


        for row in ws.iter_rows(row_offset=inicia-1):
            beneficiario = {}
            for cell in row:
                beneficiario[cell.column] = cell.value

            try:
                beneficiario_object = Beneficiario.objects.get(cedula = beneficiario['P'])

            except:
                pass

            else:

                evidencias_beneficiario = evidencias_red.filter(beneficiarios_cargados=beneficiario_object)

                id_escenciales = {}

                for id in ids:
                    id_escenciales[id['id']] = id['letter']


                for evidencia_beneficiario in evidencias_beneficiario:
                    if evidencia_beneficiario.entregable.id in list(id_escenciales.keys()):
                        id_entregable = evidencia_beneficiario.entregable.id
                        observacion = beneficiario[id_escenciales[id_entregable]]

                        if observacion == None or observacion == '':
                            pass
                        elif observacion.replace(' ','').lower() == 'ok':
                            evidencia_beneficiario.beneficiarios_validados.add(beneficiario_object)
                        else:
                            if RechazoVigencia2017.objects.filter(beneficiario_rechazo=beneficiario_object,red_id=red.id,evidencia_id=evidencia_beneficiario.id).count() == 0:
                                rechazo_object, created = RechazoVigencia2017.objects.get_or_create(
                                    beneficiario_rechazo=beneficiario_object,
                                    observacion=observacion,
                                    red_id=red.id,
                                    evidencia_id=evidencia_beneficiario.id)
                                if created:
                                    evidencia_beneficiario.beneficiarios_rechazados.add(rechazo_object)

                    else:
                        evidencia_beneficiario.completa = True
                        evidencia_beneficiario.save()


                    if evidencia_beneficiario.beneficiarios_cargados.all().count() == evidencia_beneficiario.beneficiarios_validados.all().count() + evidencia_beneficiario.beneficiarios_rechazados.all().count():
                        evidencia_beneficiario.completa = True
                        evidencia_beneficiario.save()


    red.retroalimentacion = True
    red.save()

    return "Retroalimentado RED-" + str(id_red)

@app.task
def set_pago(pagos,corte_id):
    pagos = json.loads(pagos)
    for pago in pagos:
        pago_object = Pago.objects.get(id=pago)
        pago_object.corte_id = corte_id
        pago_object.save()
    return "Corte efectuado"

@app.task
def matriz_chequeo_virtual_compilada_2017(email):
    usuario = User.objects.get(email=email)
    nombre = 'Compilado evidencias virtuales SICAN VIG 2017'

    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO.StringIO()
    dict_productos = []


    dict_productos_innovatic = [{'letter':'I','id':11,'nombre':'N1S1 - Guía: Valorando mis Competencias TIC'},
                                {'letter':'J','id':20,'nombre':'N1S1 - Hoja de Ruta para fortalecer las competencias TIC'},
                                {'letter':'K','id':262,'nombre':'N1S2 - Guía: construyendo lecciones de innovación educativa'},
                                {'letter':'L','id':15,'nombre':'N1S3 - Documento: Decálogo de un Proyecto Innovador.'},
                                {'letter':'M','id':16,'nombre':'N1S3 - Guía: Potenciando mi Experiencia Educativa de primera fase'},

                                {'letter':'N','id':267,'nombre':'N2S3 - Guía Construyendo mi PLE'},

                                {'letter':'O','id':58,'nombre':'N3S1 - Formato de recopilación de resultados: "Lo aprendido"'},
                                {'letter':'P','id':59,'nombre':'N3S2 - Enlace o imagen interactiva: Ruta de sostenibilidad'},
                                {'letter':'Q','id':60,'nombre':'N3S3 - Gráfica del PLE realizada en Mindomo'},

                                {'letter':'R','id':67,'nombre':'N4S2 - Registro fotográfico: Evento de socialización'},
                                ]



    dict_productos_directic = [{'letter':'I','id':132,'nombre':'N1S2 - Documento: presentación del capítulo  asignado a cada equipo.'},
                                {'letter':'J','id':134,'nombre':'N1S3 - El Gestor del Plan Estratégico TIC en su primer componente: Identificación Institucional.'},
                                {'letter':'K','id':142,'nombre':'N1S3 - Gestor del Plan Estratégico TIC con los dos componentes de laCaracterizacióndiligenciados: 1. Identificación institucional TIC y 2. Caracterización proyectos TIC'},
                                {'letter':'L','id':143,'nombre':'N1S4 - Gestor del Plan Estratégico TIC diligenciado en el componente Indicadores Educativos.'},
                                {'letter':'M','id':144,'nombre':'N1S5 - Gestor de Plan Estratégico de TIC: componente, valoración de la gestión institucional.'},
                                {'letter':'N','id':145,'nombre':'N1S6 - Gestor del Plan Estratégico TIC componente Matriz FODA'},

                                {'letter':'O','id':147,'nombre':'N2S1 - Gestor del Plan Estratégico TIC en el componente: Responsables y plazos.'},
                                {'letter':'P','id':156,'nombre':'N2S1 - Gestor del Plan Estratégico TIC. Componenete Comité de gestores TIC'},
                                {'letter':'Q','id':149,'nombre':'N2S2 - Gestor del Plan Estratégico TIC. Objetivo general del Plan, objetivos específicos'},
                                {'letter':'R','id':151,'nombre':'N2S3 - Gestor del Plan Estratégico TIC. Etapa: análisis y formulación del Plan los items: proyectos, actividades, áreas de gestión, y plazos (Inicio y finalización actividades y proyectos)'},

                                {'letter':'S','id':294,'nombre':'N3S1 - Gestor del Plan Estratégico TIC, componente Tipos de Cooperación.'},
                                {'letter':'T','id':164,'nombre':'N3S2 - Diligenciar la plantilla del documento compartido en Google Drive denominado: Aliados Estratégicos.'},
                                {'letter':'U','id':165,'nombre':'N3S3 - Gestor Plan Estratégico TIC'},
                                {'letter':'V','id':162,'nombre':'N3S4 - Brochure'},
                                {'letter':'W','id':166,'nombre':'N3S4 - Audio, video o Guión de la entrevista ejecutada con la en entidad pública o privada'},

                                {'letter':'X','id':171,'nombre':'N4S1 -  Dos fotografías del desarrollo de la socialización del Plan Estratégico TIC'},
                                {'letter':'Y','id':170,'nombre':'N4S2 -  Acta diligenciada y firmada por los integrantes del Comité de Gestores TIC.'},
                                ]

    dict_productos_escuela_tic_docentes_innovadores = [
                                {'letter': 'I', 'id': 304, 'nombre': 'M1 - ACTA DE COMPROMISO'},
                                {'letter': 'J', 'id': 305,'nombre': 'M1 - FORO EN LA PLATAFORMA'},
                                {'letter': 'K', 'id': 306,'nombre': 'M1 - Tarea:Actividad 3. El Coaching como estrategia de liderazgo docente.'},
                                {'letter': 'L', 'id': 307,'nombre': 'M1 - Cuestionario:Prueba Módulo 1 “Conocimientos adquiridos”'},

                                {'letter': 'M', 'id': 308,'nombre': 'M2 - Documento Word con las características esenciales de los procesos de desarrollo profesional docente.'},
                                {'letter': 'N', 'id': 309, 'nombre': 'M2 - Documento PDF, rúbrica de Categorización.'},
                                {'letter': 'O', 'id': 310,'nombre': 'M2 - Foro Wiki en la plataforma.'},
                                {'letter': 'P', 'id': 311,'nombre': 'M2 - Cuestionario:Prueba Módulo 2 Conocimientos adquiridos'},

                                {'letter': 'Q', 'id': 312,'nombre': 'M3 - Pantallazo de Polígono.'},
                                {'letter': 'R', 'id': 313,'nombre': 'M3 - Foro Wiki en la plataforma.'},
                                {'letter': 'S', 'id': 314,'nombre': 'M3 - Pantallazo de Facebook en la cual se vea reflejado el proceso de capacitación.'},
                                {'letter': 'T', 'id': 315,'nombre': 'M3 - Cuestionario:Prueba Módulo 3 “Evaluando y potenciando proyectos educativos, unidades didácticas y experiencias como Líder Tic.”'},

                                {'letter': 'U', 'id': 316, 'nombre': 'M4 - Mapa Mental Digital realizado en Mindmeister'},
                                {'letter': 'V', 'id': 317, 'nombre': 'M4 - Tarea:Actividad 13. Construyendo Comunidades de Práctica como estrategia de innovación'},
                                {'letter': 'W', 'id': 318,'nombre': 'M4 - Cuestionario:Prueba Módulo 4 “Conocimientos adquiridos”'},
                                {'letter': 'X', 'id': 319,'nombre': 'M4 - Cuestionario:Prueba Final. “Presentación de la prueba final”'},
                                ]





    wb = xlsxwriter.Workbook(output)
    ws_innovatic = wb.add_worksheet('INNOVATIC')
    ws_directic = wb.add_worksheet('DIRECTIC')
    ws_escuela_tic_docentes_innovadores = wb.add_worksheet('ESCUELATIC DOCENTES INNOVADORES')


    text = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False})

    number = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'right', 'valign':'vcenter', 'text_wrap':False,'num_format':'0'})

    validado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#00B050'})

    gris = wb.add_format({'font_name':'Arial Narrow','border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#808080', 'font_color':'#FFFFFF'})


    verde = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#008000', 'font_color':'#FFFFFF'})



    for id_diplomado in [1,3,7]:

        if id_diplomado == 1:
            ws = ws_innovatic
            dict_productos = dict_productos_innovatic

            for producto in dict_productos_innovatic:
                ws.write(producto['letter']+'1',producto['nombre'],verde)
                ws.set_column(producto['letter']+':'+producto['letter'],14)


        elif id_diplomado == 3:
            ws = ws_directic
            dict_productos = dict_productos_directic

            for producto in dict_productos_directic:
                ws.write(producto['letter']+'1',producto['nombre'],verde)
                ws.set_column(producto['letter']+':'+producto['letter'],14)


        elif id_diplomado == 7:
            ws = ws_escuela_tic_docentes_innovadores
            dict_productos = dict_productos_escuela_tic_docentes_innovadores

            for producto in dict_productos_escuela_tic_docentes_innovadores:
                ws.write(producto['letter']+'1',producto['nombre'],verde)
                ws.set_column(producto['letter']+':'+producto['letter'],14)



        ws.write('A1','REGIÓN',gris)
        ws.set_column('A:A',11)
        ws.set_row(0,90)

        ws.write('B1','DEPARTAMENTO',gris)
        ws.set_column('B:B',23)

        ws.write('C1','MUNICIPIO',gris)
        ws.set_column('C:C',18)

        ws.write('D1','NOMBRE DEL FORMADOR',gris)
        ws.set_column('D:D',36)

        ws.write('E1','NUMERO DE CEDULA DEL FORMADOR',gris)
        ws.set_column('E:E',13)

        ws.write('F1','APELLIDOS DEL BENEFICIARIO',verde)
        ws.set_column('F:F',25)

        ws.write('G1','NOMBRES DEL BENEFICIARIO',verde)
        ws.set_column('G:G',25)

        ws.write('H1','NUMERO DE CEDULA DEL BENEFICIARIO',verde)
        ws.set_column('H:H',14)

        i = 2

        for beneficiario in Beneficiario.objects.filter(grupo__diplomado__id = id_diplomado).order_by('grupo__contrato__formador'):
            ws.write('A'+str(i), beneficiario.grupo.contrato.region.nombre.upper(),text)

            ws.write('B'+str(i), beneficiario.dane_sede.municipio.departamento.nombre.upper() if beneficiario.dane_sede != None else "",text)

            ws.write('C'+str(i), beneficiario.dane_sede.municipio.nombre.upper() if beneficiario.dane_sede != None else "",text)

            ws.write('D'+str(i), beneficiario.grupo.contrato.formador.get_full_name(),text)

            ws.write('E'+str(i), beneficiario.grupo.contrato.formador.cedula,number)

            ws.write('F'+str(i), beneficiario.apellidos,text)

            ws.write('G'+str(i), beneficiario.nombres,text)

            ws.write('H'+str(i), beneficiario.cedula,number)


            for producto in dict_productos:
                entregable = Entregable.objects.get(id = producto['id'])

                evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable).order_by('-id')

                if evidencias_cargado.count() > 0:
                    ws.write( producto['letter'] + str(i) , 'SIC-' + str(evidencias_cargado[0].id), validado)


            i += 1

    wb.close()


    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"