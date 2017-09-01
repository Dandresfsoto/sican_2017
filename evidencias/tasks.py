#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from __future__ import absolute_import

from sican.celery import app
from evidencias.models import Red
import openpyxl
from sican.settings import base as settings
from StringIO import StringIO
from django.core.files import File
from matrices.models import Beneficiario
from evidencias.models import CargaMasiva
from productos.models import Entregable
from zipfile import ZipFile
from evidencias.models import Evidencia
from usuarios.models import User
import os
import shutil
from evidencias.models import Rechazo
from informes.models import InformesExcel
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

@app.task
def build_red(id_red):

    red = Red.objects.get(id = id_red)
    output = StringIO()

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ids = []
    inicia = 0

    if red.diplomado.numero == 1:
        ids = [{'id':8,'letter':'M'},
               {'id':9,'letter':'N'},
               {'id':20,'letter':'O'},
               {'id':12,'letter':'P'},
               {'id':21,'letter':'Q'},
               {'id':22,'letter':'R'},
               {'id':14,'letter':'S'},
               {'id':15,'letter':'T'},
               {'id':16,'letter':'U'},
               {'id':23,'letter':'V'},
               {'id':17,'letter':'W'},
               {'id':27,'letter':'X'},
               {'id':28,'letter':'Y'},
               {'id':40,'letter':'Z'},
               {'id':30,'letter':'AA'},
               {'id':31,'letter':'AB'},
               {'id':33,'letter':'AC'},
               {'id':34,'letter':'AD'},
               {'id':35,'letter':'AE'},
               {'id':36,'letter':'AF'},
               {'id':46,'letter':'AG'},
               {'id':58,'letter':'AH'},
               {'id':49,'letter':'AI'},
               {'id':59,'letter':'AJ'},
               {'id':52,'letter':'AK'},
               {'id':60,'letter':'AL'},
               {'id':55,'letter':'AM'},
               {'id':63,'letter':'AN'},
               {'id':64,'letter':'AO'},
               {'id':66,'letter':'AP'},
               {'id':67,'letter':'AQ'}]
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED INNOVATIC.xlsx')
        ws = wb.get_sheet_by_name('RED InnovaTIC')
        inicia = 6
    elif red.diplomado.numero == 2:
        ids = [{'id':72,'letter':'M'},
               {'id':73,'letter':'N'},
               {'id':75,'letter':'O'},
               {'id':74,'letter':'P'},
               {'id':76,'letter':'Q'},
               {'id':77,'letter':'R'},
               {'id':84,'letter':'S'},
               {'id':85,'letter':'T'},
               {'id':78,'letter':'U'},
               {'id':89,'letter':'V'},
               {'id':97,'letter':'W'},
               {'id':98,'letter':'X'},
               {'id':93,'letter':'Y'},
               {'id':92,'letter':'Z'},
               {'id':99,'letter':'AA'},
               {'id':94,'letter':'AB'},
               {'id':100,'letter':'AC'},
               {'id':95,'letter':'AD'},
               {'id':104,'letter':'AE'},
               {'id':112,'letter':'AF'},
               {'id':106,'letter':'AG'},
               {'id':109,'letter':'AH'},
               {'id':108,'letter':'AI'},
               {'id':110,'letter':'AJ'},
               {'id':119,'letter':'AK'},
               {'id':124,'letter':'AL'},
               {'id':118,'letter':'AM'},
               {'id':120,'letter':'AN'},
               {'id':121,'letter':'AO'}]
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED TECNOTIC.xlsx')
        ws = wb.get_sheet_by_name('RED TecnoTIC')
        inicia = 6
    elif red.diplomado.numero == 3:
        ids = [{'id':127,'letter':'M'},
               {'id':128,'letter':'N'},
               {'id':131,'letter':'O'},
               {'id':132,'letter':'P'},
               {'id':134,'letter':'Q'},
               {'id':133,'letter':'R'},
               {'id':142,'letter':'S'},
               {'id':143,'letter':'T'},
               {'id':135,'letter':'U'},
               {'id':144,'letter':'V'},
               {'id':137,'letter':'W'},
               {'id':140,'letter':'X'},
               {'id':139,'letter':'Y'},
               {'id':147,'letter':'Z'},
               {'id':146,'letter':'AA'},
               {'id':152,'letter':'AB'},
               {'id':148,'letter':'AC'},
               {'id':149,'letter':'AD'},
               {'id':151,'letter':'AE'},
               {'id':150,'letter':'AF'},
               {'id':156,'letter':'AG'},
               {'id':155,'letter':'AH'},
               {'id':157,'letter':'AI'},
               {'id':164,'letter':'AJ'},
               {'id':165,'letter':'AK'},
               {'id':159,'letter':'AL'},
               {'id':162,'letter':'AM'},
               {'id':161,'letter':'AN'},
               {'id':166,'letter':'AO'},
               {'id':167,'letter':'AP'},
               {'id':171,'letter':'AQ'},
               {'id':171,'letter':'AR'},
               {'id':169,'letter':'AS'}]
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED DIRECTIC.xlsx')
        ws = wb.get_sheet_by_name('RED DirecTIC')
        inicia = 6
    elif red.diplomado.numero == 4:
        ids = [{'id':221,'letter':'M'},
               {'id':221,'letter':'N'},
               {'id':221,'letter':'O'},
               {'id':224,'letter':'P'},
               {'id':228,'letter':'Q'}]

        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED FAMILIA.xlsx')
        ws = wb.get_sheet_by_name('RED Familia')
        inicia = 2

    elif red.diplomado.numero == 6:
        ids = [{'id':258,'letter':'M'},
               {'id':234,'letter':'N'},
               {'id':235,'letter':'O'},
               {'id':236,'letter':'P'},
               {'id':239,'letter':'Q'},
               {'id':242,'letter':'R'},
               {'id':244,'letter':'S'},
               {'id':247,'letter':'T'},
               {'id':248,'letter':'U'},
               {'id':255,'letter':'V'},
               {'id':256,'letter':'W'}]

        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED BOGOTA.xlsx')
        ws = wb.get_sheet_by_name('RED BOGOTA')
        inicia = 6

    evidencias_total = Evidencia.objects.filter(red_id = id_red)
    beneficiarios_id = evidencias_total.exclude(beneficiarios_cargados = None).values_list('beneficiarios_cargados__id',flat=True).distinct()


    i = 0 + inicia
    for beneficiario_id in beneficiarios_id:
        beneficiario = Beneficiario.objects.get(id = beneficiario_id)
        evidencias = evidencias_total.filter(beneficiarios_cargados__id = beneficiario_id)


        ws.cell('A'+str(i)).value = i - inicia + 1
        ws.cell('B'+str(i)).value = beneficiario.region.nombre.upper()
        ws.cell('C'+str(i)).value = beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text.upper()
        ws.cell('D'+str(i)).value = beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text.upper()
        ws.cell('E'+str(i)).value = beneficiario.ruta.upper() + '-' + beneficiario.grupo.nombre
        ws.cell('F'+str(i)).value = beneficiario.formador.get_full_name().upper()
        ws.cell('G'+str(i)).value = beneficiario.formador.cedula
        ws.cell('G'+str(i)).number_format = '0'
        ws.cell('H'+str(i)).value = beneficiario.nombres.upper()
        ws.cell('I'+str(i)).value = beneficiario.apellidos.upper()
        ws.cell('J'+str(i)).value = beneficiario.cedula

        ws.cell('J'+str(i)).number_format = '0'
        ws.cell('K'+str(i)).value = 'SICAN'
        ws.cell('L'+str(i)).value = 'I'

        for id in ids:
            evidencia = evidencias.filter(entregable__id = id['id'])
            if evidencia.count() == 1:
                if evidencia[0].subsanacion:
                    ws.cell('L'+str(i)).value = 'S'
                ws.cell( id['letter'] + str(i)).value = 'SIC-' + str(evidencia[0].id)
                ws.cell( id['letter'] + str(i)).hyperlink = 'https://sican.asoandes.org' + evidencia[0].get_archivo_url()

        i += 1


    wb.save(output)
    filename = 'RED-' + unicode(red.id) + '-'+ red.region.nombre +'.xlsx'
    red.archivo.save(filename,File(output))

    return "Generado RED-" + str(id_red)

@app.task
def carga_masiva_evidencias(id_carga_masiva,id_usuario):

    carga_masiva = CargaMasiva.objects.get(id = id_carga_masiva)
    output = StringIO()

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    wb = openpyxl.load_workbook(carga_masiva.excel.path)
    ws = wb.active

    zip = ZipFile(carga_masiva.zip.path,'r')

    i = 1

    evidencias_dict = {}

    for fila in ws.rows:
        resultado = ''

        try:
            beneficiario = Beneficiario.objects.get(cedula = fila[0].value)
        except:
            resultado = 'No existe el numero de cedula'
        else:

            try:
                entregable = Entregable.objects.get(id = fila[1].value)
            except:
                resultado = 'No existe la evidencia'
            else:

                try:
                    file_object = zip.getinfo(fila[2].value)
                except:
                    resultado = 'No existe el archivo'
                else:

                    evidencias = Evidencia.objects.filter(formador = beneficiario.formador,entregable = entregable)
                    reds = Red.objects.filter(evidencias__id__in = evidencias.values_list('id',flat=True))
                    exclude_validados = list(evidencias.exclude(beneficiarios_validados = None).values_list('beneficiarios_validados__id',flat=True))
                    exclude_enviados = []

                    for evidencia in evidencias.filter(id__in = reds.filter(retroalimentacion = False).values_list('evidencias__id',flat=True)):
                        for cargado in evidencia.beneficiarios_cargados.all():
                            exclude_enviados.append(cargado.id)

                    if True:
                    #if beneficiario.id not in exclude_validados:

                        #if beneficiario.id not in exclude_enviados:
                        if True:


                            source = zip.open(fila[2].value)

                            try:
                                filename = os.path.basename(fila[2].value)
                            except:
                                resultado = 'Error inesperado'

                            else:
                                resultado = 'Cargado correctamente'

                                target = file(os.path.join(r"C:\Temp",filename),"wb")

                                with source, target:
                                    shutil.copyfileobj(source,target)


                                if fila[2].value not in evidencias_dict.keys():
                                    evidencia_object = Evidencia.objects.create(usuario = User.objects.get(id = id_usuario),
                                                                                entregable = entregable, formador = beneficiario.formador)
                                    evidencia_object.archivo = File(open("C://Temp//" + filename, 'rb'))
                                    evidencia_object.save()
                                    evidencias_dict[fila[2].value] = evidencia_object
                                else:
                                    evidencia_object = evidencias_dict[fila[2].value]

                                os.remove("C://Temp//" + filename)

                                for evidencia in evidencias:
                                    if beneficiario in evidencia.beneficiarios_cargados.all():
                                        evidencia.beneficiarios_cargados.remove(beneficiario)

                                evidencia_object.beneficiarios_cargados.add(beneficiario)

                        else:
                            resultado = 'Se envio el entregable'
                    else:
                        resultado = 'El beneficiario ya tiene validado el entregable'




        ws_out['A'+str(i)] = fila[0].value
        ws_out['B'+str(i)] = fila[1].value
        ws_out['C'+str(i)] = fila[2].value
        ws_out['D'+str(i)] = resultado
        i += 1


    wb_out.save(output)
    filename = 'MASIVA-' + unicode(carga_masiva.id) + '.xlsx'
    carga_masiva.resultado.save(filename,File(output))

    return "Generada MASIVA-" + str(id_carga_masiva)

@app.task
def retroalimentacion_red(id_red):

    red = Red.objects.get(id = id_red)
    evidencias_red = Evidencia.objects.filter(red_id = id_red).values_list('id',flat=True)

    wb = openpyxl.load_workbook(red.archivo_retroalimentacion.file)
    ws = wb.get_active_sheet()
    ids = []
    inicia = 0

    if not red.producto_final:

        if red.diplomado.numero == 1:
            ids = [{'id':8,'letter':'M'},
                   {'id':9,'letter':'N'},
                   {'id':20,'letter':'O'},
                   {'id':12,'letter':'P'},
                   {'id':21,'letter':'Q'},
                   {'id':22,'letter':'R'},
                   {'id':14,'letter':'S'},
                   {'id':15,'letter':'T'},
                   {'id':16,'letter':'U'},
                   {'id':23,'letter':'V'},
                   {'id':17,'letter':'W'},
                   {'id':27,'letter':'X'},
                   {'id':28,'letter':'Y'},
                   {'id':40,'letter':'Z'},
                   {'id':30,'letter':'AA'},
                   {'id':31,'letter':'AB'},
                   {'id':33,'letter':'AC'},
                   {'id':34,'letter':'AD'},
                   {'id':35,'letter':'AE'},
                   {'id':36,'letter':'AF'},
                   {'id':46,'letter':'AG'},
                   {'id':58,'letter':'AH'},
                   {'id':49,'letter':'AI'},
                   {'id':59,'letter':'AJ'},
                   {'id':52,'letter':'AK'},
                   {'id':60,'letter':'AL'},
                   {'id':55,'letter':'AM'},
                   {'id':63,'letter':'AN'},
                   {'id':64,'letter':'AO'},
                   {'id':66,'letter':'AP'},
                   {'id':67,'letter':'AQ'}]

            inicia = 6
        elif red.diplomado.numero == 2:
            ids = [{'id':72,'letter':'M'},
                   {'id':73,'letter':'N'},
                   {'id':75,'letter':'O'},
                   {'id':74,'letter':'P'},
                   {'id':76,'letter':'Q'},
                   {'id':77,'letter':'R'},
                   {'id':84,'letter':'S'},
                   {'id':85,'letter':'T'},
                   {'id':78,'letter':'U'},
                   {'id':89,'letter':'V'},
                   {'id':97,'letter':'W'},
                   {'id':98,'letter':'X'},
                   {'id':93,'letter':'Y'},
                   {'id':92,'letter':'Z'},
                   {'id':99,'letter':'AA'},
                   {'id':94,'letter':'AB'},
                   {'id':100,'letter':'AC'},
                   {'id':95,'letter':'AD'},
                   {'id':104,'letter':'AE'},
                   {'id':112,'letter':'AF'},
                   {'id':106,'letter':'AG'},
                   {'id':109,'letter':'AH'},
                   {'id':108,'letter':'AI'},
                   {'id':110,'letter':'AJ'},
                   {'id':119,'letter':'AK'},
                   {'id':124,'letter':'AL'},
                   {'id':118,'letter':'AM'},
                   {'id':120,'letter':'AN'},
                   {'id':121,'letter':'AO'}]
            inicia = 6
        elif red.diplomado.numero == 3:
            ids = [{'id':127,'letter':'M'},
                   {'id':128,'letter':'N'},
                   {'id':131,'letter':'O'},
                   {'id':132,'letter':'P'},
                   {'id':134,'letter':'Q'},
                   {'id':133,'letter':'R'},
                   {'id':142,'letter':'S'},
                   {'id':143,'letter':'T'},
                   {'id':135,'letter':'U'},
                   {'id':144,'letter':'V'},
                   {'id':137,'letter':'W'},
                   {'id':140,'letter':'X'},
                   {'id':139,'letter':'Y'},
                   {'id':147,'letter':'Z'},
                   {'id':146,'letter':'AA'},
                   {'id':152,'letter':'AB'},
                   {'id':148,'letter':'AC'},
                   {'id':149,'letter':'AD'},
                   {'id':151,'letter':'AE'},
                   {'id':150,'letter':'AF'},
                   {'id':156,'letter':'AG'},
                   {'id':155,'letter':'AH'},
                   {'id':157,'letter':'AI'},
                   {'id':164,'letter':'AJ'},
                   {'id':165,'letter':'AK'},
                   {'id':159,'letter':'AL'},
                   {'id':162,'letter':'AM'},
                   {'id':161,'letter':'AN'},
                   {'id':166,'letter':'AO'},
                   {'id':167,'letter':'AP'},
                   {'id':171,'letter':'AQ'},
                   {'id':171,'letter':'AR'},
                   {'id':172,'letter':'AS'}]
            inicia = 6
        elif red.diplomado.numero == 4:
            ids = [{'id':221,'letter':'M'},
                   {'id':221,'letter':'N'},
                   {'id':221,'letter':'O'},
                   {'id':224,'letter':'P'},
                   {'id':228,'letter':'Q'}]

            inicia = 2


        for row in ws.iter_rows(row_offset=inicia-1):
            beneficiario = {}
            for cell in row:
                beneficiario[cell.column] = cell.value

            try:
                beneficiario_object = Beneficiario.objects.get(cedula = beneficiario['J'])

            except:
                pass

            else:
                for id in ids:
                    entregable = Entregable.objects.get(id = id['id'])

                    try:
                        evidencia = Evidencia.objects.filter(id__in = evidencias_red).get(entregable = entregable, formador = beneficiario_object.formador, beneficiarios_cargados = beneficiario_object)
                    except:
                        pass
                    else:

                        if beneficiario[id['letter']] == 'OK' or beneficiario[id['letter']] == 'Ok' or beneficiario[id['letter']] == 'oK' or beneficiario[id['letter']] == 'ok':
                            if beneficiario_object in evidencia.beneficiarios_cargados.all():
                                evidencia.beneficiarios_validados.add(beneficiario_object)

                        elif beneficiario[id['letter']] != '' and beneficiario[id['letter']] != None:
                            if beneficiario_object in evidencia.beneficiarios_cargados.all():
                                rechazo_object,created = Rechazo.objects.get_or_create(beneficiario_rechazo = beneficiario_object,
                                                                                       observacion = unicode(beneficiario[id['letter']]),
                                                                                       red_id = red.id,
                                                                                       evidencia_id=evidencia.id)
                                if created:
                                    evidencia.beneficiarios_rechazados.add(rechazo_object)


    red.retroalimentacion = True
    red.save()

    return "Retroalimentado RED-" + str(id_red)

@app.task
def build_consolidado_red(email):
    usuario = User.objects.get(email=email)
    nombre = "Consolidado RED"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    output = StringIO()

    ids_innovatic = [{'id':8,'letter':'M'},
               {'id':9,'letter':'N'},
               {'id':20,'letter':'O'},
               {'id':12,'letter':'P'},
               {'id':21,'letter':'Q'},
               {'id':22,'letter':'R'},
               {'id':14,'letter':'S'},
               {'id':15,'letter':'T'},
               {'id':16,'letter':'U'},
               {'id':23,'letter':'V'},
               {'id':17,'letter':'W'},
               {'id':27,'letter':'X'},
               {'id':28,'letter':'Y'},
               {'id':40,'letter':'Z'},
               {'id':30,'letter':'AA'},
               {'id':31,'letter':'AB'},
               {'id':33,'letter':'AC'},
               {'id':34,'letter':'AD'},
               {'id':35,'letter':'AE'},
               {'id':36,'letter':'AF'},
               {'id':46,'letter':'AG'},
               {'id':58,'letter':'AH'},
               {'id':49,'letter':'AI'},
               {'id':59,'letter':'AJ'},
               {'id':52,'letter':'AK'},
               {'id':60,'letter':'AL'},
               {'id':55,'letter':'AM'},
               {'id':63,'letter':'AN'},
               {'id':64,'letter':'AO'},
               {'id':66,'letter':'AP'},
               {'id':67,'letter':'AQ'}]

    ids_tecnotic = [{'id':72,'letter':'M'},
               {'id':73,'letter':'N'},
               {'id':75,'letter':'O'},
               {'id':74,'letter':'P'},
               {'id':76,'letter':'Q'},
               {'id':77,'letter':'R'},
               {'id':84,'letter':'S'},
               {'id':85,'letter':'T'},
               {'id':78,'letter':'U'},
               {'id':89,'letter':'V'},
               {'id':97,'letter':'W'},
               {'id':98,'letter':'X'},
               {'id':93,'letter':'Y'},
               {'id':92,'letter':'Z'},
               {'id':99,'letter':'AA'},
               {'id':94,'letter':'AB'},
               {'id':100,'letter':'AC'},
               {'id':95,'letter':'AD'},
               {'id':104,'letter':'AE'},
               {'id':112,'letter':'AF'},
               {'id':106,'letter':'AG'},
               {'id':109,'letter':'AH'},
               {'id':108,'letter':'AI'},
               {'id':110,'letter':'AJ'},
               {'id':119,'letter':'AK'},
               {'id':124,'letter':'AL'},
               {'id':118,'letter':'AM'},
               {'id':120,'letter':'AN'},
               {'id':121,'letter':'AO'}]

    ids_directic = [{'id':127,'letter':'M'},
               {'id':128,'letter':'N'},
               {'id':131,'letter':'O'},
               {'id':132,'letter':'P'},
               {'id':134,'letter':'Q'},
               {'id':133,'letter':'R'},
               {'id':142,'letter':'S'},
               {'id':143,'letter':'T'},
               {'id':135,'letter':'U'},
               {'id':144,'letter':'V'},
               {'id':137,'letter':'W'},
               {'id':140,'letter':'X'},
               {'id':139,'letter':'Y'},
               {'id':147,'letter':'Z'},
               {'id':146,'letter':'AA'},
               {'id':152,'letter':'AB'},
               {'id':148,'letter':'AC'},
               {'id':149,'letter':'AD'},
               {'id':151,'letter':'AE'},
               {'id':150,'letter':'AF'},
               {'id':156,'letter':'AG'},
               {'id':155,'letter':'AH'},
               {'id':157,'letter':'AI'},
               {'id':164,'letter':'AJ'},
               {'id':165,'letter':'AK'},
               {'id':159,'letter':'AL'},
               {'id':162,'letter':'AM'},
               {'id':161,'letter':'AN'},
               {'id':166,'letter':'AO'},
               {'id':167,'letter':'AP'},
               {'id':171,'letter':'AQ'},
               {'id':171,'letter':'AR'},
               {'id':169,'letter':'AS'}]

    ids_escuelatic = [{'id':221,'letter':'M'},
               {'id':221,'letter':'N'},
               {'id':221,'letter':'O'},
               {'id':224,'letter':'P'},
               {'id':228,'letter':'Q'}]

    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/CONSOLIDADO RED.xlsx')
    ws_innovatic = wb.get_sheet_by_name('RED InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('RED TecnoTIC')
    ws_directic = wb.get_sheet_by_name('RED DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('RED Familia')


    i_innovatic = 5
    i_tecnotic = 5
    i_directic = 5
    i_escuelatic = 5


    for red in Red.objects.all():

        if red.diplomado.numero == 1:
            ids = ids_innovatic
            i_innovatic += 1
            i = i_innovatic
            ws = ws_innovatic

        elif red.diplomado.numero == 2:
            ids = ids_tecnotic
            i_tecnotic += 1
            i = i_tecnotic
            ws = ws_tecnotic

        elif red.diplomado.numero == 3:
            ids = ids_directic
            i_directic += 1
            i = i_directic
            ws = ws_directic

        else:
            ids = ids_escuelatic
            i_escuelatic += 1
            i = i_escuelatic
            ws = ws_escuelatic

        ws.cell('A'+str(i)).value = 'RED-' + unicode(red.id)
        ws.cell('B'+str(i)).value = unicode(red.region.nombre)
        ws.cell('C'+str(i)).value = red.fecha

        for id in ids:
            evidencias = Evidencia.objects.filter(red_id = red.id).filter(entregable__id = id['id']).values_list('beneficiarios_cargados',flat=True).distinct()
            x = list(evidencias)
            if None in x:
                x.remove(None)
            ws.cell( id['letter'] + str(i)).value = len(x)
            ws.cell( id['letter'] + str(i)).style =Style(font=Font(name='Arial',
                                                                          size=11,
                                                                          bold=False,
                                                                          color='FF000000'
                                                                        ),
                                                                alignment=Alignment(
                                                                    horizontal='center',
                                                                    vertical='center',
                                                                    wrap_text=True
                                                                ),
                                                                number_format='0'
                                                                )


    wb.save(output)
    filename = 'CONSOLIDADOS_RED.xlsx'
    informe.archivo.save(filename,File(output))

    return "Generado consolidados RED"

@app.task
def build_consolidado_aprobacion_red(email):
    usuario = User.objects.get(email=email)
    nombre = "Consolidado aprobacion RED"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    output = StringIO()

    ids_innovatic = [{'id':8,'letter':'M'},
               {'id':9,'letter':'N'},
               {'id':20,'letter':'O'},
               {'id':12,'letter':'P'},
               {'id':21,'letter':'Q'},
               {'id':22,'letter':'R'},
               {'id':14,'letter':'S'},
               {'id':15,'letter':'T'},
               {'id':16,'letter':'U'},
               {'id':23,'letter':'V'},
               {'id':17,'letter':'W'},
               {'id':27,'letter':'X'},
               {'id':28,'letter':'Y'},
               {'id':40,'letter':'Z'},
               {'id':30,'letter':'AA'},
               {'id':31,'letter':'AB'},
               {'id':33,'letter':'AC'},
               {'id':34,'letter':'AD'},
               {'id':35,'letter':'AE'},
               {'id':36,'letter':'AF'},
               {'id':46,'letter':'AG'},
               {'id':58,'letter':'AH'},
               {'id':49,'letter':'AI'},
               {'id':59,'letter':'AJ'},
               {'id':52,'letter':'AK'},
               {'id':60,'letter':'AL'},
               {'id':55,'letter':'AM'},
               {'id':63,'letter':'AN'},
               {'id':64,'letter':'AO'},
               {'id':66,'letter':'AP'},
               {'id':67,'letter':'AQ'}]

    ids_tecnotic = [{'id':72,'letter':'M'},
               {'id':73,'letter':'N'},
               {'id':75,'letter':'O'},
               {'id':74,'letter':'P'},
               {'id':76,'letter':'Q'},
               {'id':77,'letter':'R'},
               {'id':84,'letter':'S'},
               {'id':85,'letter':'T'},
               {'id':78,'letter':'U'},
               {'id':89,'letter':'V'},
               {'id':97,'letter':'W'},
               {'id':98,'letter':'X'},
               {'id':93,'letter':'Y'},
               {'id':92,'letter':'Z'},
               {'id':99,'letter':'AA'},
               {'id':94,'letter':'AB'},
               {'id':100,'letter':'AC'},
               {'id':95,'letter':'AD'},
               {'id':104,'letter':'AE'},
               {'id':112,'letter':'AF'},
               {'id':106,'letter':'AG'},
               {'id':109,'letter':'AH'},
               {'id':108,'letter':'AI'},
               {'id':110,'letter':'AJ'},
               {'id':119,'letter':'AK'},
               {'id':124,'letter':'AL'},
               {'id':118,'letter':'AM'},
               {'id':120,'letter':'AN'},
               {'id':121,'letter':'AO'}]

    ids_directic = [{'id':127,'letter':'M'},
               {'id':128,'letter':'N'},
               {'id':131,'letter':'O'},
               {'id':132,'letter':'P'},
               {'id':134,'letter':'Q'},
               {'id':133,'letter':'R'},
               {'id':142,'letter':'S'},
               {'id':143,'letter':'T'},
               {'id':135,'letter':'U'},
               {'id':144,'letter':'V'},
               {'id':137,'letter':'W'},
               {'id':140,'letter':'X'},
               {'id':139,'letter':'Y'},
               {'id':147,'letter':'Z'},
               {'id':146,'letter':'AA'},
               {'id':152,'letter':'AB'},
               {'id':148,'letter':'AC'},
               {'id':149,'letter':'AD'},
               {'id':151,'letter':'AE'},
               {'id':150,'letter':'AF'},
               {'id':156,'letter':'AG'},
               {'id':155,'letter':'AH'},
               {'id':157,'letter':'AI'},
               {'id':164,'letter':'AJ'},
               {'id':165,'letter':'AK'},
               {'id':159,'letter':'AL'},
               {'id':162,'letter':'AM'},
               {'id':161,'letter':'AN'},
               {'id':166,'letter':'AO'},
               {'id':167,'letter':'AP'},
               {'id':171,'letter':'AQ'},
               {'id':171,'letter':'AR'},
               {'id':169,'letter':'AS'}]

    ids_escuelatic = [{'id':221,'letter':'M'},
               {'id':221,'letter':'N'},
               {'id':221,'letter':'O'},
               {'id':224,'letter':'P'},
               {'id':228,'letter':'Q'}]

    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/CONSOLIDADO RED.xlsx')
    ws_innovatic = wb.get_sheet_by_name('RED InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('RED TecnoTIC')
    ws_directic = wb.get_sheet_by_name('RED DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('RED Familia')


    i_innovatic = 5
    i_tecnotic = 5
    i_directic = 5
    i_escuelatic = 5


    for red in Red.objects.all():

        if red.diplomado.numero == 1:
            ids = ids_innovatic
            i_innovatic += 1
            i = i_innovatic
            ws = ws_innovatic

        elif red.diplomado.numero == 2:
            ids = ids_tecnotic
            i_tecnotic += 1
            i = i_tecnotic
            ws = ws_tecnotic

        elif red.diplomado.numero == 3:
            ids = ids_directic
            i_directic += 1
            i = i_directic
            ws = ws_directic

        else:
            ids = ids_escuelatic
            i_escuelatic += 1
            i = i_escuelatic
            ws = ws_escuelatic

        ws.cell('A'+str(i)).value = 'RED-' + unicode(red.id)
        ws.cell('B'+str(i)).value = unicode(red.region.nombre)
        ws.cell('C'+str(i)).value = red.fecha

        for id in ids:
            evidencias = Evidencia.objects.filter(red_id = red.id).filter(entregable__id = id['id']).values_list('beneficiarios_cargados',flat=True).distinct()
            x = list(evidencias)
            if None in x:
                x.remove(None)
            ws.cell( id['letter'] + str(i)).value = len(x)
            ws.cell( id['letter'] + str(i)).style =Style(font=Font(name='Arial',
                                                                          size=11,
                                                                          bold=False,
                                                                          color='FF000000'
                                                                        ),
                                                                alignment=Alignment(
                                                                    horizontal='center',
                                                                    vertical='center',
                                                                    wrap_text=True
                                                                ),
                                                                number_format='0'
                                                                )

    wb.save(output)
    filename = 'CONSOLIDADOS_APROBACION_RED.xlsx'
    informe.archivo.save(filename,File(output))

    return "Generado consolidados RED"



@app.task
def build_red_producto_final(id_red):

    red = Red.objects.get(id = id_red)
    output = StringIO()

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ids = []
    inicia = 0

    if red.diplomado.numero == 1:
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED PROYECTOS INNOVATIC.xlsx')
        ws = wb.get_sheet_by_name('Hoja1')
        i = 4

        for beneficiario in red.beneficiarios.all():

            ws.cell('A'+str(i)).value = i - 3
            ws.cell('B'+str(i)).value = beneficiario.region.nombre.upper()
            ws.cell('C'+str(i)).value = beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text.upper()
            ws.cell('D'+str(i)).value = beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text.upper()
            ws.cell('E'+str(i)).value = beneficiario.ie_text
            ws.cell('F'+str(i)).value = beneficiario.dane_ie_text
            ws.cell('G'+str(i)).value = beneficiario.ruta.upper() + '-' + beneficiario.grupo.nombre
            ws.cell('H'+str(i)).value = beneficiario.formador.get_full_name().upper()
            ws.cell('I'+str(i)).value = beneficiario.formador.cedula
            ws.cell('I'+str(i)).number_format = '0'
            ws.cell('J'+str(i)).value = beneficiario.nombres.upper()
            ws.cell('K'+str(i)).value = beneficiario.apellidos.upper()
            ws.cell('L'+str(i)).value = beneficiario.cedula

            ws.cell('L'+str(i)).number_format = '0'
            ws.cell('M'+str(i)).value = 'EXTERNO'
            ws.cell('N'+str(i)).value = 'I'

            ws.cell('O' + str(i)).value = beneficiario.nombre_producto_final
            ws.cell('P' + str(i)).value = beneficiario.area_basica_producto_final
            ws.cell('Q' + str(i)).value = 'N/A'
            ws.cell('R' + str(i)).value = beneficiario.link



            i += 1


    wb.save(output)
    filename = 'RED-' + unicode(red.id) + '-'+ red.region.nombre +'.xlsx'
    red.archivo.save(filename,File(output))

    return "Generado RED-" + str(id_red)