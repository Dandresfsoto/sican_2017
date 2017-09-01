#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from __future__ import absolute_import

from celery import shared_task
from formadores.models import CohortesFormadores
from mail_templated import send_mail
import openpyxl
from StringIO import StringIO
from usuarios.models import User
from cargos.models import Cargo
from formadores.models import Formador
from region.models import Region
from usuarios.tasks import send_mail_templated
import random
import string
from sican.settings.base import DEFAULT_FROM_EMAIL, API_KEY_SMS
from formadores.models import Contrato
from formadores.models import SolicitudSoportes
from zipfile import ZipFile
import os
import shutil
from django.core.files import File
from StringIO import StringIO
from django.contrib.auth.models import Group
import urllib, urllib2, json

@shared_task
def cohorte_formadores(id):
    cohorte = CohortesFormadores.objects.get(id = id)

    wb = openpyxl.load_workbook(cohorte.archivo.path)
    ws = wb.active

    output = StringIO()

    zip = ZipFile(cohorte.contratos.path,'r')

    for fila in ws.iter_rows(row_offset=2):

        email = fila[0].value

        if email != None:

            try:
                user = User.objects.get(email=email)
            except:
                first_name = fila[1].value
                last_name = fila[2].value
                fullname = first_name + last_name
                try:
                    cargo = Cargo.objects.get(id=fila[3].value)
                except:
                    cargo = None
                telefono_personal = fila[4].value
                correo_personal = fila[5].value


                if first_name != None and last_name != None and fullname != None and cargo!= None and telefono_personal != None and correo_personal != None:

                    user = User.objects.create_user(email=email,first_name=first_name,last_name=last_name,fullname=fullname,
                                               cargo=cargo,telefono_personal=telefono_personal,correo_personal=correo_personal)

                    password = "".join([random.choice(string.ascii_letters) for i in xrange(6)])
                    user.set_password(password)
                    user.save()

                    ws.cell(row=fila[0].row, column=14).value = 'Usuario creado'

                    send_mail_templated.delay('email/new_user.tpl',{'url_base' : 'https://sican.asoandes.org',
                                                                    'first_name': user.first_name,
                                                                    'last_name': user.last_name,'email': user.email,
                                                                    'password':password},DEFAULT_FROM_EMAIL,[user.email])
                else:
                    if cargo == None:
                        ws.cell(row=fila[0].row, column=14).value = 'Error: ID cargo'
                    else:
                        ws.cell(row=fila[0].row, column=14).value = 'Error: Campos vacios'

            else:
                ws.cell(row=fila[0].row, column=14).value = 'Warning: Usuario ya existe'

            user.groups.add(Group.objects.get(id=23))
            cedula = fila[7].value


            try:
                formador = Formador.objects.get(cedula = cedula)
            except:
                try:
                    region = Region.objects.get(id = fila[6].value)
                except:
                    region = None
                nombres = fila[1].value
                apellidos = fila[2].value
                correo_personal = fila[5].value
                celular_personal = fila[4].value
                cargo = Cargo.objects.get(id=fila[3].value)

                if region != None and nombres != None and apellidos != None and cedula != None and correo_personal != None and celular_personal != None and cargo != None:
                    formador = Formador.objects.create(usuario=user,nombres=nombres,apellidos=apellidos,
                                                       cedula=cedula,correo_personal=correo_personal,celular_personal=celular_personal)
                    formador.save()
                    formador.cargo.add(cargo)
                    formador.region.add(region)

                    ws.cell(row=fila[0].row, column=15).value = 'Formador creado'
                else:
                    if region == None:
                        ws.cell(row=fila[0].row, column=15).value = 'Error: ID region'
                    else:
                        ws.cell(row=fila[0].row, column=15).value = 'Error: Campos vacios'
            else:
                if formador.usuario != user:
                    formador.usuario = user
                    formador.save()
                    ws.cell(row=fila[0].row, column=15).value = 'Warning: Usuario actualizado'
                else:
                    ws.cell(row=fila[0].row, column=15).value = 'Formador ya existe'


            codigo = fila[8].value
            try:
                soportes_requeridos = SolicitudSoportes.objects.get(id = fila[9].value)
            except:
                soportes_requeridos = None
            fecha_inicio = fila[10].value
            fecha_fin = fila[11].value

            if codigo != None and soportes_requeridos != None and formador != None:

                try:
                    contrato = Contrato.objects.create(nombre=codigo,formador=formador,soportes_requeridos=soportes_requeridos,
                                                       fecha_inicio=fecha_inicio,fecha_fin=fecha_fin)
                except:
                    ws.cell(row=fila[0].row, column=16).value = 'Error: Creacion de contrato, posiblemente existe otro contrato con el mismo nombre'

                else:

                    source = zip.open(fila[12].value)

                    try:
                        filename = os.path.basename(fila[12].value)
                    except:
                        ws.cell(row=fila[0].row, column=16).value = 'Error: Path del contrato'

                    else:

                        target = file(os.path.join(r"C:\Temp",filename),"wb")

                        with source, target:
                            shutil.copyfileobj(source,target)

                        contrato.contrato_original = File(open("C://Temp//" + filename, 'rb'))
                        contrato.save()
                        ws.cell(row=fila[0].row, column=16).value = 'Contrato creado y soporte cargado'
                        send_mail_templated.delay('email/contrato_formador.tpl',{'url_base' : 'https://sican.asoandes.org',
                                                                    'first_name': user.first_name,
                                                                    'last_name': user.last_name,'email': user.email,
                                                                    'contrato':contrato.nombre},DEFAULT_FROM_EMAIL,[user.email])
                        if user.telefono_personal != None and len(unicode(user.telefono_personal)) == 10:
                            mensaje = 'ASOANDES: Se agrego el contrato '+contrato.nombre+' a tu cuenta en el sistema SICAN,' \
                                                                               ' procede a legalizarlo, dudas con recursohumano@asoandes.org'
                            parametros = urllib.urlencode({'apikey':API_KEY_SMS,'mensaje': mensaje,'numcelular': unicode(user.telefono_personal),'numregion':'57'})
                            headers = {"Content-type": "application/x-www-form-urlencoded", "Accept":"text/plain"}
                            request = urllib2.Request('http://panel.smasivos.com/api.envio.new.php', parametros, headers)
                            opener = urllib2.build_opener()
                            respuesta = opener.open(request).read()
                            j=json.loads(respuesta)
                            ws.cell(row=fila[0].row, column=17).value = unicode(j)

            else:
                ws.cell(row=fila[0].row, column=16).value = 'Error: Campos vacios'


    wb.save(output)
    filename = 'COHORTE-' + unicode(cohorte.id) + '.xlsx'
    cohorte.resultado.save(filename,File(output))


    return 'Cohorte procesado, '