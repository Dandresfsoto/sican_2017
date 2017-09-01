#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from __future__ import absolute_import

from sican.celery import app
from informes.functions import construir_reporte, cronograma_interventoria
from informes.models import InformesExcel
from django.core.files import File
from usuarios.models import User
from formadores.models import Formador, SolicitudTransporte
from formadores.models import Soporte as SoporteFormadores
from rh.models import TipoSoporte
from preinscripcion.models import DocentesPreinscritos
from formacion.models import EntradaCronograma
from isoweek import Week
import datetime
from formacion.models import Semana
from lideres.models import Lideres, Soporte
from encuestas.models import PercepcionInicial, PercepcionFinal
from radicados.models import Radicado
from formadores.models import Cortes
from formadores.models import Revision
import zipfile
import shutil
import os
from rh.models import RequerimientoPersonal
from productos.models import Contratos, ValorEntregable
from StringIO import StringIO
import openpyxl
from sican.settings import base as settings
from productos.models import Diplomado
from evidencias.models import Beneficiario
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from evidencias.models import Evidencia
from productos.models import Entregable
from evidencias.models import Red
from openpyxl.comments import Comment
from django.db.models import Q
from region.models import Region
import xlsxwriter
from formadores.models import Contrato as ContratoVigencia2017
from vigencia2017.models import Grupos as GruposVigencia2017
from vigencia2017.models import Beneficiario as BeneficiarioVigencia2017
from vigencia2017.models import Evidencia as EvidenciaVigencia2017

@app.task
def nueva_semana():
    x, created = Semana.objects.get_or_create(numero = datetime.datetime.now().isocalendar()[1]+1)
    return "Semana actualizada"

@app.task
def formadores(email):
    usuario = User.objects.get(email=email)
    nombre = "Directorio de formadores"
    proceso = "RH-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']

    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]

    contenidos = []

    for formador in Formador.objects.exclude(oculto=True):
        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.get_cargo_string() if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def reporte_requerimientos_contratacion(email):
    usuario = User.objects.get(email=email)
    nombre = "Requerimientos de contratacion"
    proceso = "RH-INF05"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Fecha Solicitud','Solicitante','Departamento','Municipios','Codigo Ruta','Encargado','Observación Solicitante',
               'Fecha Respuesta','Nombre','Cedula','Celular','Email','Hoja de vida','Observación respuesta','Fecha solicitud de contratación',
               'Observacion final','Estado']

    formatos = ['General','d/m/yy','General','General','General','General','General','General',
                'd/m/yy','General','0','General','General','General','General','d/m/yy',
                'General','General']


    ancho_columnas =  [30,15,20,15,30,15,20,30,
                       15,20,15,15,20,20,30,20,
                       30,30]

    contenidos = []

    for requerimiento in RequerimientoPersonal.objects.all():

        estado = ''

        if requerimiento.remitido_respuesta == False and requerimiento.remitido_contratacion == False and requerimiento.contratar == False and requerimiento.desierto == False and requerimiento.contratado == False and requerimiento.contrato_enviado == False:
            estado = 'Remitido a RH'

        elif requerimiento.remitido_respuesta == True and requerimiento.remitido_contratacion == False and requerimiento.contratar == False and requerimiento.desierto == False and requerimiento.contratado == False and requerimiento.contrato_enviado == False:
            estado = 'Listo para capacitar'

        elif requerimiento.remitido_respuesta == True and requerimiento.remitido_contratacion == True and requerimiento.contratar == True and requerimiento.desierto == False and requerimiento.contratado == False and requerimiento.contrato_enviado == False:
            estado = 'Proceder a contrato'

        elif requerimiento.remitido_respuesta == True and requerimiento.remitido_contratacion == True and requerimiento.contratar == False and requerimiento.desierto == True and requerimiento.contratado == False and requerimiento.contrato_enviado == False:
            estado = 'Aspirante deserta'

        elif requerimiento.contratado == False and requerimiento.contrato_enviado == True:
            estado = 'Contrato enviado'

        elif requerimiento.contratado == True:
            estado = 'Contratado'



        contenidos.append([
            'REQ-'+unicode(requerimiento.id),
            requerimiento.fecha_solicitud,
            requerimiento.solicitante.get_full_name_string(),
            requerimiento.departamento.nombre,
            requerimiento.get_municipios_string(),
            requerimiento.codigo_ruta,
            requerimiento.encargado.get_full_name_string(),
            requerimiento.observacion_solicitante,
            requerimiento.fecha_respuesta,
            requerimiento.nombre,
            requerimiento.cedula,
            requerimiento.celular,
            requerimiento.email,
            'Si' if str(requerimiento.hv) != '' else 'No',
            requerimiento.observacion_respuesta,
            requerimiento.fecha_solicitud_contratacion,
            requerimiento.observacion_final,
            estado
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def formadores_soportes(email):
    usuario = User.objects.get(email=email)
    nombre = "Soportes cargados por formador"

    proceso = "RH-INF02"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Nombres','Apellidos','Cedula','Región']

    formatos = ['General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15]

    contenidos = []

    tipos_soportes = TipoSoporte.objects.exclude(oculto=True).values_list('id','nombre')

    for tipo_soporte in tipos_soportes:
        titulos.append(tipo_soporte[1])
        formatos.append('General')
        ancho_columnas.append(30)

    for formador in Formador.objects.exclude(oculto=True):
        row =[
                'FOR-'+unicode(formador.id),
                formador.nombres,
                formador.apellidos,
                formador.cedula,
                formador.get_region_string(),
            ]
        for tipo_soporte in tipos_soportes:
            try:
                soporte = SoporteFormadores.objects.filter(formador=formador).get(tipo__id=tipo_soporte[0])
            except:
                row.append('No')
            else:
                if soporte.get_archivo_url() != "":
                    row.append('Si')
                else:
                    row.append('No')

        contenidos.append(row)

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def preinscritos(email):
    usuario = User.objects.get(email=email)
    nombre = "Reporte de docentes preinscritos"
    proceso = "FOR-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Cedula','Primer apellido','Segundo apellido','Primer nombre','Segundo nombre','Cargo','Correo',
               'Telefono fijo','Celular','Departamento','Municipio','Radicado','Base Mineducación','Fecha registro',
               'Secretaria Radicado','Municipio Radicado','Nombre Sede','Dane Sede','Tipo Sede','Ubicación']

    formatos = ['General','0','General','General','General','General','General','General',
                'General','General','General','General','0','General','d/m/yy',
                'General','General','General','General','General','General']


    ancho_columnas =  [30,20,15,15,15,15,15,30,
                       15,15,20,20,15,10,15,
                       20,20,20,20,20,20]

    contenidos = []

    for docente in DocentesPreinscritos.objects.all():
        contenidos.append([
            'PRE-'+unicode(docente.id),
            docente.cedula,
            docente.primer_apellido,
            docente.segundo_apellido,
            docente.primer_nombre,
            docente.segundo_nombre,
            docente.cargo,
            docente.correo,
            docente.telefono_fijo,
            docente.telefono_celular,
            docente.departamento.nombre if docente.departamento != None else '',
            docente.municipio.nombre if docente.municipio != None else '',
            docente.radicado.numero if docente.radicado != None else '',
            'Si' if docente.verificado else 'No',
            docente.fecha,
            docente.radicado.secretaria.nombre,
            docente.radicado.municipio.nombre,
            docente.radicado.nombre_sede,
            docente.radicado.dane_sede,
            docente.radicado.tipo,
            docente.radicado.ubicacion,
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def transportes(email):
    usuario = User.objects.get(email=email)
    nombre = "Reporte de solicitudes de desplazamiento"
    proceso = "FIN-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Formador','Cedula','Región','Cargo','Banco','Tipo de cuenta','Numero de cuenta',
               'Nombre solicitud','Fecha','Estado','Valor solicitado','Valor Aprobado','Observación']

    formatos = ['General','General','General','General','General','General','General','General',
                'General','d/m/yy h:mm:ss AM/PM','General','"$"#,##0_);("$"#,##0)','"$"#,##0_);("$"#,##0)','General']


    ancho_columnas =  [30,30,15,15,15,15,15,20,
                       50,15,20,15,15,60]

    contenidos = []

    for solicitud in SolicitudTransporte.objects.all():
        contenidos.append([
            'PRE-'+unicode(solicitud.id),
            solicitud.formador.get_full_name(),
            solicitud.formador.cedula,
            solicitud.formador.get_region_string(),
            solicitud.formador.get_cargo_string() if solicitud.formador.cargo != None else "",
            solicitud.formador.banco.nombre if solicitud.formador.banco != None else "",
            solicitud.formador.tipo_cuenta,
            solicitud.formador.numero_cuenta,
            solicitud.nombre,
            solicitud.creacion,
            solicitud.estado,
            solicitud.valor,
            solicitud.valor_aprobado,
            solicitud.observacion
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def cronograma_general(email,semana_id):
    usuario = User.objects.get(email=email)
    nombre = "Cronograma consolidado general"

    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    semana = Semana.objects.get(id=semana_id)

    innovatics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 1").order_by('formador__region').values_list('id',flat=True)
    tecnotics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 2").order_by('formador__region').values_list('id',flat=True)
    directics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 3").order_by('formador__region').values_list('id',flat=True)
    escuelatics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 4").order_by('formador__region').values_list('id',flat=True)

    inicio = Week(semana.creacion.isocalendar()[0],semana.creacion.isocalendar()[1]+1).monday()
    fin = Week(semana.creacion.isocalendar()[0],semana.creacion.isocalendar()[1]+1).sunday()
    rango = inicio.strftime("Del dia %d de %B del %Y") + ' - ' + fin.strftime(" al dia %d de %B del %Y")

    output = cronograma_interventoria(innovatics,tecnotics,directics,escuelatics,rango)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def cronograma_lider(email,semana_id):
    usuario = User.objects.get(email=email)
    nombre = "Cronograma semanal lider"

    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    semana = Semana.objects.get(id=semana_id)

    innovatics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 1",formador__lider__email = email).values_list('id',flat=True)
    tecnotics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 2",formador__lider__email = email).values_list('id',flat=True)
    directics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 3",formador__lider__email = email).values_list('id',flat=True)
    escuelatics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 4",formador__lider__email = email).values_list('id',flat=True)

    inicio = Week(semana.creacion.isocalendar()[0],semana.creacion.isocalendar()[1]+1).monday()
    fin = Week(semana.creacion.isocalendar()[0],semana.creacion.isocalendar()[1]+1).sunday()
    rango = inicio.strftime("Del dia %d de %B del %Y") + ' - ' + fin.strftime(" al dia %d de %B del %Y")

    output = cronograma_interventoria(innovatics,tecnotics,directics,escuelatics,rango)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def lideres(email):
    usuario = User.objects.get(email=email)
    nombre = "Directorio de lideres"
    proceso = "RH-INF03"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']

    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]

    contenidos = []

    for lider in Lideres.objects.exclude(oculto=True):
        contenidos.append([
            'LID-'+unicode(lider.id),
            lider.nombres,
            lider.apellidos,
            lider.cedula,
            lider.region.nombre,
            lider.correo_personal,
            lider.celular_personal,
            lider.cargo.nombre if lider.cargo != None else '',
            lider.profesion,
            lider.fecha_contratacion,
            lider.fecha_terminacion,
            lider.banco.nombre if lider.banco != None else '',
            lider.tipo_cuenta,
            lider.numero_cuenta,
            lider.eps,
            lider.pension,
            lider.arl
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def lideres_soportes(email):
    usuario = User.objects.get(email=email)
    nombre = "Soportes cargados por lider"
    proceso = "RH-INF04"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Nombres','Apellidos','Cedula','Región']

    formatos = ['General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15]

    contenidos = []

    tipos_soportes = TipoSoporte.objects.exclude(oculto=True).values_list('id','nombre')

    for tipo_soporte in tipos_soportes:
        titulos.append(tipo_soporte[1])
        formatos.append('General')
        ancho_columnas.append(30)

    for lider in Lideres.objects.exclude(oculto=True):
        row =[
                'LID-'+unicode(lider.id),
                lider.nombres,
                lider.apellidos,
                lider.cedula,
                lider.region.nombre,
            ]
        for tipo_soporte in tipos_soportes:
            try:
                soporte = Soporte.objects.filter(lider=lider).get(tipo__id=tipo_soporte[0])
            except:
                row.append('No')
            else:
                if soporte.get_archivo_url() != "":
                    row.append('Si')
                else:
                    row.append('No')

        contenidos.append(row)

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def encuesta_percepcion_inicial(email):
    usuario = User.objects.get(email=email)
    nombre = "Base de datos respuestas percepción inicial"
    proceso = "ENC-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Cedula','Primer Apellido','Segundo Apellido','Primer Nombre','Segundo Nombre','Cargo','Correo','Telefono Fijo','Telefono Celular',
               'Departamento','Municipio','Verificado','Fecha','Radicado','Secretaria','Municipio Radicado','Nombre Sede','Dane Sede','Ubicación',
               'Area','Area Otras','Antiguedad','Pregunta 1','Pregunta 1 Opcional','Pregunta 2','Pregunta 3','Pregunta 4','Pregunta 5','Pregunta 6',
               'Pregunta 6 Opcional','Pregunta 7','Pregunta 8','Pregunta 9','Pregunta 10','Pregunta 11','Pregunta 12','Pregunta 12 Opcional','Pregunta 13']

    formatos = ['General','0','General','General','General','General','General','General','General','General',
                'General','General','General','d/m/yy','0','General','General','General','0','General',
                'General','General','General','General','General','General','General','General','General','General',
                'General','General','General','General','General','General','General','General','General',]


    ancho_columnas =  [30,20,15,15,15,40,40,40,40,40,
                       40,40,40,40,40,40,40,40,40,40,
                       60,60,60,60,60,60,60,60,60,60,
                       60,60,60,60,60,60,60,60,60]

    contenidos = []

    for encuestado in PercepcionInicial.objects.all():
        contenidos.append([
            'ENC-'+unicode(encuestado.id),
            encuestado.docente_preinscrito.cedula,
            encuestado.docente_preinscrito.primer_apellido,
            encuestado.docente_preinscrito.segundo_apellido,
            encuestado.docente_preinscrito.primer_nombre,
            encuestado.docente_preinscrito.segundo_nombre,
            encuestado.docente_preinscrito.cargo,
            encuestado.docente_preinscrito.correo,
            encuestado.docente_preinscrito.telefono_fijo,
            encuestado.docente_preinscrito.telefono_celular,
            encuestado.docente_preinscrito.departamento.nombre if encuestado.docente_preinscrito.departamento != None else '',
            encuestado.docente_preinscrito.municipio.nombre if encuestado.docente_preinscrito.municipio != None else '',
            'SI' if encuestado.docente_preinscrito.verificado else 'NO',
            encuestado.docente_preinscrito.fecha,
            encuestado.docente_preinscrito.radicado.numero if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.secretaria.nombre if encuestado.docente_preinscrito.radicado.secretaria != None else '',
            encuestado.docente_preinscrito.radicado.municipio.nombre if encuestado.docente_preinscrito.radicado.municipio != None else '',
            encuestado.docente_preinscrito.radicado.nombre_sede if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.dane_sede if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.ubicacion if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.area,
            encuestado.area_1,
            encuestado.antiguedad,
            encuestado.pregunta_1,
            encuestado.pregunta_1_1,
            encuestado.pregunta_2,
            encuestado.pregunta_3,
            encuestado.pregunta_4,
            encuestado.pregunta_5,
            encuestado.pregunta_6,
            encuestado.pregunta_6_1,
            encuestado.pregunta_7,
            encuestado.pregunta_8,
            encuestado.pregunta_9,
            encuestado.pregunta_10,
            encuestado.pregunta_11,
            encuestado.pregunta_12,
            encuestado.pregunta_12_1,
            encuestado.pregunta_13,
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def encuesta_percepcion_final(email):
    usuario = User.objects.get(email=email)
    nombre = "Base de datos respuestas percepción final"
    proceso = "ENC-INF02"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Cedula','Primer Apellido','Segundo Apellido','Primer Nombre','Segundo Nombre','Cargo','Correo','Telefono Fijo','Telefono Celular',
               'Departamento','Municipio','Verificado','Fecha','Radicado','Secretaria','Municipio Radicado','Nombre Sede','Dane Sede','Ubicación',
               'Area','Tiempo Formación','Pregunta 1','Pregunta 2','Pregunta 3','Pregunta 4','Pregunta 5','Pregunta 6','Pregunta 7','Pregunta 8',
               'Pregunta 9','Pregunta 10','Pregunta 11','Pregunta 12']

    formatos = ['General','0','General','General','General','General','General','General','General','General',
                'General','General','General','d/m/yy','0','General','General','General','0','General',
                'General','General','General','General','General','General','General','General','General','General',
                'General','General','General','General']


    ancho_columnas =  [30,20,15,15,15,40,40,40,40,40,
                       40,40,40,40,40,40,40,40,40,40,
                       60,60,60,60,60,60,60,60,60,60,
                       60,60,60,60]

    contenidos = []

    for encuestado in PercepcionFinal.objects.all():
        contenidos.append([
            'ENC-'+unicode(encuestado.id),
            encuestado.docente_preinscrito.cedula,
            encuestado.docente_preinscrito.primer_apellido,
            encuestado.docente_preinscrito.segundo_apellido,
            encuestado.docente_preinscrito.primer_nombre,
            encuestado.docente_preinscrito.segundo_nombre,
            encuestado.docente_preinscrito.cargo,
            encuestado.docente_preinscrito.correo,
            encuestado.docente_preinscrito.telefono_fijo,
            encuestado.docente_preinscrito.telefono_celular,
            encuestado.docente_preinscrito.departamento.nombre if encuestado.docente_preinscrito.departamento != None else '',
            encuestado.docente_preinscrito.municipio.nombre if encuestado.docente_preinscrito.municipio != None else '',
            'SI' if encuestado.docente_preinscrito.verificado else 'NO',
            encuestado.docente_preinscrito.fecha,
            encuestado.docente_preinscrito.radicado.numero if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.secretaria.nombre if encuestado.docente_preinscrito.radicado.secretaria != None else '',
            encuestado.docente_preinscrito.radicado.municipio.nombre if encuestado.docente_preinscrito.radicado.municipio != None else '',
            encuestado.docente_preinscrito.radicado.nombre_sede if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.dane_sede if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.ubicacion if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.area,
            encuestado.tiempo_formacion,
            encuestado.pregunta_1,
            encuestado.pregunta_2,
            encuestado.pregunta_3,
            encuestado.pregunta_4,
            encuestado.pregunta_5,
            encuestado.pregunta_6,
            encuestado.pregunta_7,
            encuestado.pregunta_8,
            encuestado.pregunta_9,
            encuestado.pregunta_10,
            encuestado.pregunta_11,
            encuestado.pregunta_12,
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def radicados(email):
    usuario = User.objects.get(email=email)
    nombre = "Base de datos radicados"
    proceso = "DB-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Numero','Municipio','Departamento','Secretaria','Nombre Sede','Dane Sede','Tipo','Ubicación']

    formatos = ['General','General','General','General','General','General','General','General','General']


    ancho_columnas =  [30,30,30,30,30,30,30,30,30]

    contenidos = []

    for radicado in Radicado.objects.exclude(oculto=True):
        contenidos.append([
            'RAD-'+unicode(radicado.id),
            radicado.numero,
            radicado.municipio.nombre if radicado.municipio != None else '',
            radicado.municipio.departamento.nombre if radicado.municipio.departamento != None else '',
            radicado.secretaria.nombre if radicado.secretaria != None else '',
            radicado.nombre_sede,
            radicado.dane_sede,
            radicado.tipo,
            radicado.ubicacion
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def pagos_mensual(email):
    usuario = User.objects.get(email=email)
    nombre = "Reporte mensual de pago a formadores"
    proceso = "COR-INF02"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    cortes = Cortes.objects.all()
    fechas = []

    for corte in cortes:
        fecha_corte = '01/' + str(corte.mes) + '/' + str(corte.year)
        fecha_corte = datetime.datetime.strptime(fecha_corte,"%d/%m/%Y").date()
        fechas.append(fecha_corte)

    fechas = list(set(fechas))


    titulos = ['ID','Nombres','Apellidos','Cedula','Región',
               'Correo','Celular','Cargo','Banco','Tipo cuenta','Numero cuenta']


    formatos = ['General','General','General','General','General',
               'General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,
                       40,20,20,20,20,20]


    for fecha_corte in fechas:
        titulos.append(fecha_corte.strftime("%B de %Y"))
        formatos.append('"$"#,##0_);("$"#,##0)')
        ancho_columnas.append(30)

    contenidos = []



    for formador in Formador.objects.all():
        row =[
                'FOR-'+unicode(formador.id),
                formador.nombres,
                formador.apellidos,
                formador.cedula,
                formador.get_region_string(),
                formador.correo_personal,
                formador.celular_personal,
                formador.get_cargo_string() if formador.cargo != None else '',
                formador.banco.nombre if formador.banco != None else '',
                formador.tipo_cuenta,
                formador.numero_cuenta,
            ]
        for fecha_corte in fechas:
            valor = 0
            for corte in Cortes.objects.filter(mes = str(fecha_corte.month).zfill(2), year = fecha_corte.year):
                revisiones = Revision.objects.filter(corte = corte, formador_revision = formador)
                for revision in revisiones:
                    for producto in revision.productos.all():
                        valor += producto.cantidad * producto.valor_entregable.valor

            row.append(valor)

        contenidos.append(row)

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def zip_hv(email):

    if os.path.exists("C:\\Temp\\hv.zip"):
        os.remove("C:\\Temp\\hv.zip")

    usuario = User.objects.get(email=email)
    nombre = "Zip: Hojas de vida"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")

    zip = zipfile.ZipFile('C:\\Temp\\hv.zip',"w",allowZip64=True)

    for soporte in SoporteFormadores.objects.filter(tipo__id = 3).exclude(oculto = True):
        if str(soporte.archivo) != '':
            if os.path.exists(soporte.archivo.path):
                zip.write(soporte.archivo.path,soporte.formador.get_full_name()+'/'+os.path.basename(soporte.archivo.path))

    zip.close()
    informe.archivo = File(open('C:\\Temp\\hv.zip'))
    informe.save()

    shutil.copy('C:\\Temp\\hv.zip',informe.archivo.path)

    return "Zip creado HV"

@app.task
def zip_contrato(email):

    if os.path.exists("C:\\Temp\\contratos.zip"):
        os.remove("C:\\Temp\\contratos.zip")

    usuario = User.objects.get(email=email)
    nombre = "Zip: Contratos"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")


    zip = zipfile.ZipFile('C:\\Temp\\contratos.zip',"w",allowZip64=True)

    for soporte in SoporteFormadores.objects.filter(tipo__id = 10).exclude(oculto = True):
        if str(soporte.archivo) != '':
            if os.path.exists(soporte.archivo.path):
                zip.write(soporte.archivo.path,soporte.formador.get_full_name()+'/'+os.path.basename(soporte.archivo.path))

    zip.close()
    informe.archivo = File(open('C:\\Temp\\contratos.zip'))
    informe.save()

    shutil.copy('C:\\Temp\\contratos.zip',informe.archivo.path)

    return "Zip creado Contrato"

@app.task
def acumulado_tipo_1(email):
    usuario = User.objects.get(email=email)
    nombre = "Acumulado formadores tipo 1"
    proceso = "REV-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    contrato_1 = Contratos.objects.filter(cargo__nombre="Formador Tipo 1").get(nombre = 'Capacitación 1')
    entregables_1 = ValorEntregable.objects.filter(contrato = contrato_1)

    contrato_2 = Contratos.objects.filter(cargo__nombre="Formador Tipo 1").get(nombre = 'Capacitación 2')
    entregables_2 = ValorEntregable.objects.filter(contrato = contrato_2)


    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']



    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]


    if list(entregables_1.values_list('entregable__id',flat=True)) == list(entregables_2.values_list('entregable__id',flat=True)):
        for entregable in entregables_1:
            titulos.append(entregable.entregable.sesion.nivel.nombre + ' - ' +
                           entregable.entregable.sesion.nombre + ' - ID:' +
                           str(entregable.entregable.id))
            formatos.append('General')
            ancho_columnas.append(30)
    else:
        raise ValueError('Los contratos no tienen los mismos entregables')


    titulos.append('Valor pagado')
    formatos.append('"$"#,##0_);("$"#,##0)')
    ancho_columnas.append(30)


    contenidos = []

    for formador in Formador.objects.filter(cargo__nombre="Formador Tipo 1").exclude(oculto=True):

        cantidad_list = []
        valor = 0
        for entregable in entregables_1:
            productos = Revision.objects.filter(formador_revision = formador).filter(productos__valor_entregable__entregable__id = entregable.entregable.id)
            cantidad = productos.values_list('productos__cantidad',flat=True)
            cantidad_list.append(sum(cantidad))
            valor_entregable = 0

            for x in productos.values_list('productos__cantidad','productos__valor_entregable__valor'):
                valor_entregable += x[0]*x[1]

            valor += valor_entregable

        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.get_cargo_string() if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ] + cantidad_list + [valor])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def acumulado_tipo_2(email):
    usuario = User.objects.get(email=email)
    nombre = "Acumulado formadores tipo 2"
    proceso = "REV-INF02"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    contrato_1 = Contratos.objects.filter(cargo__nombre="Formador Tipo 2").get(nombre = 'Capacitación 1')
    entregables_1 = ValorEntregable.objects.filter(contrato = contrato_1)

    contrato_2 = Contratos.objects.filter(cargo__nombre="Formador Tipo 2").get(nombre = 'Capacitación 2')
    entregables_2 = ValorEntregable.objects.filter(contrato = contrato_2)


    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']



    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]


    if list(entregables_1.values_list('entregable__id',flat=True)) == list(entregables_2.values_list('entregable__id',flat=True)):
        for entregable in entregables_1:
            titulos.append(entregable.entregable.sesion.nivel.nombre + ' - ' +
                           entregable.entregable.sesion.nombre + ' - ID:' +
                           str(entregable.entregable.id))
            formatos.append('General')
            ancho_columnas.append(30)
    else:
        raise ValueError('Los contratos no tienen los mismos entregables')


    titulos.append('Valor pagado')
    formatos.append('"$"#,##0_);("$"#,##0)')
    ancho_columnas.append(30)


    contenidos = []

    for formador in Formador.objects.filter(cargo__nombre="Formador Tipo 2").exclude(oculto=True):

        cantidad_list = []
        valor = 0
        for entregable in entregables_1:
            productos = Revision.objects.filter(formador_revision = formador).filter(productos__valor_entregable__entregable__id = entregable.entregable.id)
            cantidad = productos.values_list('productos__cantidad',flat=True)
            cantidad_list.append(sum(cantidad))
            valor_entregable = 0

            for x in productos.values_list('productos__cantidad','productos__valor_entregable__valor'):
                valor_entregable += x[0]*x[1]

            valor += valor_entregable

        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.get_cargo_string() if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ] + cantidad_list + [valor])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def acumulado_tipo_3(email):
    usuario = User.objects.get(email=email)
    nombre = "Acumulado formadores tipo 3"
    proceso = "REV-INF03"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    contrato_1 = Contratos.objects.filter(cargo__nombre="Formador Tipo 3").get(nombre = 'Capacitación 1')
    entregables_1 = ValorEntregable.objects.filter(contrato = contrato_1)

    contrato_2 = Contratos.objects.filter(cargo__nombre="Formador Tipo 3").get(nombre = 'Capacitación 2')
    entregables_2 = ValorEntregable.objects.filter(contrato = contrato_2)


    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']



    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]


    if list(entregables_1.values_list('entregable__id',flat=True)) == list(entregables_2.values_list('entregable__id',flat=True)):
        for entregable in entregables_1:
            titulos.append(entregable.entregable.sesion.nivel.nombre + ' - ' +
                           entregable.entregable.sesion.nombre + ' - ID:' +
                           str(entregable.entregable.id))
            formatos.append('General')
            ancho_columnas.append(30)
    else:
        raise ValueError('Los contratos no tienen los mismos entregables')


    titulos.append('Valor pagado')
    formatos.append('"$"#,##0_);("$"#,##0)')
    ancho_columnas.append(30)


    contenidos = []

    for formador in Formador.objects.filter(cargo__nombre="Formador Tipo 3").exclude(oculto=True):

        cantidad_list = []
        valor = 0
        for entregable in entregables_1:
            productos = Revision.objects.filter(formador_revision = formador).filter(productos__valor_entregable__entregable__id = entregable.entregable.id)
            cantidad = productos.values_list('productos__cantidad',flat=True)
            cantidad_list.append(sum(cantidad))
            valor_entregable = 0

            for x in productos.values_list('productos__cantidad','productos__valor_entregable__valor'):
                valor_entregable += x[0]*x[1]

            valor += valor_entregable

        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.get_cargo_string() if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ] + cantidad_list + [valor])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def acumulado_tipo_4(email):
    usuario = User.objects.get(email=email)
    nombre = "Acumulado formadores tipo 4"
    proceso = "REV-INF04"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    contrato_1 = Contratos.objects.filter(cargo__nombre="Formador Tipo 4").get(nombre = 'Capacitación 1')
    entregables_1 = ValorEntregable.objects.filter(contrato = contrato_1)

    contrato_2 = Contratos.objects.filter(cargo__nombre="Formador Tipo 4").get(nombre = 'Capacitación 2')
    entregables_2 = ValorEntregable.objects.filter(contrato = contrato_2)


    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']



    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]


    if list(entregables_1.values_list('entregable__id',flat=True)) == list(entregables_2.values_list('entregable__id',flat=True)):
        for entregable in entregables_1:
            titulos.append(entregable.entregable.sesion.nivel.nombre + ' - ' +
                           entregable.entregable.sesion.nombre + ' - ID:' +
                           str(entregable.entregable.id))
            formatos.append('General')
            ancho_columnas.append(30)
    else:
        raise ValueError('Los contratos no tienen los mismos entregables')


    titulos.append('Valor pagado')
    formatos.append('"$"#,##0_);("$"#,##0)')
    ancho_columnas.append(30)


    contenidos = []

    for formador in Formador.objects.filter(cargo__nombre="Formador Tipo 4").exclude(oculto=True):

        cantidad_list = []
        valor = 0
        for entregable in entregables_1:
            productos = Revision.objects.filter(formador_revision = formador).filter(productos__valor_entregable__entregable__id = entregable.entregable.id)
            cantidad = productos.values_list('productos__cantidad',flat=True)
            cantidad_list.append(sum(cantidad))
            valor_entregable = 0

            for x in productos.values_list('productos__cantidad','productos__valor_entregable__valor'):
                valor_entregable += x[0]*x[1]

            valor += valor_entregable

        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.get_cargo_string() if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ] + cantidad_list + [valor])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_chequeo(email,id_diplomado):
    usuario = User.objects.get(email=email)
    nombre = ''

    if id_diplomado == '1':
        nombre = "Matriz lista de chequeo de productos InnovaTIC"
    elif id_diplomado == '2':
        nombre = "Matriz lista de chequeo de productos TecnoTIC"
    elif id_diplomado == '3':
        nombre = "Matriz lista de chequeo de productos DirecTIC"
    elif id_diplomado == '4':
        nombre = "Matriz lista de chequeo de productos EscuelaTIC"


    proceso = "REV-INF05"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []

    if id_diplomado == '1':
        #wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Innovatic.xlsx')
        #ws = wb.get_sheet_by_name('InnovaTIC')
        diplomado = Diplomado.objects.get(id = 1)
        i = 6

        dict_productos = [{'letter':'Y','id':8},
                          {'letter':'Z','id':9},
                          {'letter':'AA','id':10},
                          {'letter':'AC','id':11},
                          {'letter':'AF','id':12},
                          {'letter':'AG','id':13},
                          {'letter':'AK','id':14},
                          {'letter':'AL','id':15},
                          {'letter':'AN','id':16},
                          {'letter':'AR','id':17},
                          {'letter':'AS','id':18},
                          {'letter':'AT','id':19},
                          {'letter':'AD','id':20},
                          {'letter':'AH','id':21},
                          {'letter':'AJ','id':22},
                          {'letter':'AO','id':23},
                          {'letter':'AQ','id':24},
                          {'letter':'AU','id':25},
                          {'letter':'AV','id':26},
                          {'letter':'AW','id':27},
                          {'letter':'AX','id':28},
                          {'letter':'AY','id':29},
                          {'letter':'BB','id':30},
                          {'letter':'BC','id':31},
                          {'letter':'BD','id':32},
                          {'letter':'BG','id':33},
                          {'letter':'BH','id':34},
                          {'letter':'BJ','id':35},
                          {'letter':'BN','id':36},
                          {'letter':'BO','id':37},
                          {'letter':'BP','id':38},
                          {'letter':'AZ','id':39},
                          {'letter':'BA','id':40},
                          {'letter':'BF','id':41},
                          {'letter':'BL','id':42},
                          {'letter':'BM','id':43},
                          {'letter':'BQ','id':44},
                          {'letter':'BR','id':45},
                          {'letter':'BS','id':46},
                          {'letter':'BT','id':47},
                          {'letter':'BU','id':48},
                          {'letter':'BX','id':49},
                          {'letter':'BY','id':50},
                          {'letter':'BZ','id':51},
                          {'letter':'CB','id':52},
                          {'letter':'CC','id':53},
                          {'letter':'CD','id':54},
                          {'letter':'CG','id':55},
                          {'letter':'CH','id':56},
                          {'letter':'CI','id':57},
                          {'letter':'BV','id':58},
                          {'letter':'CA','id':59},
                          {'letter':'CE','id':60},
                          {'letter':'CJ','id':61},
                          {'letter':'CK','id':62},
                          {'letter':'CL','id':63},
                          {'letter':'CM','id':64},
                          {'letter':'CO','id':65},
                          {'letter':'CR','id':66},
                          {'letter':'CS','id':67},
                          {'letter':'CN','id':68},
                          {'letter':'CQ','id':69},
                          ]

    elif id_diplomado == '2':
        #wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Tecnotic.xlsx')
        #ws = wb.get_sheet_by_name('TecnoTIC')
        diplomado = Diplomado.objects.get(id = 2)
        i = 6
        dict_productos = [{'letter':'Y','id':72},
                          {'letter':'AB','id':73},
                          {'letter':'AF','id':74},
                          {'letter':'AE','id':75},
                          {'letter':'AJ','id':76},
                          {'letter':'AI','id':77},
                          {'letter':'AP','id':78},
                          {'letter':'AO','id':79},
                          {'letter':'AC','id':80},
                          {'letter':'AD','id':81},
                          {'letter':'AG','id':82},
                          {'letter':'AH','id':83},
                          {'letter':'AK','id':84},
                          {'letter':'AM','id':85},
                          {'letter':'AQ','id':86},
                          {'letter':'AR','id':87},
                          {'letter':'AS','id':88},
                          {'letter':'AV','id':89},
                          {'letter':'AT','id':90},
                          {'letter':'AU','id':91},
                          {'letter':'AZ','id':92},
                          {'letter':'AY','id':93},
                          {'letter':'BB','id':94},
                          {'letter':'BE','id':95},
                          {'letter':'BD','id':96},
                          {'letter':'AW','id':97},
                          {'letter':'AX','id':98},
                          {'letter':'BA','id':99},
                          {'letter':'BC','id':100},
                          {'letter':'BF','id':101},
                          {'letter':'BG','id':102},
                          {'letter':'BH','id':103},
                          {'letter':'BJ','id':104},
                          {'letter':'BI','id':105},
                          {'letter':'BO','id':106},
                          {'letter':'BM','id':107},
                          {'letter':'BR','id':108},
                          {'letter':'BQ','id':109},
                          {'letter':'BV','id':110},
                          {'letter':'BT','id':111},
                          {'letter':'BK','id':112},
                          {'letter':'BP','id':113},
                          {'letter':'BS','id':114},
                          {'letter':'BW','id':115},
                          {'letter':'BX','id':116},
                          {'letter':'BY','id':117},
                          {'letter':'CB','id':118},
                          {'letter':'BZ','id':119},
                          {'letter':'CG','id':120},
                          {'letter':'CI','id':121},
                          {'letter':'CC','id':122},
                          {'letter':'CD','id':123},
                          {'letter':'CE','id':124}
                          ]
    elif id_diplomado == '3':
        #wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Directic.xlsx')
        #ws = wb.get_sheet_by_name('DirecTIC')
        diplomado = Diplomado.objects.get(id = 3)
        i = 6
        dict_productos = [{'letter':'Y','id':127},
                          {'letter':'Z','id':128},
                          {'letter':'AF','id':129},
                          {'letter':'AB','id':130},
                          {'letter':'AD','id':131},
                          {'letter':'AE','id':132},
                          {'letter':'AI','id':133},
                          {'letter':'AG','id':134},
                          {'letter':'AN','id':135},
                          {'letter':'AL','id':136},
                          {'letter':'AR','id':137},
                          {'letter':'AQ','id':138},
                          {'letter':'AW','id':139},
                          {'letter':'AU','id':140},
                          {'letter':'AF','id':141},
                          {'letter':'AJ','id':142},
                          {'letter':'AO','id':143},
                          {'letter':'AS','id':144},
                          {'letter':'AX','id':145},
                          {'letter':'AZ','id':146},
                          {'letter':'AY','id':147},
                          {'letter':'BE','id':148},
                          {'letter':'BD','id':149},
                          {'letter':'BH','id':150},
                          {'letter':'BG','id':151},
                          {'letter':'BA','id':152},
                          {'letter':'BF','id':153},
                          {'letter':'BI','id':154},
                          {'letter':'BL','id':155},
                          {'letter':'BJ','id':156},
                          {'letter':'BN','id':157},
                          {'letter':'BO','id':158},
                          {'letter':'BT','id':159},
                          {'letter':'BR','id':160},
                          {'letter':'BX','id':161},
                          {'letter':'BW','id':162},
                          {'letter':'BM','id':163},
                          {'letter':'BP','id':164},
                          {'letter':'BU','id':165},
                          {'letter':'BY','id':166},
                          {'letter':'CA','id':167},
                          {'letter':'CB','id':168},
                          {'letter':'CE','id':169},
                          {'letter':'CD','id':170},
                          {'letter':'CC','id':171},
                          {'letter':'CF','id':172}
                          ]
    else:
        #wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz EscuelaTIC.xlsx')
        #ws = wb.get_sheet_by_name('EscuelaTIC')
        diplomado = Diplomado.objects.get(id = 4)
        i = 10
        dict_productos = [{'letter':'Z','id':221},
                          {'letter':'AD','id':233},
                          {'letter':'AI','id':224},
                          {'letter':'AP','id':228},
        ]

    number = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='right',vertical='center',wrap_text=False),
                   number_format='0',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
             )

    text = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='left',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                 )

    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet('Hoja1')

    text = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'center', 'text_wrap':False})

    number = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'right', 'valign':'center', 'text_wrap':False,'num_format':'0'})

    validado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'center', 'text_wrap':False, 'pattern':1, 'bg_color':'#00B050'})

    enviado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'center', 'text_wrap':False, 'pattern':1, 'bg_color':'#FFC000'})

    cargado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'center', 'text_wrap':False})

    rechazado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'center', 'text_wrap':False, 'pattern':1, 'bg_color':'#FF0000'})


    for beneficiario in Beneficiario.objects.filter(diplomado = diplomado).order_by('formador'):
        ws.write('A'+str(i),beneficiario.region.nombre.upper(),text)
        #ws.cell('A'+str(i)).style = text

        ws.write('B'+str(i), beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text,text)
        #ws.cell('B'+str(i)).style = text

        ws.write('C'+str(i), beneficiario.radicado.secretaria.nombre.upper() if beneficiario.radicado != None else beneficiario.secretaria_text,text)
        #ws.cell('C'+str(i)).style = text

        ws.write('D'+str(i), beneficiario.radicado.numero if beneficiario.radicado != None else '',text)
        #ws.cell('D'+str(i)).style = number

        ws.write('E'+str(i), beneficiario.dane_ie_text,text)
        #ws.cell('E'+str(i)).style = number

        ws.write('F'+str(i), beneficiario.ie_text,text)
        #ws.cell('F'+str(i)).style = text

        ws.write('G'+str(i), beneficiario.radicado.dane_sede if beneficiario.radicado != None else beneficiario.dane_sede_text,text)
        #ws.cell('G'+str(i)).style = number

        ws.write('H'+str(i), beneficiario.radicado.nombre_sede.upper() if beneficiario.radicado != None else beneficiario.sede_text,text)
        #ws.cell('H'+str(i)).style = text

        ws.write('I'+str(i), beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text,text)
        #ws.cell('I'+str(i)).style = text

        ws.write('J'+str(i), beneficiario.radicado.ubicacion if beneficiario.radicado != None else '',text)
        #ws.cell('J'+str(i)).style = text

        ws.write('K'+str(i), beneficiario.get_grupo(),text)
        #ws.cell('K'+str(i)).style = text

        ws.write('L'+str(i), beneficiario.formador.get_full_name(),text)
        #ws.cell('L'+str(i)).style = text

        ws.write('M'+str(i), beneficiario.formador.cedula,text)
        #ws.cell('M'+str(i)).style = number

        ws.write('N'+str(i), beneficiario.apellidos,text)
        #ws.cell('N'+str(i)).style = text

        ws.write('O'+str(i), beneficiario.nombres,text)
        #ws.cell('O'+str(i)).style = text

        ws.write('P'+str(i), beneficiario.cedula,text)
        #ws.cell('P'+str(i)).style = number

        ws.write('Q'+str(i), beneficiario.correo,text)
        #ws.cell('Q'+str(i)).style = text

        ws.write('R'+str(i), beneficiario.telefono_fijo,text)
        #ws.cell('R'+str(i)).style = text

        ws.write('S'+str(i), beneficiario.telefono_celular,text)
        #ws.cell('S'+str(i)).style = text

        ws.write('T'+str(i), beneficiario.area.nombre.upper() if beneficiario.area != None else '',text)
        #ws.cell('T'+str(i)).style = text

        ws.write('U'+str(i), beneficiario.grado.nombre.upper() if beneficiario.grado != None else '',text)
        #ws.cell('U'+str(i)).style = text

        ws.write('V'+str(i), '',text)
        #ws.cell('V'+str(i)).style = text

        ws.write('W'+str(i), beneficiario.genero,text)
        #ws.cell('W'+str(i)).style = text

        ws.write('X'+str(i), beneficiario.estado,text)
        #ws.cell('X'+str(i)).style = text

        ws.write('Y'+str(i), '')
        #ws.cell('Y'+str(i)).style = text

        for producto in dict_productos:
            entregable = Entregable.objects.get(id = producto['id'])

            evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable)
            evidencias_validado = Evidencia.objects.filter(beneficiarios_validados = beneficiario,entregable = entregable)

            q = Q(evidencias__id__in = evidencias_cargado.values_list('id',flat=True)) | \
                Q(evidencias__id__in = evidencias_validado.values_list('id',flat=True))

            reds = Red.objects.filter(q)

            if evidencias_validado.count() > 0:
                #red = Red.objects.get(evidencias__id = evidencias_validado[0].id)
                #ws.write_url( producto['letter'] + str(i) , 'https://sican.asoandes.org' + evidencias_validado[0].get_archivo_url() , validado, 'RED-' + str(red.id))
                ws.write( producto['letter'] + str(i) , 'SIC-' + str(evidencias_validado[0].id), validado )


            elif evidencias_validado.count() == 0 and evidencias_cargado.count() > 0:
                ultima_evidencia_cargada = evidencias_cargado[len(evidencias_cargado)-1]

                try:
                    red = Red.objects.get(id = ultima_evidencia_cargada.red_id)
                except:
                    # si ningun red contiene la evidencia
                    ws.write( producto['letter'] + str(i), 'SIC-' + str(ultima_evidencia_cargada.id), cargado )

                else:
                    # si hay un red que contiene la evidencia
                    if red.retroalimentacion:
                        # si el red fue retroalimentado

                        evidencias_rechazado = Evidencia.objects.filter(beneficiarios_rechazados__beneficiario_rechazo = beneficiario,
                                                                        beneficiarios_rechazados__red_id = red.id,
                                                                        beneficiarios_rechazados__evidencia_id = ultima_evidencia_cargada.id)

                        if evidencias_rechazado > 0:
                            try:
                                causa = evidencias_rechazado[0].beneficiarios_rechazados.get(beneficiario_rechazo = beneficiario).observacion
                            except:
                                causa = ''
                            ws.write( producto['letter'] + str(i) , 'SIC-' + str(ultima_evidencia_cargada.id), rechazado )
                            ws.write_comment( producto['letter'] + str(i) , 'RED-' + str(red.id) + ':\n' + causa )
                        else:
                            pass


                    else:
                        #si el red no ha sido retroalimentado
                        ws.write( producto['letter'] + str(i) , 'SIC-' + str(ultima_evidencia_cargada.id), enviado )
                        ws.write_comment( producto['letter'] + str(i) , 'RED-' + str(ultima_evidencia_cargada.red_id) )


            else:
                pass

        i += 1

    wb.close()

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_chequeo_formador(email,id_formador):
    usuario = User.objects.get(email=email)

    formador = Formador.objects.get(id = int(id_formador))
    id_diplomado = ''
    nombre = ''

    if formador.get_cargo_string() == 'Formador Tipo 1':
        id_diplomado = '1'
    elif formador.get_cargo_string() == 'Formador Tipo 2':
        id_diplomado = '2'
    elif formador.get_cargo_string() == 'Formador Tipo 3':
        id_diplomado = '3'
    elif formador.get_cargo_string() == 'Formador Tipo 4':
        id_diplomado = '4'


    if id_diplomado == '1':
        nombre = "Matriz lista de chequeo de productos InnovaTIC - " + formador.get_full_name()
    elif id_diplomado == '2':
        nombre = "Matriz lista de chequeo de productos TecnoTIC - " + formador.get_full_name()
    elif id_diplomado == '3':
        nombre = "Matriz lista de chequeo de productos DirecTIC - " + formador.get_full_name()
    elif id_diplomado == '4':
        nombre = "Matriz lista de chequeo de productos EscuelaTIC - " + formador.get_full_name()


    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []

    if id_diplomado == '1':
        #wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Innovatic.xlsx')
        #ws = wb.get_sheet_by_name('InnovaTIC')
        diplomado = Diplomado.objects.get(id = 1)
        i = 6

        dict_productos = [{'letter':'Y','id':8},
                          {'letter':'Z','id':9},
                          {'letter':'AA','id':10},
                          {'letter':'AC','id':11},
                          {'letter':'AF','id':12},
                          {'letter':'AG','id':13},
                          {'letter':'AK','id':14},
                          {'letter':'AL','id':15},
                          {'letter':'AN','id':16},
                          {'letter':'AR','id':17},
                          {'letter':'AS','id':18},
                          {'letter':'AT','id':19},
                          {'letter':'AD','id':20},
                          {'letter':'AI','id':21},
                          {'letter':'AJ','id':22},
                          {'letter':'AO','id':23},
                          {'letter':'AQ','id':24},
                          {'letter':'AU','id':25},
                          {'letter':'AV','id':26},
                          {'letter':'AW','id':27},
                          {'letter':'AX','id':28},
                          {'letter':'AY','id':29},
                          {'letter':'BB','id':30},
                          {'letter':'BC','id':31},
                          {'letter':'BD','id':32},
                          {'letter':'BG','id':33},
                          {'letter':'BH','id':34},
                          {'letter':'BJ','id':35},
                          {'letter':'BN','id':36},
                          {'letter':'BO','id':37},
                          {'letter':'BP','id':38},
                          {'letter':'AZ','id':39},
                          {'letter':'BA','id':40},
                          {'letter':'BF','id':41},
                          {'letter':'BL','id':42},
                          {'letter':'BM','id':43},
                          {'letter':'BQ','id':44},
                          {'letter':'BR','id':45},
                          {'letter':'BS','id':46},
                          {'letter':'BT','id':47},
                          {'letter':'BU','id':48},
                          {'letter':'BX','id':49},
                          {'letter':'BY','id':50},
                          {'letter':'BZ','id':51},
                          {'letter':'CB','id':52},
                          {'letter':'CC','id':53},
                          {'letter':'CD','id':54},
                          {'letter':'CG','id':55},
                          {'letter':'CH','id':56},
                          {'letter':'CI','id':57},
                          {'letter':'BV','id':58},
                          {'letter':'CA','id':59},
                          {'letter':'CE','id':60},
                          {'letter':'CJ','id':61},
                          {'letter':'CK','id':62},
                          {'letter':'CL','id':63},
                          {'letter':'CM','id':64},
                          {'letter':'CO','id':65},
                          {'letter':'CR','id':66},
                          {'letter':'CS','id':67},
                          {'letter':'CN','id':68},
                          {'letter':'CQ','id':69},
                          ]

    elif id_diplomado == '2':
        #wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Tecnotic.xlsx')
        #ws = wb.get_sheet_by_name('TecnoTIC')
        diplomado = Diplomado.objects.get(id = 2)
        i = 6
        dict_productos = [{'letter':'Y','id':72},
                          {'letter':'AB','id':73},
                          {'letter':'AF','id':74},
                          {'letter':'AE','id':75},
                          {'letter':'AJ','id':76},
                          {'letter':'AI','id':77},
                          {'letter':'AP','id':78},
                          {'letter':'AO','id':79},
                          {'letter':'AC','id':80},
                          {'letter':'AD','id':81},
                          {'letter':'AG','id':82},
                          {'letter':'AH','id':83},
                          {'letter':'AK','id':84},
                          {'letter':'AM','id':85},
                          {'letter':'AQ','id':86},
                          {'letter':'AR','id':87},
                          {'letter':'AS','id':88},
                          {'letter':'AV','id':89},
                          {'letter':'AT','id':90},
                          {'letter':'AU','id':91},
                          {'letter':'AZ','id':92},
                          {'letter':'AY','id':93},
                          {'letter':'BB','id':94},
                          {'letter':'BE','id':95},
                          {'letter':'BD','id':96},
                          {'letter':'AW','id':97},
                          {'letter':'AX','id':98},
                          {'letter':'BA','id':99},
                          {'letter':'BC','id':100},
                          {'letter':'BF','id':101},
                          {'letter':'BG','id':102},
                          {'letter':'BH','id':103},
                          {'letter':'BJ','id':104},
                          {'letter':'BI','id':105},
                          {'letter':'BO','id':106},
                          {'letter':'BM','id':107},
                          {'letter':'BR','id':108},
                          {'letter':'BQ','id':109},
                          {'letter':'BV','id':110},
                          {'letter':'BT','id':111},
                          {'letter':'BK','id':112},
                          {'letter':'BP','id':113},
                          {'letter':'BS','id':114},
                          {'letter':'BW','id':115},
                          {'letter':'BX','id':116},
                          {'letter':'BY','id':117},
                          {'letter':'CB','id':118},
                          {'letter':'BZ','id':119},
                          {'letter':'CG','id':120},
                          {'letter':'CI','id':121},
                          {'letter':'CC','id':122},
                          {'letter':'CD','id':123},
                          {'letter':'CE','id':124}
                          ]
    elif id_diplomado == '3':
        #wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Directic.xlsx')
        #ws = wb.get_sheet_by_name('DirecTIC')
        diplomado = Diplomado.objects.get(id = 3)
        i = 6
        dict_productos = [{'letter':'Y','id':127},
                          {'letter':'Z','id':128},
                          {'letter':'AF','id':129},
                          {'letter':'AB','id':130},
                          {'letter':'AD','id':131},
                          {'letter':'AE','id':132},
                          {'letter':'AI','id':133},
                          {'letter':'AG','id':134},
                          {'letter':'AN','id':135},
                          {'letter':'AL','id':136},
                          {'letter':'AR','id':137},
                          {'letter':'AQ','id':138},
                          {'letter':'AW','id':139},
                          {'letter':'AU','id':140},
                          {'letter':'AF','id':141},
                          {'letter':'AJ','id':142},
                          {'letter':'AO','id':143},
                          {'letter':'AS','id':144},
                          {'letter':'AX','id':145},
                          {'letter':'AZ','id':146},
                          {'letter':'AY','id':147},
                          {'letter':'BE','id':148},
                          {'letter':'BD','id':149},
                          {'letter':'BH','id':150},
                          {'letter':'BG','id':151},
                          {'letter':'BA','id':152},
                          {'letter':'BF','id':153},
                          {'letter':'BI','id':154},
                          {'letter':'BL','id':155},
                          {'letter':'BJ','id':156},
                          {'letter':'BN','id':157},
                          {'letter':'BO','id':158},
                          {'letter':'BT','id':159},
                          {'letter':'BR','id':160},
                          {'letter':'BX','id':161},
                          {'letter':'BW','id':162},
                          {'letter':'BM','id':163},
                          {'letter':'BP','id':164},
                          {'letter':'BU','id':165},
                          {'letter':'BY','id':166},
                          {'letter':'CA','id':167},
                          {'letter':'CB','id':168},
                          {'letter':'CE','id':169},
                          {'letter':'CD','id':170},
                          {'letter':'CC','id':171},
                          {'letter':'CF','id':172}
                          ]
    else:
        #wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz EscuelaTIC.xlsx')
        #ws = wb.get_sheet_by_name('EscuelaTIC')
        diplomado = Diplomado.objects.get(id = 4)
        i = 10
        dict_productos = [{'letter':'Z','id':221},
                          {'letter':'AD','id':233},
                          {'letter':'AI','id':224},
                          {'letter':'AP','id':228},
        ]

    """
    number = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='right',vertical='center',wrap_text=False),
                   number_format='0',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
             )

    text = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='left',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                 )


    validado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FF00B050',end_color='FF00B050')
                 )

    enviado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FFFFC000',end_color='FFFFC000')
                 )

    cargado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                 )


    rechazado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FFFF0000',end_color='FFFF0000')
                 )

    """

    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet()

    for beneficiario in Beneficiario.objects.filter(diplomado = diplomado).order_by('formador').filter(formador__id = id_formador):
        ws.write('A'+str(i), beneficiario.region.nombre.upper())
        #ws.cell('A'+str(i)).style = text

        ws.write('B'+str(i), beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text)
        #ws.cell('B'+str(i)).style = text

        ws.write('C'+str(i), beneficiario.radicado.secretaria.nombre.upper() if beneficiario.radicado != None else beneficiario.secretaria_text)
        #ws.cell('C'+str(i)).style = text

        ws.write('D'+str(i), beneficiario.radicado.numero if beneficiario.radicado != None else '')
        #ws.cell('D'+str(i)).style = number

        ws.write('E'+str(i), beneficiario.dane_ie_text)
        #ws.cell('E'+str(i)).style = number

        ws.write('F'+str(i), beneficiario.ie_text)
        #ws.cell('F'+str(i)).style = text

        ws.write('G'+str(i), beneficiario.radicado.dane_sede if beneficiario.radicado != None else beneficiario.dane_sede_text)
        #ws.cell('G'+str(i)).style = number

        ws.write('H'+str(i), beneficiario.radicado.nombre_sede.upper() if beneficiario.radicado != None else beneficiario.sede_text)
        #ws.cell('H'+str(i)).style = text

        ws.write('I'+str(i), beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text)
        #ws.cell('I'+str(i)).style = text

        ws.write('J'+str(i), beneficiario.radicado.ubicacion if beneficiario.radicado != None else '')
        #ws.cell('J'+str(i)).style = text

        ws.write('K'+str(i), beneficiario.get_grupo())
        #ws.cell('K'+str(i)).style = text

        ws.write('L'+str(i), beneficiario.formador.get_full_name())
        #ws.cell('L'+str(i)).style = text

        ws.write('M'+str(i), beneficiario.formador.cedula)
        #ws.cell('M'+str(i)).style = number

        ws.write('N'+str(i), beneficiario.apellidos)
        #ws.cell('N'+str(i)).style = text

        ws.write('O'+str(i), beneficiario.nombres)
        #ws.cell('O'+str(i)).style = text

        ws.write('P'+str(i), beneficiario.cedula)
        #ws.cell('P'+str(i)).style = number

        ws.write('Q'+str(i), beneficiario.correo)
        #ws.cell('Q'+str(i)).style = text

        ws.write('R'+str(i), beneficiario.telefono_fijo)
        #ws.cell('R'+str(i)).style = text

        ws.write('S'+str(i), beneficiario.telefono_celular)
        #ws.cell('S'+str(i)).style = text

        ws.write('T'+str(i), beneficiario.area.nombre.upper() if beneficiario.area != None else '')
        #ws.cell('T'+str(i)).style = text

        ws.write('U'+str(i), beneficiario.grado.nombre.upper() if beneficiario.grado != None else '')
        #ws.cell('U'+str(i)).style = text

        ws.write('V'+str(i), '')
        #ws.cell('V'+str(i)).style = text

        ws.write('W'+str(i), beneficiario.genero)
        #ws.cell('W'+str(i)).style = text

        ws.write('X'+str(i), beneficiario.estado)
        #ws.cell('X'+str(i)).style = text

        ws.write('Y'+str(i), '')
        #ws.cell('Y'+str(i)).style = text

        """

        for producto in dict_productos:
            entregable = Entregable.objects.get(id = producto['id'])

            evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable)
            evidencias_validado = Evidencia.objects.filter(beneficiarios_validados = beneficiario,entregable = entregable)

            q = Q(evidencias__id__in = evidencias_cargado.values_list('id',flat=True)) | \
                Q(evidencias__id__in = evidencias_validado.values_list('id',flat=True))

            reds = Red.objects.filter(q)

            if evidencias_validado.count() > 0:
                red = Red.objects.get(evidencias__id = evidencias_validado[0].id)
                ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                ws.cell(producto['letter']+str(i)).style = validado
                ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + evidencias_validado[0].get_archivo_url()
                ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(evidencias_validado[0].id),'SICAN')


            elif evidencias_validado.count() == 0 and evidencias_cargado.count() > 0:
                ultima_evidencia_cargada = evidencias_cargado[len(evidencias_cargado)-1]

                try:
                    red = reds.get(evidencias__id = ultima_evidencia_cargada.id)
                except:
                    # si ningun red contiene la evidencia
                    ws.cell(producto['letter']+str(i)).value = 'SIC-' + str(ultima_evidencia_cargada.id)
                    ws.cell(producto['letter']+str(i)).style = cargado
                    ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()

                else:
                    # si hay un red que contiene la evidencia
                    if red.retroalimentacion:
                        # si el red fue retroalimentado

                        evidencias_rechazado = Evidencia.objects.filter(beneficiarios_rechazados__beneficiario_rechazo = beneficiario,
                                                                        beneficiarios_rechazados__red_id = red.id,
                                                                        beneficiarios_rechazados__evidencia_id = ultima_evidencia_cargada.id)

                        if evidencias_rechazado > 0:
                            try:
                                causa = evidencias_rechazado[0].beneficiarios_rechazados.get(beneficiario_rechazo = beneficiario).observacion
                            except:
                                causa = ''
                            ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                            ws.cell(producto['letter']+str(i)).style = rechazado
                            ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()
                            ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(ultima_evidencia_cargada.id) + ':\n' + causa,'SICAN')
                        else:
                            pass


                    else:
                        #si el red no ha sido retroalimentado
                        ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                        ws.cell(producto['letter']+str(i)).style = enviado
                        ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()
                        ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(ultima_evidencia_cargada.id),'SICAN')


            else:
                ws.cell(producto['letter']+str(i)).style = text
        """
        i += 1

    wb.close()

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def zip_ss(email):

    if os.path.exists("C:\\Temp\\ss.zip"):
        os.remove("C:\\Temp\\ss.zip")

    usuario = User.objects.get(email=email)
    nombre = "Zip: Seguridad Social"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")

    ids = [{'id':13,'mes':'Julio'},
           {'id':8,'mes':'Agosto'},
           {'id':17,'mes':'Septiembre'},
           {'id':18,'mes':'Octubre'},
           {'id':19,'mes':'Noviembre'},
           {'id':20,'mes':'Diciembre'},
    ]

    zip = zipfile.ZipFile('C:\\Temp\\ss.zip',"w",allowZip64=True)

    for id in ids:
        for soporte in SoporteFormadores.objects.filter(tipo__id = id['id']).exclude(oculto = True):
            if str(soporte.archivo) != '':
                if os.path.exists(soporte.archivo.path):
                    zip.write(soporte.archivo.path,id['mes']+'/'+soporte.formador.get_full_name()+'/'+os.path.basename(soporte.archivo.path))

    zip.close()
    informe.archivo = File(open('C:\\Temp\\ss.zip'))
    informe.save()

    shutil.copy('C:\\Temp\\ss.zip',informe.archivo.path)

    return "Zip creado Seguridad Social"

@app.task
def descargas_certificados_escuelatic(email):
    usuario = User.objects.get(email=email)
    nombre = "Reporte descargas certificaados Escuela TIC"
    proceso = "FOR-INF010"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Región','Departamento','Municipios','Formador','Nombres','Apellidos','Cedula','IP','Fecha Descarga']

    formatos = ['General','General','General','General','General','General','General','0','General','d/m/yy']


    ancho_columnas =  [30,20,20,20,20,20,20,20,20,20]

    contenidos = []

    for beneficiario in Beneficiario.objects.filter(diplomado__numero = 4).exclude(fecha_descarga = None):
        contenidos.append([
            'BENE-'+unicode(beneficiario.id),
            beneficiario.region.nombre,
            beneficiario.departamento_text if beneficiario.radicado == None else beneficiario.radicado.municipio.departamento.nombre.upper(),
            beneficiario.municipio_text if beneficiario.radicado == None else beneficiario.radicado.municipio.nombre.upper(),
            beneficiario.formador.get_full_name(),
            beneficiario.nombres,
            beneficiario.apellidos,
            beneficiario.cedula,
            beneficiario.ip_descarga,
            beneficiario.fecha_descarga
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def progreso_listados_actas(email):
    usuario = User.objects.get(email=email)
    nombre = "Progreso de carga listados de asistencia y actas de compromiso"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    output = StringIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/PROGRESO.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')

    dict_productos_t1 = [{'letter':'B','id':8},
                          {'letter':'C','id':9},
                          {'letter':'D','id':12},
                          {'letter':'E','id':14},
                          {'letter':'F','id':17},
                          {'letter':'K','id':27},
                          {'letter':'L','id':30},
                          {'letter':'M','id':33},
                          {'letter':'N','id':36},
                          {'letter':'S','id':46},
                          {'letter':'T','id':49},
                          {'letter':'U','id':52},
                          {'letter':'V','id':55},
                          {'letter':'AA','id':63},
                          {'letter':'AB','id':66},
                          ]

    dict_productos_t2 = [{'letter':'B','id':72},
                          {'letter':'C','id':73},
                          {'letter':'D','id':74},
                          {'letter':'E','id':76},
                          {'letter':'F','id':78},
                          {'letter':'K','id':89},
                          {'letter':'L','id':92},
                          {'letter':'M','id':94},
                          {'letter':'N','id':95},
                          {'letter':'S','id':104},
                          {'letter':'T','id':106},
                          {'letter':'U','id':108},
                          {'letter':'V','id':110},
                          {'letter':'AA','id':118},
                          {'letter':'AB','id':120},
                          ]

    dict_productos_t3 = [{'letter':'B','id':127},
                          {'letter':'C','id':128},
                          {'letter':'D','id':131},
                          {'letter':'E','id':133},
                          {'letter':'F','id':135},
                          {'letter':'G','id':137},
                          {'letter':'H','id':139},
                          {'letter':'O','id':146},
                          {'letter':'P','id':148},
                          {'letter':'Q','id':150},
                          {'letter':'U','id':155},
                          {'letter':'V','id':157},
                          {'letter':'W','id':159},
                          {'letter':'X','id':161},
                          {'letter':'AC','id':167},
                          {'letter':'AD','id':169},
                          ]

    dict_productos_t4 = [ {'letter':'B','id':224},
                          {'letter':'C','id':228},
        ]

    for producto in dict_productos_t1:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 5
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_cargados__region__id = region.id).values_list('beneficiarios_cargados',flat=True).distinct().count()
            i += 1

    for producto in dict_productos_t2:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 16
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_cargados__region__id = region.id).values_list('beneficiarios_cargados',flat=True).distinct().count()
            i += 1

    for producto in dict_productos_t3:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 27
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_cargados__region__id = region.id).values_list('beneficiarios_cargados',flat=True).distinct().count()
            i += 1


    for producto in dict_productos_t4:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 38
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_cargados__region__id = region.id).values_list('beneficiarios_cargados',flat=True).distinct().count()
            i += 1

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def progreso_listados_actas_aprobadas(email):
    usuario = User.objects.get(email=email)
    nombre = "Aprobación listados de asistencia y actas de compromiso"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    output = StringIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/PROGRESO.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')

    dict_productos_t1 = [{'letter':'B','id':8},
                          {'letter':'C','id':9},
                          {'letter':'D','id':12},
                          {'letter':'E','id':14},
                          {'letter':'F','id':17},
                          {'letter':'K','id':27},
                          {'letter':'L','id':30},
                          {'letter':'M','id':33},
                          {'letter':'N','id':36},
                          {'letter':'S','id':46},
                          {'letter':'T','id':49},
                          {'letter':'U','id':52},
                          {'letter':'V','id':55},
                          {'letter':'AA','id':63},
                          {'letter':'AB','id':66},
                          ]

    dict_productos_t2 = [{'letter':'B','id':72},
                          {'letter':'C','id':73},
                          {'letter':'D','id':74},
                          {'letter':'E','id':76},
                          {'letter':'F','id':78},
                          {'letter':'K','id':89},
                          {'letter':'L','id':92},
                          {'letter':'M','id':94},
                          {'letter':'N','id':95},
                          {'letter':'S','id':104},
                          {'letter':'T','id':106},
                          {'letter':'U','id':108},
                          {'letter':'V','id':110},
                          {'letter':'AA','id':118},
                          {'letter':'AB','id':120},
                          ]

    dict_productos_t3 = [{'letter':'B','id':127},
                          {'letter':'C','id':128},
                          {'letter':'D','id':131},
                          {'letter':'E','id':133},
                          {'letter':'F','id':135},
                          {'letter':'G','id':137},
                          {'letter':'H','id':139},
                          {'letter':'O','id':146},
                          {'letter':'P','id':148},
                          {'letter':'Q','id':150},
                          {'letter':'U','id':155},
                          {'letter':'V','id':157},
                          {'letter':'W','id':159},
                          {'letter':'X','id':161},
                          {'letter':'AC','id':167},
                          {'letter':'AD','id':169},
                          ]

    dict_productos_t4 = [ {'letter':'B','id':224},
                          {'letter':'C','id':228},
        ]

    for producto in dict_productos_t1:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 5
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_validados__region__id = region.id).values_list('beneficiarios_validados',flat=True).distinct().count()
            i += 1

    for producto in dict_productos_t2:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 16
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_validados__region__id = region.id).values_list('beneficiarios_validados',flat=True).distinct().count()
            i += 1

    for producto in dict_productos_t3:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 27
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_validados__region__id = region.id).values_list('beneficiarios_validados',flat=True).distinct().count()
            i += 1


    for producto in dict_productos_t4:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 38
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_validados__region__id = region.id).values_list('beneficiarios_validados',flat=True).distinct().count()
            i += 1

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def progreso_virtuales(email):
    usuario = User.objects.get(email=email)
    nombre = "Progreso de carga actividades virtuales"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    output = StringIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/PROGRESO_VIRTUAL.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')

    dict_productos_t1 = [{'letter':'B','id':20},
                          {'letter':'C','id':21},
                          {'letter':'D','id':22},
                          {'letter':'E','id':15},
                          {'letter':'F','id':16},
                          {'letter':'G','id':23},

                          {'letter':'N','id':28},
                          {'letter':'O','id':40},
                          {'letter':'P','id':31},
                          {'letter':'Q','id':34},

                          {'letter':'V','id':58},
                          {'letter':'W','id':59},
                          {'letter':'X','id':60},

                          {'letter':'AB','id':64},
                          {'letter':'AC','id':67},

                          ]

    dict_productos_t2 = [{'letter':'B','id':75},
                          {'letter':'C','id':77},
                          {'letter':'D','id':84},
                          {'letter':'E','id':85},

                          {'letter':'J','id':97},
                          {'letter':'K','id':98},
                          {'letter':'L','id':93},
                          {'letter':'M','id':99},
                          {'letter':'N','id':100},

                          {'letter':'T','id':112},
                          {'letter':'U','id':109},

                          {'letter':'X','id':119},
                          {'letter':'Y','id':124},
                          {'letter':'Z','id':121},
                          ]

    dict_productos_t3 = [{'letter':'B','id':132},
                          {'letter':'C','id':134},
                          {'letter':'D','id':142},
                          {'letter':'E','id':143},
                          {'letter':'F','id':138},
                          {'letter':'G','id':140},

                          {'letter':'N','id':147},
                          {'letter':'O','id':152},
                          {'letter':'P','id':149},
                          {'letter':'Q','id':151},

                          {'letter':'V','id':156},
                          {'letter':'W','id':164},
                          {'letter':'X','id':165},
                          {'letter':'Y','id':162},
                          {'letter':'Z','id':166},

                          {'letter':'AF','id':171},
                          ]


    for producto in dict_productos_t1:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 6
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_cargados__region__id = region.id).values_list('beneficiarios_cargados',flat=True).distinct().count()
            i += 1

    for producto in dict_productos_t2:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 18
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_cargados__region__id = region.id).values_list('beneficiarios_cargados',flat=True).distinct().count()
            i += 1

    for producto in dict_productos_t3:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 30
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_cargados__region__id = region.id).values_list('beneficiarios_cargados',flat=True).distinct().count()
            i += 1


    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def progreso_virtuales_aprobadas(email):
    usuario = User.objects.get(email=email)
    nombre = "Aprobacion actividades virtuales"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    output = StringIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/PROGRESO_VIRTUAL.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')

    dict_productos_t1 = [{'letter':'B','id':20},
                          {'letter':'C','id':21},
                          {'letter':'D','id':22},
                          {'letter':'E','id':15},
                          {'letter':'F','id':16},
                          {'letter':'G','id':23},

                          {'letter':'N','id':28},
                          {'letter':'O','id':40},
                          {'letter':'P','id':31},
                          {'letter':'Q','id':34},

                          {'letter':'V','id':58},
                          {'letter':'W','id':59},
                          {'letter':'X','id':60},

                          {'letter':'AB','id':64},
                          {'letter':'AC','id':67},

                          ]

    dict_productos_t2 = [{'letter':'B','id':75},
                          {'letter':'C','id':77},
                          {'letter':'D','id':84},
                          {'letter':'E','id':85},

                          {'letter':'J','id':97},
                          {'letter':'K','id':98},
                          {'letter':'L','id':93},
                          {'letter':'M','id':99},
                          {'letter':'N','id':100},

                          {'letter':'T','id':112},
                          {'letter':'U','id':109},

                          {'letter':'X','id':119},
                          {'letter':'Y','id':124},
                          {'letter':'Z','id':121},
                          ]

    dict_productos_t3 = [{'letter':'B','id':132},
                          {'letter':'C','id':134},
                          {'letter':'D','id':142},
                          {'letter':'E','id':143},
                          {'letter':'F','id':138},
                          {'letter':'G','id':140},

                          {'letter':'N','id':147},
                          {'letter':'O','id':152},
                          {'letter':'P','id':149},
                          {'letter':'Q','id':151},

                          {'letter':'V','id':156},
                          {'letter':'W','id':164},
                          {'letter':'X','id':165},
                          {'letter':'Y','id':162},
                          {'letter':'Z','id':166},

                          {'letter':'AF','id':171},
                          ]


    for producto in dict_productos_t1:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 6
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_validados__region__id = region.id).values_list('beneficiarios_validados',flat=True).distinct().count()
            i += 1

    for producto in dict_productos_t2:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 18
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_validados__region__id = region.id).values_list('beneficiarios_validados',flat=True).distinct().count()
            i += 1

    for producto in dict_productos_t3:
        evidencias = Evidencia.objects.filter(entregable__id = producto['id'])
        i = 30
        for region in Region.objects.filter(id__in=[1,2]):
            ws.cell('A' + str(i)).value = region.nombre.upper()
            ws.cell(producto['letter'] + str(i)).value = evidencias.filter(beneficiarios_validados__region__id = region.id).values_list('beneficiarios_validados',flat=True).distinct().count()
            i += 1


    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_chequeo_actividad(email,id_actividad):
    usuario = User.objects.get(email=email)
    nombre = ''

    id_diplomado = str(Entregable.objects.get(id = id_actividad).sesion.nivel.diplomado.id)

    if id_diplomado == '1':
        nombre = "Matriz lista de chequeo de productos InnovaTIC actividad: " + str(id_actividad)
    elif id_diplomado == '2':
        nombre = "Matriz lista de chequeo de productos TecnoTIC actividad: " + str(id_actividad)
    elif id_diplomado == '3':
        nombre = "Matriz lista de chequeo de productos DirecTIC actividad: " + str(id_actividad)
    elif id_diplomado == '4':
        nombre = "Matriz lista de chequeo de productos EscuelaTIC actividad: " + str(id_actividad)


    proceso = "REV-INF05"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []

    if id_diplomado == '1':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Innovatic.xlsx')
        ws = wb.get_sheet_by_name('InnovaTIC')
        diplomado = Diplomado.objects.get(id = 1)
        i = 6

        dict_productos = [{'letter':'Y','id':8},
                          {'letter':'Z','id':9},
                          {'letter':'AA','id':10},
                          {'letter':'AC','id':11},
                          {'letter':'AF','id':12},
                          {'letter':'AG','id':13},
                          {'letter':'AK','id':14},
                          {'letter':'AL','id':15},
                          {'letter':'AN','id':16},
                          {'letter':'AR','id':17},
                          {'letter':'AS','id':18},
                          {'letter':'AT','id':19},
                          {'letter':'AD','id':20},
                          {'letter':'AI','id':21},
                          {'letter':'AJ','id':22},
                          {'letter':'AO','id':23},
                          {'letter':'AQ','id':24},
                          {'letter':'AU','id':25},
                          {'letter':'AV','id':26},
                          {'letter':'AW','id':27},
                          {'letter':'AX','id':28},
                          {'letter':'AY','id':29},
                          {'letter':'BB','id':30},
                          {'letter':'BC','id':31},
                          {'letter':'BD','id':32},
                          {'letter':'BG','id':33},
                          {'letter':'BH','id':34},
                          {'letter':'BJ','id':35},
                          {'letter':'BN','id':36},
                          {'letter':'BO','id':37},
                          {'letter':'BP','id':38},
                          {'letter':'AZ','id':39},
                          {'letter':'BA','id':40},
                          {'letter':'BF','id':41},
                          {'letter':'BL','id':42},
                          {'letter':'BM','id':43},
                          {'letter':'BQ','id':44},
                          {'letter':'BR','id':45},
                          {'letter':'BS','id':46},
                          {'letter':'BT','id':47},
                          {'letter':'BU','id':48},
                          {'letter':'BX','id':49},
                          {'letter':'BY','id':50},
                          {'letter':'BZ','id':51},
                          {'letter':'CB','id':52},
                          {'letter':'CC','id':53},
                          {'letter':'CD','id':54},
                          {'letter':'CG','id':55},
                          {'letter':'CH','id':56},
                          {'letter':'CI','id':57},
                          {'letter':'BV','id':58},
                          {'letter':'CA','id':59},
                          {'letter':'CE','id':60},
                          {'letter':'CJ','id':61},
                          {'letter':'CK','id':62},
                          {'letter':'CL','id':63},
                          {'letter':'CM','id':64},
                          {'letter':'CO','id':65},
                          {'letter':'CR','id':66},
                          {'letter':'CS','id':67},
                          {'letter':'CN','id':68},
                          {'letter':'CQ','id':69},
                          ]

    elif id_diplomado == '2':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Tecnotic.xlsx')
        ws = wb.get_sheet_by_name('TecnoTIC')
        diplomado = Diplomado.objects.get(id = 2)
        i = 6
        dict_productos = [{'letter':'Y','id':72},
                          {'letter':'AB','id':73},
                          {'letter':'AF','id':74},
                          {'letter':'AE','id':75},
                          {'letter':'AJ','id':76},
                          {'letter':'AI','id':77},
                          {'letter':'AP','id':78},
                          {'letter':'AO','id':79},
                          {'letter':'AC','id':80},
                          {'letter':'AD','id':81},
                          {'letter':'AG','id':82},
                          {'letter':'AH','id':83},
                          {'letter':'AK','id':84},
                          {'letter':'AM','id':85},
                          {'letter':'AQ','id':86},
                          {'letter':'AR','id':87},
                          {'letter':'AS','id':88},
                          {'letter':'AV','id':89},
                          {'letter':'AT','id':90},
                          {'letter':'AU','id':91},
                          {'letter':'AZ','id':92},
                          {'letter':'AY','id':93},
                          {'letter':'BB','id':94},
                          {'letter':'BE','id':95},
                          {'letter':'BD','id':96},
                          {'letter':'AW','id':97},
                          {'letter':'AX','id':98},
                          {'letter':'BA','id':99},
                          {'letter':'BC','id':100},
                          {'letter':'BF','id':101},
                          {'letter':'BG','id':102},
                          {'letter':'BH','id':103},
                          {'letter':'BJ','id':104},
                          {'letter':'BI','id':105},
                          {'letter':'BO','id':106},
                          {'letter':'BM','id':107},
                          {'letter':'BR','id':108},
                          {'letter':'BQ','id':109},
                          {'letter':'BV','id':110},
                          {'letter':'BT','id':111},
                          {'letter':'BK','id':112},
                          {'letter':'BP','id':113},
                          {'letter':'BS','id':114},
                          {'letter':'BW','id':115},
                          {'letter':'BX','id':116},
                          {'letter':'BY','id':117},
                          {'letter':'CB','id':118},
                          {'letter':'BZ','id':119},
                          {'letter':'CG','id':120},
                          {'letter':'CI','id':121},
                          {'letter':'CC','id':122},
                          {'letter':'CD','id':123},
                          {'letter':'CE','id':124}
                          ]
    elif id_diplomado == '3':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Directic.xlsx')
        ws = wb.get_sheet_by_name('DirecTIC')
        diplomado = Diplomado.objects.get(id = 3)
        i = 6
        dict_productos = [{'letter':'Y','id':127},
                          {'letter':'Z','id':128},
                          {'letter':'AF','id':129},
                          {'letter':'AB','id':130},
                          {'letter':'AD','id':131},
                          {'letter':'AE','id':132},
                          {'letter':'AI','id':133},
                          {'letter':'AG','id':134},
                          {'letter':'AN','id':135},
                          {'letter':'AL','id':136},
                          {'letter':'AR','id':137},
                          {'letter':'AQ','id':138},
                          {'letter':'AW','id':139},
                          {'letter':'AU','id':140},
                          {'letter':'AF','id':141},
                          {'letter':'AJ','id':142},
                          {'letter':'AO','id':143},
                          {'letter':'AS','id':144},
                          {'letter':'AX','id':145},
                          {'letter':'AZ','id':146},
                          {'letter':'AY','id':147},
                          {'letter':'BE','id':148},
                          {'letter':'BD','id':149},
                          {'letter':'BH','id':150},
                          {'letter':'BG','id':151},
                          {'letter':'BA','id':152},
                          {'letter':'BF','id':153},
                          {'letter':'BI','id':154},
                          {'letter':'BL','id':155},
                          {'letter':'BJ','id':156},
                          {'letter':'BN','id':157},
                          {'letter':'BO','id':158},
                          {'letter':'BT','id':159},
                          {'letter':'BR','id':160},
                          {'letter':'BX','id':161},
                          {'letter':'BW','id':162},
                          {'letter':'BM','id':163},
                          {'letter':'BP','id':164},
                          {'letter':'BU','id':165},
                          {'letter':'BY','id':166},
                          {'letter':'CA','id':167},
                          {'letter':'CB','id':168},
                          {'letter':'CE','id':169},
                          {'letter':'CD','id':170},
                          {'letter':'CC','id':171},
                          {'letter':'CF','id':172}
                          ]
    else:
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz EscuelaTIC.xlsx')
        ws = wb.get_sheet_by_name('EscuelaTIC')
        diplomado = Diplomado.objects.get(id = 4)
        i = 10
        dict_productos = [{'letter':'Z','id':221},
                          {'letter':'AD','id':233},
                          {'letter':'AI','id':224},
                          {'letter':'AP','id':228},
        ]

    number = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='right',vertical='center',wrap_text=False),
                   number_format='0',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
             )

    text = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='left',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                 )


    validado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FF00B050',end_color='FF00B050')
                 )

    enviado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FFFFC000',end_color='FFFFC000')
                 )

    cargado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                 )


    rechazado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FFFF0000',end_color='FFFF0000')
                 )

    for beneficiario in Beneficiario.objects.filter(diplomado = diplomado).order_by('formador'):
        ws.cell('A'+str(i)).value = beneficiario.region.nombre.upper()
        ws.cell('A'+str(i)).style = text

        ws.cell('B'+str(i)).value = beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text
        ws.cell('B'+str(i)).style = text

        ws.cell('C'+str(i)).value = beneficiario.radicado.secretaria.nombre.upper() if beneficiario.radicado != None else beneficiario.secretaria_text
        ws.cell('C'+str(i)).style = text

        ws.cell('D'+str(i)).value = beneficiario.radicado.numero if beneficiario.radicado != None else ''
        ws.cell('D'+str(i)).style = number

        ws.cell('E'+str(i)).value = beneficiario.dane_ie_text
        ws.cell('E'+str(i)).style = number

        ws.cell('F'+str(i)).value = beneficiario.ie_text
        ws.cell('F'+str(i)).style = text

        ws.cell('G'+str(i)).value = beneficiario.radicado.dane_sede if beneficiario.radicado != None else beneficiario.dane_sede_text
        ws.cell('G'+str(i)).style = number

        ws.cell('H'+str(i)).value = beneficiario.radicado.nombre_sede.upper() if beneficiario.radicado != None else beneficiario.sede_text
        ws.cell('H'+str(i)).style = text

        ws.cell('I'+str(i)).value = beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text
        ws.cell('I'+str(i)).style = text

        ws.cell('J'+str(i)).value = beneficiario.radicado.ubicacion if beneficiario.radicado != None else ''
        ws.cell('J'+str(i)).style = text

        ws.cell('K'+str(i)).value = beneficiario.get_grupo()
        ws.cell('K'+str(i)).style = text

        ws.cell('L'+str(i)).value = beneficiario.formador.get_full_name()
        ws.cell('L'+str(i)).style = text

        ws.cell('M'+str(i)).value = beneficiario.formador.cedula
        ws.cell('M'+str(i)).style = number

        ws.cell('N'+str(i)).value = beneficiario.apellidos
        ws.cell('N'+str(i)).style = text

        ws.cell('O'+str(i)).value = beneficiario.nombres
        ws.cell('O'+str(i)).style = text

        ws.cell('P'+str(i)).value = beneficiario.cedula
        ws.cell('P'+str(i)).style = number

        ws.cell('Q'+str(i)).value = beneficiario.correo
        ws.cell('Q'+str(i)).style = text

        ws.cell('R'+str(i)).value = beneficiario.telefono_fijo
        ws.cell('R'+str(i)).style = text

        ws.cell('S'+str(i)).value = beneficiario.telefono_celular
        ws.cell('S'+str(i)).style = text

        ws.cell('T'+str(i)).value = beneficiario.area.nombre.upper() if beneficiario.area != None else ''
        ws.cell('T'+str(i)).style = text

        ws.cell('U'+str(i)).value = beneficiario.grado.nombre.upper() if beneficiario.grado != None else ''
        ws.cell('U'+str(i)).style = text

        ws.cell('V'+str(i)).value = ''
        ws.cell('V'+str(i)).style = text

        ws.cell('W'+str(i)).value = beneficiario.genero
        ws.cell('W'+str(i)).style = text

        ws.cell('X'+str(i)).value = beneficiario.estado
        ws.cell('X'+str(i)).style = text

        ws.cell('Y'+str(i)).value = ''
        ws.cell('Y'+str(i)).style = text


        for producto in dict_productos:
            if producto['id'] == int(id_actividad):
                entregable = Entregable.objects.get(id = producto['id'])

                evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable)
                evidencias_validado = Evidencia.objects.filter(beneficiarios_validados = beneficiario,entregable = entregable)

                q = Q(evidencias__id__in = evidencias_cargado.values_list('id',flat=True)) | \
                    Q(evidencias__id__in = evidencias_validado.values_list('id',flat=True))

                reds = Red.objects.filter(q)

                if evidencias_validado.count() > 0:
                    red = Red.objects.get(evidencias__id = evidencias_validado[0].id)
                    ws.cell(producto['letter']+str(i)).value = 'SIC-' + str(evidencias_validado[0].id)
                    #ws.cell(producto['letter']+str(i)).style = validado
                    #ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + evidencias_validado[0].get_archivo_url()
                    #ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(evidencias_validado[0].id),'SICAN')


                elif evidencias_validado.count() == 0 and evidencias_cargado.count() > 0:
                    ultima_evidencia_cargada = evidencias_cargado[len(evidencias_cargado)-1]

                    try:
                        red = reds.get(evidencias__id = ultima_evidencia_cargada.id)
                    except:
                        # si ningun red contiene la evidencia
                        ws.cell(producto['letter']+str(i)).value = 'SIC-' + str(ultima_evidencia_cargada.id)
                        #ws.cell(producto['letter']+str(i)).style = cargado
                        #ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()

                    else:
                        # si hay un red que contiene la evidencia
                        if red.retroalimentacion:
                            # si el red fue retroalimentado

                            evidencias_rechazado = Evidencia.objects.filter(beneficiarios_rechazados__beneficiario_rechazo = beneficiario,
                                                                            beneficiarios_rechazados__red_id = red.id,
                                                                            beneficiarios_rechazados__evidencia_id = ultima_evidencia_cargada.id)

                            if evidencias_rechazado > 0:
                                try:
                                    causa = evidencias_rechazado[0].beneficiarios_rechazados.get(beneficiario_rechazo = beneficiario).observacion
                                except:
                                    causa = ''
                                ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                                #ws.cell(producto['letter']+str(i)).style = rechazado
                                #ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()
                                #ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(ultima_evidencia_cargada.id) + ':\n' + causa,'SICAN')
                            else:
                                pass


                        else:
                            #si el red no ha sido retroalimentado
                            ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                            #ws.cell(producto['letter']+str(i)).style = enviado
                            #ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()
                            #ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(ultima_evidencia_cargada.id),'SICAN')


                else:
                    ws.cell(producto['letter']+str(i)).style = text
        i += 1

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def aprobados_niveles(email):
    usuario = User.objects.get(email=email)
    nombre = "Aprobación de listados de asistencia por nivel"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    output = StringIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/PROGRESO_DESEMBOLSO.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')

    dict_productos_t1 = [{'letter':'B','aprueba':3,'id':[9,12,14,17]},
                        {'letter':'D','aprueba':3,'id':[27,30,33,36]},
                        {'letter':'F','aprueba':3,'id':[46,49,52,55]},
                        {'letter':'H','aprueba':2,'id':[63,66]}
                        ]

    dict_productos_t2 = [{'letter':'B','aprueba':3,'id':[73,74,76,78]},
                         {'letter':'D','aprueba':3,'id':[89,92,94,95]},
                         {'letter':'F','aprueba':3,'id':[104,106,108,110]},
                         {'letter':'H','aprueba':2,'id':[118,120]},
                          ]

    dict_productos_t3 = [{'letter':'B','aprueba':5,'id':[128,131,133,135,137,139]},
                         {'letter':'D','aprueba':2,'id':[146,148,150]},
                         {'letter':'F','aprueba':3,'id':[155,157,159,161]},
                         {'letter':'H','aprueba':2,'id':[167,169]},
                        ]

    dict_productos_t4 = [ {'letter':'B','aprueba':2,'id':[224,228]},
        ]

    for producto in dict_productos_t1:

        i = 5
        for region in Region.objects.filter(id__in=[1,2]):

            cantidad = 0

            ws.cell('A' + str(i)).value = region.nombre.upper()

            evidencias = list(Evidencia.objects.filter(beneficiarios_cargados__region__id = region.id, entregable__id__in = producto['id']).values_list('beneficiarios_validados',flat=True).distinct())

            parciales = []

            for item in producto['id']:
                parciales.append(list(Evidencia.objects.filter(beneficiarios_cargados__region__id = region.id, entregable__id = item).values_list('beneficiarios_validados',flat=True).distinct()))

            for evidencia in evidencias:
                j = 0
                for parcial in parciales:
                    if evidencia in parcial:
                        j += 1
                if j >= producto['aprueba']:
                    cantidad += 1

            ws.cell(producto['letter'] + str(i)).value = cantidad
            i += 1

    for producto in dict_productos_t2:
        i = 16
        for region in Region.objects.filter(id__in=[1,2]):

            cantidad = 0

            ws.cell('A' + str(i)).value = region.nombre.upper()

            evidencias = list(Evidencia.objects.filter(beneficiarios_cargados__region__id = region.id, entregable__id__in = producto['id']).values_list('beneficiarios_validados',flat=True).distinct())

            parciales = []

            for item in producto['id']:
                parciales.append(list(Evidencia.objects.filter(beneficiarios_cargados__region__id = region.id, entregable__id = item).values_list('beneficiarios_validados',flat=True).distinct()))

            for evidencia in evidencias:
                j = 0
                for parcial in parciales:
                    if evidencia in parcial:
                        j += 1
                if j >= producto['aprueba']:
                    cantidad += 1

            ws.cell(producto['letter'] + str(i)).value = cantidad
            i += 1

    for producto in dict_productos_t3:
        i = 27
        for region in Region.objects.filter(id__in=[1,2]):

            cantidad = 0

            ws.cell('A' + str(i)).value = region.nombre.upper()

            evidencias = list(Evidencia.objects.filter(beneficiarios_cargados__region__id = region.id, entregable__id__in = producto['id']).values_list('beneficiarios_validados',flat=True).distinct())

            parciales = []

            for item in producto['id']:
                parciales.append(list(Evidencia.objects.filter(beneficiarios_cargados__region__id = region.id, entregable__id = item).values_list('beneficiarios_validados',flat=True).distinct()))

            for evidencia in evidencias:
                j = 0
                for parcial in parciales:
                    if evidencia in parcial:
                        j += 1
                if j >= producto['aprueba']:
                    cantidad += 1

            ws.cell(producto['letter'] + str(i)).value = cantidad
            i += 1


    for producto in dict_productos_t4:
        i = 38
        for region in Region.objects.filter(id__in=[1,2]):

            cantidad = 0

            ws.cell('A' + str(i)).value = region.nombre.upper()

            evidencias = list(Evidencia.objects.filter(beneficiarios_cargados__region__id = region.id, entregable__id__in = producto['id']).values_list('beneficiarios_validados',flat=True).distinct())

            parciales = []

            for item in producto['id']:
                parciales.append(list(Evidencia.objects.filter(beneficiarios_cargados__region__id = region.id, entregable__id = item).values_list('beneficiarios_validados',flat=True).distinct()))

            for evidencia in evidencias:
                j = 0
                for parcial in parciales:
                    if evidencia in parcial:
                        j += 1
                if j >= producto['aprueba']:
                    cantidad += 1

            ws.cell(producto['letter'] + str(i)).value = cantidad
            i += 1

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def actividades_virtuales(id_red):

    red = Red.objects.get(id = id_red)
    output = StringIO()

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()

    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/RED VIRTUAL.xlsx')
    ws_innovatic = wb.get_sheet_by_name('InnovaTIC')
    inicia_innovatic = 6
    ws_tecnotic = wb.get_sheet_by_name('TecnoTIC')
    inicia_tecnotic = 6
    ws_directic = wb.get_sheet_by_name('DirecTIC')
    inicia_directic = 6


    ids_innovatic = [{'id':20,'letter':'O'},
                     {'id':21,'letter':'Q'},
                     {'id':22,'letter':'R'},
                     {'id':15,'letter':'T'},
                     {'id':16,'letter':'U'},
                     {'id':23,'letter':'V'},
                     {'id':28,'letter':'Y'},
                     {'id':40,'letter':'Z'},
                     {'id':31,'letter':'AB'},
                     {'id':34,'letter':'AD'},
                     {'id':35,'letter':'AE'},
                     {'id':58,'letter':'AH'},
                     {'id':59,'letter':'AJ'},
                     {'id':60,'letter':'AL'},
                     {'id':64,'letter':'AO'},
                     {'id':67,'letter':'AQ'}]

    ids_tecnotic = [{'id':75,'letter':'O'},
                    {'id':77,'letter':'R'},
                    {'id':84,'letter':'S'},
                    {'id':85,'letter':'T'},
                    {'id':97,'letter':'W'},
                    {'id':98,'letter':'X'},
                    {'id':93,'letter':'Y'},
                    {'id':99,'letter':'AA'},
                    {'id':100,'letter':'AC'},
                    {'id':112,'letter':'AF'},
                    {'id':109,'letter':'AH'},
                    {'id':119,'letter':'AK'},
                    {'id':124,'letter':'AL'},
                    {'id':121,'letter':'AO'}]

    ids_directic = [{'id':132,'letter':'P'},
               {'id':134,'letter':'Q'},
               {'id':142,'letter':'S'},
               {'id':143,'letter':'T'},
               {'id':144,'letter':'V'},
               {'id':140,'letter':'X'},
               {'id':147,'letter':'Z'},
               {'id':152,'letter':'AB'},
               {'id':149,'letter':'AD'},
               {'id':151,'letter':'AE'},
               {'id':156,'letter':'AG'},
               {'id':164,'letter':'AJ'},
               {'id':165,'letter':'AK'},
               {'id':162,'letter':'AM'},
               {'id':166,'letter':'AO'},
               {'id':171,'letter':'AQ'},
               {'id':171,'letter':'AR'}]


    evidencias_total = Evidencia.objects.filter(red_id = id_red)
    beneficiarios_id = evidencias_total.exclude(beneficiarios_cargados = None).values_list('beneficiarios_cargados__id',flat=True).distinct()


    i = 0 + inicia
    for beneficiario_id in beneficiarios_id:
        beneficiario = Beneficiario.objects.get(id = beneficiario_id)
        evidencias = evidencias_total.filter(beneficiarios_cargados__id = beneficiario_id)


        ws.cell('A'+str(i)).value = i - inicia + 1
        ws.cell('B'+str(i)).value = beneficiario.region.nombre.upper()
        ws.cell('C'+str(i)).value = beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text.upper()
        ws.cell('D'+str(i)).value = beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text.upper()
        ws.cell('E'+str(i)).value = beneficiario.ruta.upper() + '-' + beneficiario.grupo.nombre
        ws.cell('F'+str(i)).value = beneficiario.formador.get_full_name().upper()
        ws.cell('G'+str(i)).value = beneficiario.formador.cedula
        ws.cell('G'+str(i)).number_format = '0'
        ws.cell('H'+str(i)).value = beneficiario.nombres.upper()
        ws.cell('I'+str(i)).value = beneficiario.apellidos.upper()
        ws.cell('J'+str(i)).value = beneficiario.cedula

        ws.cell('J'+str(i)).number_format = '0'
        ws.cell('K'+str(i)).value = 'SICAN'
        ws.cell('L'+str(i)).value = 'I'

        for id in ids:
            evidencia = evidencias.filter(entregable__id = id['id'])
            if evidencia.count() == 1:
                if evidencia[0].subsanacion:
                    ws.cell('L'+str(i)).value = 'S'
                ws.cell( id['letter'] + str(i)).value = 'SIC-' + str(evidencia[0].id)
                ws.cell( id['letter'] + str(i)).hyperlink = 'https://sican.asoandes.org' + evidencia[0].get_archivo_url()

        i += 1


    wb.save(output)
    filename = 'RED-' + unicode(red.id) + '-'+ red.region.nombre +'.xlsx'
    red.archivo.save(filename,File(output))

    return "Generado RED-" + str(id_red)

@app.task
def matriz_chequeo_compilada(email):
    usuario = User.objects.get(email=email)
    nombre = 'Compilado listados de asistencia SICAN'

    proceso = "REV-INF05"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []

    dict_productos_innovatic = [{'letter':'I','id':8,'red':'J','aprobacion':'K'},
                                {'letter':'L','id':9,'red':'M','aprobacion':'N'},
                                {'letter':'O','id':12,'red':'P','aprobacion':'Q'},
                                {'letter':'R','id':14,'red':'S','aprobacion':'T'},
                                {'letter':'U','id':17,'red':'V','aprobacion':'W'},
                                {'letter':'X','id':27,'red':'Y','aprobacion':'Z'},
                                {'letter':'AA','id':30,'red':'AB','aprobacion':'AC'},
                                {'letter':'AD','id':33,'red':'AE','aprobacion':'AF'},
                                {'letter':'AG','id':36,'red':'AH','aprobacion':'AI'},
                                {'letter':'AJ','id':46,'red':'AK','aprobacion':'AL'},
                                {'letter':'AM','id':49,'red':'AN','aprobacion':'AO'},
                                {'letter':'AP','id':52,'red':'AQ','aprobacion':'AR'},
                                {'letter':'AS','id':55,'red':'AT','aprobacion':'AU'},
                                {'letter':'AV','id':63,'red':'AW','aprobacion':'AX'},
                                {'letter':'AY','id':66,'red':'AZ','aprobacion':'BA'},
                                ]

    dict_productos_tecnotic = [   {'letter':'I','id':72,'red':'J','aprobacion':'K'},
                                  {'letter':'L','id':73,'red':'M','aprobacion':'N'},
                                  {'letter':'O','id':74,'red':'P','aprobacion':'Q'},
                                  {'letter':'R','id':76,'red':'S','aprobacion':'T'},
                                  {'letter':'U','id':78,'red':'V','aprobacion':'W'},
                                  {'letter':'X','id':89,'red':'Y','aprobacion':'Z'},
                                  {'letter':'AA','id':92,'red':'AB','aprobacion':'AC'},
                                  {'letter':'AD','id':94,'red':'AE','aprobacion':'AF'},
                                  {'letter':'AG','id':95,'red':'AH','aprobacion':'AI'},
                                  {'letter':'AJ','id':104,'red':'AK','aprobacion':'AL'},
                                  {'letter':'AM','id':106,'red':'AN','aprobacion':'AO'},
                                  {'letter':'AP','id':108,'red':'AQ','aprobacion':'AR'},
                                  {'letter':'AS','id':110,'red':'AT','aprobacion':'AU'},
                                  {'letter':'AV','id':118,'red':'AW','aprobacion':'AX'},
                                  {'letter':'AY','id':120,'red':'AZ','aprobacion':'BA'},
                                  ]


    dict_productos_directic = [   {'letter':'I','id':127,'red':'J','aprobacion':'K'},
                                  {'letter':'L','id':128,'red':'M','aprobacion':'N'},
                                  {'letter':'O','id':131,'red':'P','aprobacion':'Q'},
                                  {'letter':'R','id':133,'red':'S','aprobacion':'T'},
                                  {'letter':'U','id':135,'red':'V','aprobacion':'W'},
                                  {'letter':'X','id':137,'red':'Y','aprobacion':'Z'},
                                  {'letter':'AA','id':139,'red':'AB','aprobacion':'AC'},
                                  {'letter':'AD','id':146,'red':'AE','aprobacion':'AF'},
                                  {'letter':'AG','id':148,'red':'AH','aprobacion':'AI'},
                                  {'letter':'AJ','id':150,'red':'AK','aprobacion':'AL'},
                                  {'letter':'AM','id':155,'red':'AN','aprobacion':'AO'},
                                  {'letter':'AP','id':157,'red':'AQ','aprobacion':'AR'},
                                  {'letter':'AS','id':159,'red':'AT','aprobacion':'AU'},
                                  {'letter':'AV','id':161,'red':'AW','aprobacion':'AX'},
                                  {'letter':'AY','id':167,'red':'AZ','aprobacion':'BA'},
                                  {'letter':'BB','id':169,'red':'BC','aprobacion':'BD'},
                                  ]



    dict_productos_escuelatic = [{'letter':'I','id':224,'red':'J','aprobacion':'K'},
                                 {'letter':'L','id':228,'red':'M','aprobacion':'N'},
                                ]


    wb = xlsxwriter.Workbook(output)
    ws_innovatic = wb.add_worksheet('INNOVATIC')
    ws_tecnotic = wb.add_worksheet('TECNOTIC')
    ws_directic = wb.add_worksheet('DIRECTIC')
    ws_escuelatic = wb.add_worksheet('ESCUELATIC')

    text = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False})

    number = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'right', 'valign':'vcenter', 'text_wrap':False,'num_format':'0'})

    validado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#00B050'})

    enviado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#FFC000'})

    cargado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False})

    rechazado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#FF0000'})

    gris = wb.add_format({'font_name':'Arial Narrow','border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#808080', 'font_color':'#FFFFFF'})

    verde = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#008000', 'font_color':'#FFFFFF'})

    azul = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#000080', 'font_color':'#FFFFFF'})

    morado = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#7030A0', 'font_color':'#FFFFFF'})

    agua_marina = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#00B0F0', 'font_color':'#FFFFFF'})



    for id_diplomado in [1]:

        if id_diplomado == 1:
            ws = ws_innovatic
            dict_productos = dict_productos_innovatic

            ws.write('I1','ACTA COMPROMISO',azul)
            ws.set_column('I:I',12)

            ws.write('J1','RED',azul)
            ws.set_column('J:J',12)

            ws.write('K1','APROBACIÓN',azul)
            ws.set_column('K:K',12)

            j = 11

            for k in range(len(dict_productos_innovatic)-1):
                ws.write(0,j,'ASISTENCIA ' + str(k+1),verde)
                ws.set_column(j,j,14)

                ws.write(0,j+1,'RED',verde)
                ws.set_column(j+1,j+1,14)

                ws.write(0,j+2,'APROBACIÓN',verde)
                ws.set_column(j+2,j+2,14)

                j += 3

            fin = j

            ws.write(0,fin,'APROBACIÓN N1',morado)
            ws.set_column(0,fin,20)

            ws.write(0,fin+1,'APROBACIÓN N2',morado)
            ws.set_column(0,fin+1,20)

            ws.write(0,fin+2,'APROBACIÓN N3',morado)
            ws.set_column(0,fin+2,20)

            ws.write(0,fin+3,'APROBACIÓN N4',morado)
            ws.set_column(0,fin+3,20)


            ws.write(0,fin+4,'APROBACIÓN N1 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+4,20)

            ws.write(0,fin+5,'APROBACIÓN N2 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+5,20)

            ws.write(0,fin+6,'APROBACIÓN N3 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+6,20)

            ws.write(0,fin+7,'APROBACIÓN N4 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+7,20)


        elif id_diplomado == 2:
            ws = ws_tecnotic
            dict_productos = dict_productos_tecnotic

            ws.write('I1','ACTA COMPROMISO',azul)
            ws.set_column('I:I',12)

            ws.write('J1','RED',azul)
            ws.set_column('J:J',12)

            ws.write('K1','APROBACIÓN',azul)
            ws.set_column('K:K',12)

            j = 11

            for k in range(len(dict_productos_tecnotic)-1):
                ws.write(0,j,'ASISTENCIA ' + str(k+1),verde)
                ws.set_column(j,j,14)

                ws.write(0,j+1,'RED',verde)
                ws.set_column(j+1,j+1,14)

                ws.write(0,j+2,'APROBACIÓN',verde)
                ws.set_column(j+2,j+2,14)

                j += 3

            fin = j

            ws.write(0,fin,'APROBACIÓN N1',morado)
            ws.set_column(0,fin,20)

            ws.write(0,fin+1,'APROBACIÓN N2',morado)
            ws.set_column(0,fin+1,20)

            ws.write(0,fin+2,'APROBACIÓN N3',morado)
            ws.set_column(0,fin+2,20)

            ws.write(0,fin+3,'APROBACIÓN N4',morado)
            ws.set_column(0,fin+3,20)


            ws.write(0,fin+4,'APROBACIÓN N1 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+4,20)

            ws.write(0,fin+5,'APROBACIÓN N2 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+5,20)

            ws.write(0,fin+6,'APROBACIÓN N3 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+6,20)

            ws.write(0,fin+7,'APROBACIÓN N4 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+7,20)

        elif id_diplomado == 3:
            ws = ws_directic
            dict_productos = dict_productos_directic

            ws.write('I1','ACTA COMPROMISO',azul)
            ws.set_column('I:I',12)

            ws.write('J1','RED',azul)
            ws.set_column('J:J',12)

            ws.write('K1','APROBACIÓN',azul)
            ws.set_column('K:K',12)

            j = 11

            for k in range(len(dict_productos_directic)-1):
                ws.write(0,j,'ASISTENCIA ' + str(k+1),verde)
                ws.set_column(j,j,14)

                ws.write(0,j+1,'RED',verde)
                ws.set_column(j+1,j+1,14)

                ws.write(0,j+2,'APROBACIÓN',verde)
                ws.set_column(j+2,j+2,14)

                j += 3

            fin = j

            ws.write(0,fin,'APROBACIÓN N1',morado)
            ws.set_column(0,fin,20)

            ws.write(0,fin+1,'APROBACIÓN N2',morado)
            ws.set_column(0,fin+1,20)

            ws.write(0,fin+2,'APROBACIÓN N3',morado)
            ws.set_column(0,fin+2,20)

            ws.write(0,fin+3,'APROBACIÓN N4',morado)
            ws.set_column(0,fin+3,20)


            ws.write(0,fin+4,'APROBACIÓN N1 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+4,20)

            ws.write(0,fin+5,'APROBACIÓN N2 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+5,20)

            ws.write(0,fin+6,'APROBACIÓN N3 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+6,20)

            ws.write(0,fin+7,'APROBACIÓN N4 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+7,20)

        else:
            ws = ws_escuelatic
            dict_productos = dict_productos_escuelatic

            ws.write('I1','LISTADO DE ASISTENCIA N2 S1',azul)
            ws.set_column('I:I',12)

            ws.write('J1','RED',azul)
            ws.set_column('J:J',12)

            ws.write('K1','APROBACIÓN',azul)
            ws.set_column('K:K',12)


            ws.write('L1','LISTADO DE ASISTENCIA N2 S2',azul)
            ws.set_column('L:L',12)

            ws.write('M1','RED',azul)
            ws.set_column('M:M',12)

            ws.write('N1','APROBACIÓN',azul)
            ws.set_column('N:N',12)

            fin = 14

            ws.write('O1','APROBACIÓN N2',morado)
            ws.set_column('O:O',20)



        ws.write('A1','REGIÓN',gris)
        ws.set_column('A:A',11)
        ws.set_row(0,90)

        ws.write('B1','DEPARTAMENTO',gris)
        ws.set_column('B:B',23)

        ws.write('C1','MUNICIPIO',gris)
        ws.set_column('C:C',18)

        ws.write('D1','NOMBRE DEL FORMADOR',gris)
        ws.set_column('D:D',36)

        ws.write('E1','NUMERO DE CEDULA DEL FORMADOR',gris)
        ws.set_column('E:E',13)

        ws.write('F1','APELLIDOS DEL BENEFICIARIO',verde)
        ws.set_column('F:F',25)

        ws.write('G1','NOMBRES DEL BENEFICIARIO',verde)
        ws.set_column('G:G',25)

        ws.write('H1','NUMERO DE CEDULA DEL BENEFICIARIO',verde)
        ws.set_column('H:H',14)

        i = 2

        for beneficiario in Beneficiario.objects.filter(diplomado__id = id_diplomado).order_by('formador'):
            ws.write('A'+str(i), beneficiario.region.nombre.upper(),text)

            ws.write('B'+str(i), beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text,text)

            ws.write('C'+str(i), beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text,text)

            ws.write('D'+str(i), beneficiario.formador.get_full_name(),text)

            ws.write('E'+str(i), beneficiario.formador.cedula,number)

            ws.write('F'+str(i), beneficiario.apellidos,text)

            ws.write('G'+str(i), beneficiario.nombres,text)

            ws.write('H'+str(i), beneficiario.cedula,number)


            for producto in dict_productos:
                entregable = Entregable.objects.get(id = producto['id'])

                evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable)
                evidencias_validado = Evidencia.objects.filter(beneficiarios_validados = beneficiario,entregable = entregable)


                if evidencias_validado.count() > 0:
                    ws.write( producto['letter'] + str(i) , 'SIC-' + str(evidencias_validado[0].id), validado )
                    ws.write( producto['red'] + str(i) , 'RED-' + str(evidencias_validado[0].red_id), text )
                    ws.write( producto['aprobacion'] + str(i) , 'SI', text )


                elif evidencias_validado.count() == 0 and evidencias_cargado.count() > 0:
                    ultima_evidencia_cargada = evidencias_cargado[len(evidencias_cargado)-1]

                    try:
                        red = Red.objects.get(id = ultima_evidencia_cargada.red_id)
                    except:
                        # si ningun red contiene la evidencia
                        ws.write( producto['letter'] + str(i), 'SIC-' + str(ultima_evidencia_cargada.id), cargado )
                        if ultima_evidencia_cargada.red_id != None:
                            ws.write( producto['red'] + str(i) , 'RED-' + str(ultima_evidencia_cargada.red_id), text )

                    else:
                        # si hay un red que contiene la evidencia
                        if red.retroalimentacion:
                            # si el red fue retroalimentado

                            evidencias_rechazado = Evidencia.objects.filter(beneficiarios_rechazados__beneficiario_rechazo = beneficiario,
                                                                            beneficiarios_rechazados__red_id = red.id,
                                                                            beneficiarios_rechazados__evidencia_id = ultima_evidencia_cargada.id)

                            if evidencias_rechazado > 0:
                                ws.write( producto['letter'] + str(i) , 'SIC-' + str(ultima_evidencia_cargada.id), rechazado )
                                if ultima_evidencia_cargada.red_id != None:
                                    ws.write( producto['red'] + str(i) , 'RED-' + str(ultima_evidencia_cargada.red_id), text )
                            else:
                                pass


                        else:
                            #si el red no ha sido retroalimentado
                            ws.write( producto['letter'] + str(i) , 'SIC-' + str(ultima_evidencia_cargada.id), enviado )
                            if ultima_evidencia_cargada.red_id != None:
                                    ws.write( producto['red'] + str(i) , 'RED-' + str(ultima_evidencia_cargada.red_id), text )


                else:
                    pass

                if id_diplomado == 1:
                    ws.write_formula(i-1,fin,'=IF(IF(N'+ str(i) +'="SI",1,0)+IF(Q'+ str(i) +'="SI",1,0)+IF(T'+ str(i) +'="SI",1,0)+IF(W' + str(i) + '="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+1,'=IF(IF(Z'+ str(i) +'="SI",1,0)+IF(AC'+ str(i) +'="SI",1,0)+IF(AF'+ str(i) +'="SI",1,0)+IF(AI'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+2,'=IF(IF(AL'+ str(i) +'="SI",1,0)+IF(AO'+ str(i) +'="SI",1,0)+IF(AR'+ str(i) +'="SI",1,0)+IF(AU'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+3,'=IF(IF(AX'+ str(i) +'="SI",1,0)+IF(BA'+ str(i) +'="SI",1,0)>=1,"SI","NO")',text)


                    ws.write_formula(i-1,fin+4,'=IF($K'+ str(i) +'="SI",BB'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+5,'=IF($K'+ str(i) +'="SI",BC'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+6,'=IF($K'+ str(i) +'="SI",BD'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+7,'=IF($K'+ str(i) +'="SI",BE'+ str(i) +',"NO ACTA")',text)


                if id_diplomado == 2:
                    ws.write_formula(i-1,fin,'=IF(IF(N'+ str(i) +'="SI",1,0)+IF(Q'+ str(i) +'="SI",1,0)+IF(T'+ str(i) +'="SI",1,0)+IF(W' + str(i) + '="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+1,'=IF(IF(Z'+ str(i) +'="SI",1,0)+IF(AC'+ str(i) +'="SI",1,0)+IF(AF'+ str(i) +'="SI",1,0)+IF(AI'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+2,'=IF(IF(AL'+ str(i) +'="SI",1,0)+IF(AO'+ str(i) +'="SI",1,0)+IF(AR'+ str(i) +'="SI",1,0)+IF(AU'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+3,'=IF(IF(AX'+ str(i) +'="SI",1,0)+IF(BA'+ str(i) +'="SI",1,0)>=1,"SI","NO")',text)


                    ws.write_formula(i-1,fin+4,'=IF($K'+ str(i) +'="SI",BB'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+5,'=IF($K'+ str(i) +'="SI",BC'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+6,'=IF($K'+ str(i) +'="SI",BD'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+7,'=IF($K'+ str(i) +'="SI",BE'+ str(i) +',"NO ACTA")',text)

                if id_diplomado == 3:
                    ws.write_formula(i-1,fin,'=IF(IF(N'+ str(i) +'="SI",1,0)+IF(Q'+ str(i) +'="SI",1,0)+IF(T'+ str(i) +'="SI",1,0)+IF(W'+ str(i) +'="SI",1,0)+IF(Z'+ str(i) +'="SI",1,0)+IF(AC'+ str(i) +'="SI",1,0)>=4,"SI","NO")',text)
                    ws.write_formula(i-1,fin+1,'=IF(IF(AF'+ str(i) +'="SI",1,0)+IF(AI'+ str(i) +'="SI",1,0)+IF(AL'+ str(i) +'="SI",1,0)>=2,"SI","NO")',text)
                    ws.write_formula(i-1,fin+2,'=IF(IF(AO'+ str(i) +'="IF",1,0)+IF(AR'+ str(i) +'="SI",1,0)+IF(AU'+ str(i) +'="SI",1,0)+IF(AX'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+3,'=IF(IF(BA'+ str(i) +'="SI",1,0)+IF(BD'+ str(i) +'="SI",1,0)>=1,"SI","NO")',text)


                    ws.write_formula(i-1,fin+4,'=IF($K'+ str(i) +'="SI",BE'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+5,'=IF($K'+ str(i) +'="SI",BF'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+6,'=IF($K'+ str(i) +'="SI",BG'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+7,'=IF($K'+ str(i) +'="SI",BH'+ str(i) +',"NO ACTA")',text)

                if id_diplomado == 4:
                    ws.write_formula(i-1,fin,'=IF(IF(K'+ str(i) +'="SI",1,0)+IF(N'+ str(i) +'="SI",1,0)=2,"SI","NO")',text)

            i += 1

    wb.close()

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_chequeo_virtual_compilada(email):
    usuario = User.objects.get(email=email)
    nombre = 'Compilado evidencias virtuales SICAN'

    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []

    dict_productos_innovatic = [{'letter':'I','id':20,'nombre':'N1S1 - Hoja de ruta'},
                                {'letter':'J','id':21,'nombre':'N1S2 - Guía: construyendo lecciones de innovación educativa'},
                                {'letter':'K','id':22,'nombre':'N1S2 - Taller práctico: 10 claves para la implementación de tendencias y enfoques innovadores'},
                                {'letter':'L','id':15,'nombre':'N1S3 - Documento: Decálogo de un Proyecto Innovador'},
                                {'letter':'M','id':16,'nombre':'N1S3 - Guía: Potenciando mi Experiencia Educativa de primera fase'},
                                {'letter':'N','id':23,'nombre':'N1S4 - Guía: implementando la Experiencia Educativa Innovadora'},

                                {'letter':'O','id':28,'nombre':'N2S1 - Quiz realizado en Blendspace (Fomato Word)'},
                                {'letter':'P','id':40,'nombre':'N2S1 - Mapa mental'},
                                {'letter':'Q','id':34,'nombre':'N2S3 - Guía: Construyendo mi PLE'},

                                {'letter':'R','id':58,'nombre':'N3S1 - Formato de recopilación de resultados: "Lo aprendido"'},
                                {'letter':'S','id':59,'nombre':'N3S2 - Enlace o imagen interactiva: Ruta de sostenibilidad'},
                                {'letter':'T','id':60,'nombre':'N3S3 - Gráfica del PLE realizada en Mindomo'},

                                {'letter':'U','id':64,'nombre':'N4S1 - Documento de Google Drive 8 puntos para exponer mi PLE + herramienta de presentación, cargada en la Blackboard'},
                                {'letter':'V','id':67,'nombre':'N4S2 - Registro fotográfico: Evento de socialización'},
                                ]

    dict_productos_tecnotic = [{'letter':'I','id':75,'nombre':'N1S2 - Taller práctico 1'},
                                {'letter':'J','id':77,'nombre':'N1S3 - Ficha. Documentación de apps educativas'},
                                {'letter':'K','id':84,'nombre':'N1S3 - Realimentar las apps educativas de mínimo dos compañeros'},
                                {'letter':'L','id':85,'nombre':'N1S3 - Guía: implementación de apps educativas en prácticas de aula'},

                                {'letter':'M','id':97,'nombre':'N2S1 - Campaña de expectativa'},
                                {'letter':'N','id':98,'nombre':'N2S2 - Bitácora 1'},
                                {'letter':'O','id':93,'nombre':'N2S2 - Bitácora 2'},
                                {'letter':'P','id':99,'nombre':'N2S2 - Bitácora 3'},
                                {'letter':'Q','id':100,'nombre':'N2S3 - Bitácora 4'},

                                {'letter':'R','id':112,'nombre':'N3S1 - Bitácora 5'},
                                {'letter':'S','id':109,'nombre':'N3S2 - Bitácora 6'},

                                {'letter':'T','id':119,'nombre':'N4S1 - Modelo de presentación del Proyecto Educativo App'},
                                {'letter':'U','id':124,'nombre':'N4S1 - Realimentar la participación de un compañer'},
                                {'letter':'V','id':121,'nombre':'N4S2 - Registro fotográfico del evento'},
                                ]


    dict_productos_directic = [{'letter':'I','id':132,'nombre':'N1S2 - Documento: presentación del capítulo  asignado a cada equipo'},
                                {'letter':'J','id':134,'nombre':'N1S3 - El Gestor del Plan Estratégico TIC en su primer componente'},
                                {'letter':'K','id':142,'nombre':'N1S3 -  Gestor del Plan Estratégico TIC con los dos componentes de laCaracterizacióndiligenciados'},
                                {'letter':'L','id':143,'nombre':'N1S4 -  Gestor del Plan Estratégico TIC diligenciado en el componente Indicadores Educativos'},
                                {'letter':'M','id':144,'nombre':'N1S5 -  Gestor de Plan Estratégico de TIC: componente, valoración de la gestión institucional'},
                                {'letter':'N','id':140,'nombre':'N1S6 - Gestor del Plan Estratégico TIC componente Matriz FODA'},

                                {'letter':'O','id':147,'nombre':'N2S1 - Gestor del Plan Estratégico TIC en el componente: Responsables y plazos'},
                                {'letter':'P','id':152,'nombre':'N2S1 - Gestor del Plan Estratégico TIC. Componenete Comité de gestores TIC'},
                                {'letter':'Q','id':149,'nombre':'N2S2 - Gestor del Plan Estratégico TIC. Objetivo general del Plan, objetivos específicos'},
                                {'letter':'R','id':151,'nombre':'N2S3 - Gestor del Plan Estratégico TIC. Etapa: análisis y formulación del Plan'},

                                {'letter':'S','id':156,'nombre':'N3S1 - Gestor del Plan Estratégico TIC, componente Tipos de Cooperación.'},
                                {'letter':'T','id':164,'nombre':'N3S2 - Diligenciar la plantilla del documento compartido en Google Drive denominado: Aliados Estratégicos.'},
                                {'letter':'U','id':165,'nombre':'N3S3 - Gestor Plan Estratégico TIC'},
                                {'letter':'V','id':162,'nombre':'N3S4 - Brochure'},
                                {'letter':'W','id':166,'nombre':'N3S4 -  Audio, video o Guión de la entrevista ejecutada con la en entidad pública o privada'},

                                {'letter':'X','id':171,'nombre':'N4S1 -  Dos fotografías del desarrollo de la socialización del Plan Estratégico TIC'},
                               {'letter':'Y','id':171,'nombre':'N4S2 -  Acta diligenciada y firmada por los integrantes del Comité de Gestores TIC.'},
                                ]





    wb = xlsxwriter.Workbook(output)
    ws_innovatic = wb.add_worksheet('INNOVATIC')
    ws_tecnotic = wb.add_worksheet('TECNOTIC')
    ws_directic = wb.add_worksheet('DIRECTIC')


    text = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False})

    number = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'right', 'valign':'vcenter', 'text_wrap':False,'num_format':'0'})

    validado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#00B050'})

    gris = wb.add_format({'font_name':'Arial Narrow','border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#808080', 'font_color':'#FFFFFF'})


    verde = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#008000', 'font_color':'#FFFFFF'})



    for id_diplomado in [1,2,3]:

        if id_diplomado == 1:
            ws = ws_innovatic
            dict_productos = dict_productos_innovatic

            for producto in dict_productos_innovatic:
                ws.write(producto['letter']+'1',producto['nombre'],verde)
                ws.set_column(producto['letter']+':'+producto['letter'],14)



        elif id_diplomado == 2:
            ws = ws_tecnotic
            dict_productos = dict_productos_tecnotic

            for producto in dict_productos_tecnotic:
                ws.write(producto['letter']+'1',producto['nombre'],verde)
                ws.set_column(producto['letter']+':'+producto['letter'],14)

        elif id_diplomado == 3:
            ws = ws_directic
            dict_productos = dict_productos_directic

            for producto in dict_productos_directic:
                ws.write(producto['letter']+'1',producto['nombre'],verde)
                ws.set_column(producto['letter']+':'+producto['letter'],14)



        ws.write('A1','REGIÓN',gris)
        ws.set_column('A:A',11)
        ws.set_row(0,90)

        ws.write('B1','DEPARTAMENTO',gris)
        ws.set_column('B:B',23)

        ws.write('C1','MUNICIPIO',gris)
        ws.set_column('C:C',18)

        ws.write('D1','NOMBRE DEL FORMADOR',gris)
        ws.set_column('D:D',36)

        ws.write('E1','NUMERO DE CEDULA DEL FORMADOR',gris)
        ws.set_column('E:E',13)

        ws.write('F1','APELLIDOS DEL BENEFICIARIO',verde)
        ws.set_column('F:F',25)

        ws.write('G1','NOMBRES DEL BENEFICIARIO',verde)
        ws.set_column('G:G',25)

        ws.write('H1','NUMERO DE CEDULA DEL BENEFICIARIO',verde)
        ws.set_column('H:H',14)

        i = 2

        for beneficiario in Beneficiario.objects.filter(diplomado__id = id_diplomado).order_by('formador'):
            ws.write('A'+str(i), beneficiario.region.nombre.upper(),text)

            ws.write('B'+str(i), beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text,text)

            ws.write('C'+str(i), beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text,text)

            ws.write('D'+str(i), beneficiario.formador.get_full_name(),text)

            ws.write('E'+str(i), beneficiario.formador.cedula,number)

            ws.write('F'+str(i), beneficiario.apellidos,text)

            ws.write('G'+str(i), beneficiario.nombres,text)

            ws.write('H'+str(i), beneficiario.cedula,number)


            for producto in dict_productos:
                entregable = Entregable.objects.get(id = producto['id'])

                evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable)
                evidencias_validado = Evidencia.objects.filter(beneficiarios_validados = beneficiario,entregable = entregable)

                if evidencias_cargado.count() > 0:
                    ws.write( producto['letter'] + str(i) , 'SIC-' + str(evidencias_cargado[0].id), validado)


            i += 1

    wb.close()

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def compilado_matriz_chequeo():
    matriz_chequeo_compilada.delay('sistemas@asoandes.org')
    return "Reporte diario generado"

@app.task
def reporte_actividades_virtuales(email):
    usuario = User.objects.get(email=email)
    nombre = 'Compilado actividades virtuales SICAN'

    proceso = "REV-INF05"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []

    dict_productos_innovatic =   [{'letter':'AL','id':15},
                                  {'letter':'AN','id':16},
                                  {'letter':'AD','id':20},
                                  {'letter':'AH','id':21},
                                  {'letter':'AJ','id':22},
                                  {'letter':'AO','id':23},

                                  {'letter':'AX','id':28},
                                  {'letter':'BH','id':34},
                                  {'letter':'BA','id':40},

                                  {'letter':'BV','id':58},
                                  {'letter':'CA','id':59},
                                  {'letter':'CE','id':60},

                                  {'letter':'CM','id':64},
                                  {'letter':'CS','id':67},
                                  ]

    dict_productos_tecnotic =    [{'letter':'AE','id':75},
                                  {'letter':'AI','id':77},
                                  {'letter':'AM','id':85},

                                  {'letter':'AW','id':97},
                                  {'letter':'AX','id':98},
                                  {'letter':'AY','id':93},
                                  {'letter':'BA','id':99},
                                  {'letter':'BC','id':100},

                                  {'letter':'BK','id':112},
                                  {'letter':'BQ','id':109},


                                  {'letter':'BZ','id':119},
                                  {'letter':'CE','id':124},
                                  {'letter':'CI','id':121}
                                  ]


    dict_productos_directic =    [  {'letter':'AE','id':132},
                                    {'letter':'AU','id':140},

                                    {'letter':'BG','id':151},

                                    {'letter':'BP','id':164},
                                    {'letter':'BU','id':165},
                                    {'letter':'BW','id':162},
                                    {'letter':'BY','id':166},

                                    {'letter':'CC','id':171},
                                  ]


    wb = xlsxwriter.Workbook(output)
    ws_innovatic = wb.add_worksheet('INNOVATIC')
    ws_tecnotic = wb.add_worksheet('TECNOTIC')
    ws_directic = wb.add_worksheet('DIRECTIC')

    text = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False})

    number = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'right', 'valign':'vcenter', 'text_wrap':False,'num_format':'0'})

    validado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#00B050'})

    enviado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#FFC000'})

    cargado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False})

    rechazado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#FF0000'})

    gris = wb.add_format({'font_name':'Arial Narrow','border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#808080', 'font_color':'#FFFFFF'})

    verde = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#008000', 'font_color':'#FFFFFF'})

    azul = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#000080', 'font_color':'#FFFFFF'})

    morado = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#7030A0', 'font_color':'#FFFFFF'})

    agua_marina = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#00B0F0', 'font_color':'#FFFFFF'})



    for id_diplomado in [1,2,3]:

        if id_diplomado == 1:
            ws = ws_innovatic
            dict_productos = dict_productos_innovatic

            ws.write('I1','ACTA COMPROMISO',azul)
            ws.set_column('I:I',12)

            ws.write('J1','RED',azul)
            ws.set_column('J:J',12)

            ws.write('K1','APROBACIÓN',azul)
            ws.set_column('K:K',12)

            j = 11

            for k in range(len(dict_productos_innovatic)-1):
                ws.write(0,j,'ASISTENCIA ' + str(k+1),verde)
                ws.set_column(j,j,14)

                ws.write(0,j+1,'RED',verde)
                ws.set_column(j+1,j+1,14)

                ws.write(0,j+2,'APROBACIÓN',verde)
                ws.set_column(j+2,j+2,14)

                j += 3

            fin = j

            ws.write(0,fin,'APROBACIÓN N1',morado)
            ws.set_column(0,fin,20)

            ws.write(0,fin+1,'APROBACIÓN N2',morado)
            ws.set_column(0,fin+1,20)

            ws.write(0,fin+2,'APROBACIÓN N3',morado)
            ws.set_column(0,fin+2,20)

            ws.write(0,fin+3,'APROBACIÓN N4',morado)
            ws.set_column(0,fin+3,20)


            ws.write(0,fin+4,'APROBACIÓN N1 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+4,20)

            ws.write(0,fin+5,'APROBACIÓN N2 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+5,20)

            ws.write(0,fin+6,'APROBACIÓN N3 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+6,20)

            ws.write(0,fin+7,'APROBACIÓN N4 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+7,20)


        elif id_diplomado == 2:
            ws = ws_tecnotic
            dict_productos = dict_productos_tecnotic

            ws.write('I1','ACTA COMPROMISO',azul)
            ws.set_column('I:I',12)

            ws.write('J1','RED',azul)
            ws.set_column('J:J',12)

            ws.write('K1','APROBACIÓN',azul)
            ws.set_column('K:K',12)

            j = 11

            for k in range(len(dict_productos_tecnotic)-1):
                ws.write(0,j,'ASISTENCIA ' + str(k+1),verde)
                ws.set_column(j,j,14)

                ws.write(0,j+1,'RED',verde)
                ws.set_column(j+1,j+1,14)

                ws.write(0,j+2,'APROBACIÓN',verde)
                ws.set_column(j+2,j+2,14)

                j += 3

            fin = j

            ws.write(0,fin,'APROBACIÓN N1',morado)
            ws.set_column(0,fin,20)

            ws.write(0,fin+1,'APROBACIÓN N2',morado)
            ws.set_column(0,fin+1,20)

            ws.write(0,fin+2,'APROBACIÓN N3',morado)
            ws.set_column(0,fin+2,20)

            ws.write(0,fin+3,'APROBACIÓN N4',morado)
            ws.set_column(0,fin+3,20)


            ws.write(0,fin+4,'APROBACIÓN N1 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+4,20)

            ws.write(0,fin+5,'APROBACIÓN N2 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+5,20)

            ws.write(0,fin+6,'APROBACIÓN N3 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+6,20)

            ws.write(0,fin+7,'APROBACIÓN N4 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+7,20)

        elif id_diplomado == 3:
            ws = ws_directic
            dict_productos = dict_productos_directic

            ws.write('I1','ACTA COMPROMISO',azul)
            ws.set_column('I:I',12)

            ws.write('J1','RED',azul)
            ws.set_column('J:J',12)

            ws.write('K1','APROBACIÓN',azul)
            ws.set_column('K:K',12)

            j = 11

            for k in range(len(dict_productos_directic)-1):
                ws.write(0,j,'ASISTENCIA ' + str(k+1),verde)
                ws.set_column(j,j,14)

                ws.write(0,j+1,'RED',verde)
                ws.set_column(j+1,j+1,14)

                ws.write(0,j+2,'APROBACIÓN',verde)
                ws.set_column(j+2,j+2,14)

                j += 3

            fin = j

            ws.write(0,fin,'APROBACIÓN N1',morado)
            ws.set_column(0,fin,20)

            ws.write(0,fin+1,'APROBACIÓN N2',morado)
            ws.set_column(0,fin+1,20)

            ws.write(0,fin+2,'APROBACIÓN N3',morado)
            ws.set_column(0,fin+2,20)

            ws.write(0,fin+3,'APROBACIÓN N4',morado)
            ws.set_column(0,fin+3,20)


            ws.write(0,fin+4,'APROBACIÓN N1 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+4,20)

            ws.write(0,fin+5,'APROBACIÓN N2 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+5,20)

            ws.write(0,fin+6,'APROBACIÓN N3 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+6,20)

            ws.write(0,fin+7,'APROBACIÓN N4 (CON ACTAS)',agua_marina)
            ws.set_column(0,fin+7,20)





        ws.write('A1','REGIÓN',gris)
        ws.set_column('A:A',11)
        ws.set_row(0,90)

        ws.write('B1','NOMBRE DEL FORMADOR',gris)
        ws.set_column('B:B',36)

        ws.write('C1','NUMERO DE CEDULA DEL FORMADOR',gris)
        ws.set_column('C:C',13)

        ws.write('D1','APELLIDOS DEL BENEFICIARIO',verde)
        ws.set_column('D:D',25)

        ws.write('E1','NOMBRES DEL BENEFICIARIO',verde)
        ws.set_column('E:E',25)

        ws.write('F1','NUMERO DE CEDULA DEL BENEFICIARIO',verde)
        ws.set_column('F:F',14)

        i = 2

        for beneficiario in Beneficiario.objects.filter(diplomado__id = id_diplomado).order_by('formador'):
            ws.write('A'+str(i), beneficiario.region.nombre.upper(),text)

            ws.write('B'+str(i), beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else beneficiario.departamento_text,text)

            ws.write('C'+str(i), beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else beneficiario.municipio_text,text)

            ws.write('D'+str(i), beneficiario.formador.get_full_name(),text)

            ws.write('E'+str(i), beneficiario.formador.cedula,number)

            ws.write('F'+str(i), beneficiario.apellidos,text)

            ws.write('G'+str(i), beneficiario.nombres,text)

            ws.write('H'+str(i), beneficiario.cedula,number)


            for producto in dict_productos:
                entregable = Entregable.objects.get(id = producto['id'])

                evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable)
                evidencias_validado = Evidencia.objects.filter(beneficiarios_validados = beneficiario,entregable = entregable)


                if evidencias_validado.count() > 0:
                    ws.write( producto['letter'] + str(i) , 'SIC-' + str(evidencias_validado[0].id), validado )
                    ws.write( producto['red'] + str(i) , 'RED-' + str(evidencias_validado[0].red_id), text )
                    ws.write( producto['aprobacion'] + str(i) , 'SI', text )


                elif evidencias_validado.count() == 0 and evidencias_cargado.count() > 0:
                    ultima_evidencia_cargada = evidencias_cargado[len(evidencias_cargado)-1]

                    try:
                        red = Red.objects.get(id = ultima_evidencia_cargada.red_id)
                    except:
                        # si ningun red contiene la evidencia
                        ws.write( producto['letter'] + str(i), 'SIC-' + str(ultima_evidencia_cargada.id), cargado )
                        if ultima_evidencia_cargada.red_id != None:
                            ws.write( producto['red'] + str(i) , 'RED-' + str(ultima_evidencia_cargada.red_id), text )

                    else:
                        # si hay un red que contiene la evidencia
                        if red.retroalimentacion:
                            # si el red fue retroalimentado

                            evidencias_rechazado = Evidencia.objects.filter(beneficiarios_rechazados__beneficiario_rechazo = beneficiario,
                                                                            beneficiarios_rechazados__red_id = red.id,
                                                                            beneficiarios_rechazados__evidencia_id = ultima_evidencia_cargada.id)

                            if evidencias_rechazado > 0:
                                ws.write( producto['letter'] + str(i) , 'SIC-' + str(ultima_evidencia_cargada.id), rechazado )
                                if ultima_evidencia_cargada.red_id != None:
                                    ws.write( producto['red'] + str(i) , 'RED-' + str(ultima_evidencia_cargada.red_id), text )
                            else:
                                pass


                        else:
                            #si el red no ha sido retroalimentado
                            ws.write( producto['letter'] + str(i) , 'SIC-' + str(ultima_evidencia_cargada.id), enviado )
                            if ultima_evidencia_cargada.red_id != None:
                                    ws.write( producto['red'] + str(i) , 'RED-' + str(ultima_evidencia_cargada.red_id), text )


                else:
                    pass

                if id_diplomado == 1:
                    ws.write_formula(i-1,fin,'=IF(IF(N'+ str(i) +'="SI",1,0)+IF(Q'+ str(i) +'="SI",1,0)+IF(T'+ str(i) +'="SI",1,0)+IF(W' + str(i) + '="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+1,'=IF(IF(Z'+ str(i) +'="SI",1,0)+IF(AC'+ str(i) +'="SI",1,0)+IF(AF'+ str(i) +'="SI",1,0)+IF(AI'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+2,'=IF(IF(AL'+ str(i) +'="SI",1,0)+IF(AO'+ str(i) +'="SI",1,0)+IF(AR'+ str(i) +'="SI",1,0)+IF(AU'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+3,'=IF(IF(AX'+ str(i) +'="SI",1,0)+IF(BA'+ str(i) +'="SI",1,0)>=1,"SI","NO")',text)


                    ws.write_formula(i-1,fin+4,'=IF($K'+ str(i) +'="SI",BB'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+5,'=IF($K'+ str(i) +'="SI",BC'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+6,'=IF($K'+ str(i) +'="SI",BD'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+7,'=IF($K'+ str(i) +'="SI",BE'+ str(i) +',"NO ACTA")',text)


                if id_diplomado == 2:
                    ws.write_formula(i-1,fin,'=IF(IF(N'+ str(i) +'="SI",1,0)+IF(Q'+ str(i) +'="SI",1,0)+IF(T'+ str(i) +'="SI",1,0)+IF(W' + str(i) + '="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+1,'=IF(IF(Z'+ str(i) +'="SI",1,0)+IF(AC'+ str(i) +'="SI",1,0)+IF(AF'+ str(i) +'="SI",1,0)+IF(AI'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+2,'=IF(IF(AL'+ str(i) +'="SI",1,0)+IF(AO'+ str(i) +'="SI",1,0)+IF(AR'+ str(i) +'="SI",1,0)+IF(AU'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+3,'=IF(IF(AX'+ str(i) +'="SI",1,0)+IF(BA'+ str(i) +'="SI",1,0)>=1,"SI","NO")',text)


                    ws.write_formula(i-1,fin+4,'=IF($K'+ str(i) +'="SI",BB'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+5,'=IF($K'+ str(i) +'="SI",BC'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+6,'=IF($K'+ str(i) +'="SI",BD'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+7,'=IF($K'+ str(i) +'="SI",BE'+ str(i) +',"NO ACTA")',text)

                if id_diplomado == 3:
                    ws.write_formula(i-1,fin,'=IF(IF(N'+ str(i) +'="SI",1,0)+IF(Q'+ str(i) +'="SI",1,0)+IF(T'+ str(i) +'="SI",1,0)+IF(W'+ str(i) +'="SI",1,0)+IF(Z'+ str(i) +'="SI",1,0)+IF(AC'+ str(i) +'="SI",1,0)>=4,"SI","NO")',text)
                    ws.write_formula(i-1,fin+1,'=IF(IF(AF'+ str(i) +'="SI",1,0)+IF(AI'+ str(i) +'="SI",1,0)+IF(AL'+ str(i) +'="SI",1,0)>=2,"SI","NO")',text)
                    ws.write_formula(i-1,fin+2,'=IF(IF(AO'+ str(i) +'="IF",1,0)+IF(AR'+ str(i) +'="SI",1,0)+IF(AU'+ str(i) +'="SI",1,0)+IF(AX'+ str(i) +'="SI",1,0)>=3,"SI","NO")',text)
                    ws.write_formula(i-1,fin+3,'=IF(IF(BA'+ str(i) +'="SI",1,0)+IF(BD'+ str(i) +'="SI",1,0)>=1,"SI","NO")',text)


                    ws.write_formula(i-1,fin+4,'=IF($K'+ str(i) +'="SI",BE'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+5,'=IF($K'+ str(i) +'="SI",BF'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+6,'=IF($K'+ str(i) +'="SI",BG'+ str(i) +',"NO ACTA")',text)
                    ws.write_formula(i-1,fin+7,'=IF($K'+ str(i) +'="SI",BH'+ str(i) +',"NO ACTA")',text)

                if id_diplomado == 4:
                    ws.write_formula(i-1,fin,'=IF(IF(K'+ str(i) +'="SI",1,0)+IF(N'+ str(i) +'="SI",1,0)=2,"SI","NO")',text)

            i += 1

    wb.close()

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def reporte_sed_bogota(email):
    usuario = User.objects.get(email=email)
    nombre = 'Reporte secretario bogota'

    proceso = "REV-INF05"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []



    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet('SECRETARIA BOGOTA')

    text = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False})

    number = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'right', 'valign':'vcenter', 'text_wrap':False,'num_format':'0'})

    validado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#00B050'})

    enviado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#FFC000'})

    cargado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False})

    rechazado = wb.add_format({'font_name':'Calibri', 'font_size':12 ,'align':'left', 'valign':'vcenter', 'text_wrap':False, 'pattern':1, 'bg_color':'#FF0000'})

    gris = wb.add_format({'font_name':'Arial Narrow','border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#808080', 'font_color':'#FFFFFF'})

    verde = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#008000', 'font_color':'#FFFFFF'})

    azul = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#000080', 'font_color':'#FFFFFF'})

    morado = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#7030A0', 'font_color':'#FFFFFF'})

    agua_marina = wb.add_format({'font_name':'Arial Narrow', 'border':1, 'bold':1, 'font_size':10 ,'align':'center', 'valign':'vcenter', 'text_wrap':True, 'pattern':1, 'bg_color':'#00B0F0', 'font_color':'#FFFFFF'})

    ws.write('A1','CEDULA',azul)
    ws.set_column('A:A',11)

    ws.write('B1','PRIMER APELLIDO',azul)
    ws.set_column('B:B',20)

    ws.write('C1','SEGUNDO APELLIDO',azul)
    ws.set_column('C:C',20)

    ws.write('D1','PRIMER NOMBRE',azul)
    ws.set_column('D:D',20)

    ws.write('E1','SEGUNDO NOMBRE',azul)
    ws.set_column('E:E',20)

    ws.write('F1','CARGO',azul)
    ws.set_column('F:F',18)

    ws.write('G1','CORREO',azul)
    ws.set_column('G:G',20)

    ws.write('H1','TELEFONO FIJO',azul)
    ws.set_column('H:H',20)

    ws.write('I1','TELEFONO CELULAR',azul)
    ws.set_column('I:I',20)

    ws.write('J1','DEPARTAMENTO',azul)
    ws.set_column('J:J',20)

    ws.write('K1','MUNICIPIO',azul)
    ws.set_column('K:K',20)

    ws.write('L1','RADICADO',azul)
    ws.set_column('L:L',20)

    ws.write('M1','SEDE',azul)
    ws.set_column('M:M',20)

    ws.write('N1','CODIGO DANE',azul)
    ws.set_column('N:N',20)

    ws.write('O1','FECHA',azul)
    ws.set_column('O:O',20)

    ws.write('P1','AREA',azul)
    ws.set_column('P:P',20)

    ws.write('Q1','LOCALIDAD',azul)
    ws.set_column('Q:Q',20)

    i = 2

    for beneficiario in DocentesPreinscritos.objects.filter(docentes_innovadores = True):

        ws.write('A'+str(i),beneficiario.cedula,text)
        ws.write('B'+str(i),beneficiario.primer_apellido,text)
        ws.write('C'+str(i),beneficiario.segundo_apellido,text)
        ws.write('D'+str(i),beneficiario.primer_nombre,text)
        ws.write('E'+str(i),beneficiario.segundo_nombre,text)
        ws.write('F'+str(i),beneficiario.cargo,text)
        ws.write('G'+str(i),beneficiario.correo,text)
        ws.write('H'+str(i),beneficiario.telefono_fijo,text)
        ws.write('I'+str(i),beneficiario.telefono_celular,text)
        ws.write('J'+str(i),beneficiario.departamento.nombre,text)
        ws.write('K'+str(i),beneficiario.municipio.nombre,text)
        ws.write('L'+str(i),beneficiario.radicado.numero,text)
        ws.write('M'+str(i),beneficiario.radicado.nombre_sede,text)
        ws.write('N'+str(i),beneficiario.radicado.dane_sede,text)
        ws.write('O'+str(i),beneficiario.fecha,text)
        ws.write('P'+str(i),beneficiario.area,text)
        ws.write('Q'+str(i),beneficiario.localidad,text)


        i += 1




    wb.close()

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def reporte_legalizacion_contrato_formadores(email):
    usuario = User.objects.get(email=email)
    nombre = "Reporte de legalización de contratos"
    proceso = "RH-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Código','Formador','Cedula','Región','Agrupación de soportes']

    formatos = ['General','General','General','General','General','General']


    ancho_columnas =  [30,20,30,30,30,30]

    contenidos = []

    max = 0

    for contrato in Contrato.objects.all():

        soportes = SoporteFormadores.objects.filter(contrato = contrato)

        restante = []

        for soporte in soportes:
            restante.append(soporte.tipo.nombre)
            restante.append('https://sican.asoandes.org'+soporte.get_archivo_url())

        contenido = ['CONT-FOR-'+unicode(contrato.id),
                     contrato.nombre,
                     contrato.formador.get_full_name(),
                     contrato.formador.cedula,
                     contrato.formador.get_region_string(),
                     contrato.soportes_requeridos.nombre] + restante

        if len(contenido) > max:
            max = len(contenido)

        if len(titulos) < max:
            while(len(titulos)<max):
                titulos.append('')

        if len(formatos) < max:
            while(len(formatos)<max):
                formatos.append('General')

        if len(ancho_columnas) < max:
            while(len(ancho_columnas)<max):
                ancho_columnas.append(30)

        contenidos.append(contenido)

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"
