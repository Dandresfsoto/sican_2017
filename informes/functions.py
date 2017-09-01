from __future__ import unicode_literals
import openpyxl
from StringIO import StringIO
from sican.settings import base as settings
from openpyxl.drawing.image import Image
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from dateutil.tz import tzutc, tzlocal
from formacion.models import EntradaCronograma

def construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso):
    if len(ancho_columnas) != len(formatos) != len(titulos) != len(contenidos[0]):
        raise Exception('El arreglo de filas y columnas tienen distinta longitud')
    else:
        output = StringIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Formato_informe.xlsx')
        ws = wb.get_sheet_by_name('Hoja1')
        logo_andes = Image(settings.STATICFILES_DIRS[0]+'/img/andes_logo_informe.png',size=(145,145))
        logo_andes.drawing.top = 17
        logo_andes.drawing.left = 8
        ws.add_image(logo_andes)
        ws.cell('B1').value = "   Nombre: " + nombre
        ws.cell('B3').value = "   Fecha: " + fecha.astimezone(tzlocal()).strftime("%d/%m/%Y - %X")
        ws.cell('B5').value = "   Usuario: " + usuario.email
        ws.cell('B7').value = "   Proceso: " + proceso


        row_num = 9
        for col_num in xrange(len(titulos)):
            ws.cell(row=row_num, column=col_num+1).value = titulos[col_num]
            if col_num != 0:
                ws.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = ancho_columnas[col_num]

            ws.cell(row=row_num,column=col_num+1).style = Style(font=Font(name='Arial',
                                                                          size=11,
                                                                          bold=True,
                                                                          color='FFFFFFFF'
                                                                        ),
                                                                fill=PatternFill(
                                                                    fill_type='solid',
                                                                    start_color='CB2E0F',
                                                                    end_color='FF000000'
                                                                ),
                                                                alignment=Alignment(
                                                                    horizontal='center',
                                                                    vertical='center',
                                                                    wrap_text=True
                                                                ),
                                                                number_format='General'
                                                                )

        for contenido in contenidos:
            row_num += 1
            for col_num in xrange(len(contenido)):
                if contenido[col_num] == True:
                    ws.cell(row=row_num,column=col_num+1).value = "SI"
                if contenido[col_num] == False:
                    ws.cell(row=row_num,column=col_num+1).value = "NO"
                if contenido[col_num] == None:
                    ws.cell(row=row_num,column=col_num+1).value = ""
                else:
                    ws.cell(row=row_num,column=col_num+1).value = contenido[col_num]

                ws.cell(row=row_num,column=col_num+1).style = Style(font=Font(name='Arial',size=10),
                                                                        alignment=Alignment(
                                                                            horizontal='center',
                                                                            vertical='center',
                                                                            wrap_text=True
                                                                        ),
                                                                    number_format=formatos[col_num]
                                                                    )

        wb.save(output)
        return output


def cronograma_interventoria(innovatics,tecnotics,directics,escuelatics,rango):

    output = StringIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Ruteo y Cronograma.xlsx')
    ws_innovatic = wb.get_sheet_by_name('InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('TecnoTIC')
    ws_directic = wb.get_sheet_by_name('DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('ESCUELA TIC FAMILIA')

    logo_andes = Image(settings.STATICFILES_DIRS[0]+'/img/silva.jpg',size=(110,110))
    logo_andes.drawing.top = 2
    logo_andes.drawing.left = 2
    ws_innovatic.add_image(logo_andes)
    ws_tecnotic.add_image(logo_andes)
    ws_directic.add_image(logo_andes)
    ws_escuelatic.add_image(logo_andes)

    ws_innovatic.cell('F4').value = "ANDES"
    ws_tecnotic.cell('F4').value = "ANDES"
    ws_directic.cell('F4').value = "ANDES"
    ws_escuelatic.cell('F4').value = "ANDES"

    ws_innovatic.cell('M5').value = rango
    ws_tecnotic.cell('M5').value = rango
    ws_directic.cell('M5').value = rango
    ws_escuelatic.cell('M5').value = rango

    #------------------------------------------tecnotic-------------------------------------------------------------

    row_num = 10
    for innovatic in EntradaCronograma.objects.filter(id__in=innovatics).order_by('formador__region__numero'):
        for nivel in innovatic.nivel.all():
            row_num += 1

            actividades = ''
            for actividad in innovatic.actividades_entrada.filter(sesion__nivel=nivel):
                actividades += str(actividad.numero) + ','

            actividades = actividades[:-1]

            ws_innovatic.cell(row=row_num,column=1).value = innovatic.formador.get_interventoria_region()
            ws_innovatic.cell(row=row_num,column=2).value = innovatic.departamento.nombre.upper()
            ws_innovatic.cell(row=row_num,column=3).value = innovatic.municipio.nombre.upper()
            ws_innovatic.cell(row=row_num,column=4).value = innovatic.secretaria.nombre.upper()
            ws_innovatic.cell(row=row_num,column=5).value = innovatic.formador.cedula
            ws_innovatic.cell(row=row_num,column=6).value = innovatic.formador.get_full_name().upper()
            ws_innovatic.cell(row=row_num,column=7).value = innovatic.formador.celular_personal
            ws_innovatic.cell(row=row_num,column=9).value = innovatic.formador.correo_personal
            ws_innovatic.cell(row=row_num,column=10).value = innovatic.formador.codigo_ruta.upper() + '-' + innovatic.grupo.nombre.upper()
            ws_innovatic.cell(row=row_num,column=11).value = innovatic.numero_sedes
            ws_innovatic.cell(row=row_num,column=12).value = nivel.nombre.replace(' ','') + 'I'
            ws_innovatic.cell(row=row_num,column=13).value = actividades
            ws_innovatic.cell(row=row_num,column=14).value = innovatic.beneficiados
            ws_innovatic.cell(row=row_num,column=15).value = innovatic.fecha
            ws_innovatic.cell(row=row_num,column=16).value = innovatic.institucion.upper()
            ws_innovatic.cell(row=row_num,column=17).value = innovatic.direccion.upper()
            ws_innovatic.cell(row=row_num,column=18).value = innovatic.telefono
            ws_innovatic.cell(row=row_num,column=19).value = innovatic.hora_inicio
            ws_innovatic.cell(row=row_num,column=20).value = innovatic.hora_finalizacion
            ws_innovatic.cell(row=row_num,column=21).value = innovatic.ubicacion.upper()
            ws_innovatic.cell(row=row_num,column=22).value = innovatic.observaciones.upper()

    row_num = 10
    for tecnotic in EntradaCronograma.objects.filter(id__in=tecnotics).order_by('formador__region__numero'):
        for nivel in tecnotic.nivel.all():
            row_num += 1

            actividades = ''
            for actividad in tecnotic.actividades_entrada.filter(sesion__nivel=nivel):
                actividades += str(actividad.numero) + ','

            actividades = actividades[:-1]

            ws_tecnotic.cell(row=row_num,column=1).value = tecnotic.formador.get_interventoria_region()
            ws_tecnotic.cell(row=row_num,column=2).value = tecnotic.departamento.nombre.upper()
            ws_tecnotic.cell(row=row_num,column=3).value = tecnotic.municipio.nombre.upper()
            ws_tecnotic.cell(row=row_num,column=4).value = tecnotic.secretaria.nombre.upper()
            ws_tecnotic.cell(row=row_num,column=5).value = tecnotic.formador.cedula
            ws_tecnotic.cell(row=row_num,column=6).value = tecnotic.formador.get_full_name().upper()
            ws_tecnotic.cell(row=row_num,column=7).value = tecnotic.formador.celular_personal
            ws_tecnotic.cell(row=row_num,column=9).value = tecnotic.formador.correo_personal
            ws_tecnotic.cell(row=row_num,column=10).value = tecnotic.formador.codigo_ruta.upper() + '-' + tecnotic.grupo.nombre.upper()
            ws_tecnotic.cell(row=row_num,column=11).value = tecnotic.numero_sedes
            ws_tecnotic.cell(row=row_num,column=12).value = nivel.nombre.replace(' ','') + 'T'
            ws_tecnotic.cell(row=row_num,column=13).value = actividades
            ws_tecnotic.cell(row=row_num,column=14).value = tecnotic.beneficiados
            ws_tecnotic.cell(row=row_num,column=15).value = tecnotic.fecha
            ws_tecnotic.cell(row=row_num,column=16).value = tecnotic.institucion.upper()
            ws_tecnotic.cell(row=row_num,column=17).value = tecnotic.direccion.upper()
            ws_tecnotic.cell(row=row_num,column=18).value = tecnotic.telefono
            ws_tecnotic.cell(row=row_num,column=19).value = tecnotic.hora_inicio
            ws_tecnotic.cell(row=row_num,column=20).value = tecnotic.hora_finalizacion
            ws_tecnotic.cell(row=row_num,column=21).value = tecnotic.ubicacion.upper()
            ws_tecnotic.cell(row=row_num,column=22).value = tecnotic.observaciones.upper()

    row_num = 10
    for directic in EntradaCronograma.objects.filter(id__in=directics).order_by('formador__region__numero'):
        for nivel in directic.nivel.all():
            row_num += 1

            actividades = ''
            for actividad in directic.actividades_entrada.filter(sesion__nivel=nivel):
                actividades += str(actividad.numero) + ','

            actividades = actividades[:-1]

            ws_directic.cell(row=row_num,column=1).value = directic.formador.get_interventoria_region()
            ws_directic.cell(row=row_num,column=2).value = directic.departamento.nombre.upper()
            ws_directic.cell(row=row_num,column=3).value = directic.municipio.nombre.upper()
            ws_directic.cell(row=row_num,column=4).value = directic.secretaria.nombre.upper()
            ws_directic.cell(row=row_num,column=5).value = directic.formador.cedula
            ws_directic.cell(row=row_num,column=6).value = directic.formador.get_full_name().upper()
            ws_directic.cell(row=row_num,column=7).value = directic.formador.celular_personal
            ws_directic.cell(row=row_num,column=9).value = directic.formador.correo_personal
            ws_directic.cell(row=row_num,column=10).value = directic.formador.codigo_ruta.upper() + '-' + directic.grupo.nombre.upper()
            ws_directic.cell(row=row_num,column=11).value = directic.numero_sedes
            ws_directic.cell(row=row_num,column=12).value = nivel.nombre.replace(' ','') + 'D'
            ws_directic.cell(row=row_num,column=13).value = actividades
            ws_directic.cell(row=row_num,column=14).value = directic.beneficiados
            ws_directic.cell(row=row_num,column=15).value = directic.fecha
            ws_directic.cell(row=row_num,column=16).value = directic.institucion.upper()
            ws_directic.cell(row=row_num,column=17).value = directic.direccion.upper()
            ws_directic.cell(row=row_num,column=18).value = directic.telefono
            ws_directic.cell(row=row_num,column=19).value = directic.hora_inicio
            ws_directic.cell(row=row_num,column=20).value = directic.hora_finalizacion
            ws_directic.cell(row=row_num,column=21).value = directic.ubicacion.upper()
            ws_directic.cell(row=row_num,column=22).value = directic.observaciones.upper()

    row_num = 10
    for escuelatic in EntradaCronograma.objects.filter(id__in=escuelatics).order_by('formador__region__numero'):
        for nivel in escuelatic.nivel.all():
            row_num += 1

            actividades = ''
            for actividad in escuelatic.actividades_entrada.filter(sesion__nivel=nivel):
                actividades += str(actividad.numero) + ','

            actividades = actividades[:-1]

            ws_escuelatic.cell(row=row_num,column=1).value = escuelatic.formador.get_interventoria_region()
            ws_escuelatic.cell(row=row_num,column=2).value = escuelatic.departamento.nombre.upper()
            ws_escuelatic.cell(row=row_num,column=3).value = escuelatic.municipio.nombre.upper()
            ws_escuelatic.cell(row=row_num,column=4).value = escuelatic.secretaria.nombre.upper()
            ws_escuelatic.cell(row=row_num,column=5).value = escuelatic.formador.cedula
            ws_escuelatic.cell(row=row_num,column=6).value = escuelatic.formador.get_full_name().upper()
            ws_escuelatic.cell(row=row_num,column=7).value = escuelatic.formador.celular_personal
            ws_escuelatic.cell(row=row_num,column=9).value = escuelatic.formador.correo_personal
            ws_escuelatic.cell(row=row_num,column=10).value = escuelatic.formador.codigo_ruta.upper() + '-' + escuelatic.grupo.nombre.upper()
            ws_escuelatic.cell(row=row_num,column=11).value = escuelatic.numero_sedes
            ws_escuelatic.cell(row=row_num,column=12).value = nivel.nombre.replace(' ','') + 'E'
            ws_escuelatic.cell(row=row_num,column=13).value = actividades
            ws_escuelatic.cell(row=row_num,column=14).value = escuelatic.beneficiados
            ws_escuelatic.cell(row=row_num,column=15).value = escuelatic.fecha
            ws_escuelatic.cell(row=row_num,column=16).value = escuelatic.institucion.upper()
            ws_escuelatic.cell(row=row_num,column=17).value = escuelatic.direccion.upper()
            ws_escuelatic.cell(row=row_num,column=18).value = escuelatic.telefono
            ws_escuelatic.cell(row=row_num,column=19).value = escuelatic.hora_inicio
            ws_escuelatic.cell(row=row_num,column=20).value = escuelatic.hora_finalizacion
            ws_escuelatic.cell(row=row_num,column=21).value = escuelatic.ubicacion.upper()
            ws_escuelatic.cell(row=row_num,column=22).value = escuelatic.observaciones.upper()

    wb.save(output)
    return output